"""
EconAI FastAPI backend.

Serves LabelMe JSON + images from any folder on disk.
Accepts corrections and writes them back to the JSON.

Run:
  python -m uvicorn app.server:app --reload --port 8000
or via econai.py (coming soon).
"""

from __future__ import annotations

import json
import os
import posixpath
import re
import asyncio
import tempfile
import threading
import time
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List

from app.pipeline import (
    list_projects,
    load_config,
    load_pipeline,
    save_pipeline,
    stages_for,
    project_dir,
    advance_stage,
    set_stage,
)

from contextlib import asynccontextmanager

# ---------------------------------------------------------------------------
# Atomic JSON writes — concurrent handlers (SSE OCR/LLM runs, row edits,
# shape patches) may write the same page JSON; a plain write_text can leave
# an interleaved/corrupt file.  Serialize writes and use temp-file + replace
# so readers always see a complete document.
# ---------------------------------------------------------------------------
_JSON_WRITE_LOCK = threading.Lock()


def _write_json(path, obj):
    path = Path(path)
    text = json.dumps(obj, indent=2, ensure_ascii=False)
    with _JSON_WRITE_LOCK:
        fd, tmp = tempfile.mkstemp(dir=str(path.parent),
                                   prefix=path.name + ".", suffix=".tmp")
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as f:
                f.write(text)
            # Windows: os.replace fails while ANYTHING holds the target open —
            # Dropbox sync and antivirus scanners routinely do for a second or
            # two.  Retry with backoff for up to ~10 s before giving up.
            delay = 0.05
            for attempt in range(12):
                try:
                    os.replace(tmp, str(path))
                    return
                except PermissionError:
                    if attempt == 11:
                        break
                    time.sleep(delay)
                    delay = min(delay * 1.7, 2.0)
            # Last resort: direct write.  Not crash-atomic, but we hold the
            # global write lock so no other server thread interleaves, and it
            # beats losing the data outright.
            print(f"[write_json] os.replace blocked for {path.name} — "
                  f"falling back to direct write", flush=True)
            path.write_text(text, encoding="utf-8")
            try:
                os.unlink(tmp)
            except OSError:
                pass
        except BaseException:
            try:
                os.unlink(tmp)
            except OSError:
                pass
            raise


@asynccontextmanager
async def lifespan(app):
    # Warm up EasyOCR on startup so the first user request is fast
    import asyncio, concurrent.futures
    def _warmup():
        try:
            _get_easyocr_reader(["en", "hu"])
            print("EasyOCR warmed up (GPU)", flush=True)
        except Exception as e:
            print(f"EasyOCR warmup failed: {e}", flush=True)
    asyncio.get_event_loop().run_in_executor(
        concurrent.futures.ThreadPoolExecutor(max_workers=1), _warmup
    )
    yield

app = FastAPI(title="Dedust", version="0.1.0", lifespan=lifespan)

from app.validator import router as _validator_router
app.include_router(_validator_router)

import os as _os

# Allow the browser (same host, any port) to call the API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Remote-access guard (Phase A). INERT unless the ECONAI_TOKEN environment
# variable is set — with no token, behavior is identical to before (and the
# default bind is 127.0.0.1, so nothing is reachable remotely anyway).
# With a token set:
#   * requests from localhost stay unrestricted (local workflow unchanged),
#   * remote requests need the token (Bearer header, X-Econai-Token, cookie
#     set by /static/login.html, or ?token=),
#   * remote requests may only touch folders under <repo>/projects/ (the
#     `folder` param otherwise accepts arbitrary filesystem paths).
# ---------------------------------------------------------------------------
import contextvars
import secrets as _secrets

_REMOTE_REQ = contextvars.ContextVar("econai_remote_request", default=False)
PROJECTS_ROOT = (Path(__file__).parent.parent / "projects").resolve()

_AUTH_EXEMPT = ("/static/login.html", "/api/login", "/static/manifest.json",
                "/static/sw.js", "/static/icon-192.png", "/static/icon-512.png")


def _req_token(request) -> str:
    auth = request.headers.get("authorization", "")
    if auth.lower().startswith("bearer "):
        return auth[7:].strip()
    return (request.headers.get("x-econai-token")
            or request.cookies.get("econai_token")
            or request.query_params.get("token") or "")


@app.middleware("http")
async def _remote_guard(request, call_next):
    token = os.environ.get("ECONAI_TOKEN")
    if not token:
        return await call_next(request)          # guard disabled — legacy behavior
    host = request.client.host if request.client else ""
    if host in ("127.0.0.1", "::1"):
        return await call_next(request)          # local stays unrestricted
    path = request.url.path
    if path in _AUTH_EXEMPT:
        return await call_next(request)
    if _secrets.compare_digest(_req_token(request), token):
        tok = _REMOTE_REQ.set(True)              # authorized remote → caged folders
        try:
            return await call_next(request)
        finally:
            _REMOTE_REQ.reset(tok)
    from fastapi.responses import JSONResponse, RedirectResponse
    wants_html = "text/html" in request.headers.get("accept", "")
    if request.method == "GET" and wants_html:
        return RedirectResponse("/static/login.html")
    return JSONResponse({"detail": "Missing or invalid token"}, status_code=401)


class LoginBody(BaseModel):
    token: str


@app.post("/api/login")
def api_login(body: LoginBody):
    """Exchange the shared token for a long-lived cookie (used by the login
    page so remote browsers authenticate once per device)."""
    real = os.environ.get("ECONAI_TOKEN")
    if not real:
        raise HTTPException(status_code=400, detail="Remote access is not enabled on this server")
    if not _secrets.compare_digest(body.token, real):
        raise HTTPException(status_code=401, detail="Wrong token")
    from fastapi.responses import JSONResponse
    resp = JSONResponse({"ok": True})
    resp.set_cookie("econai_token", real, max_age=60 * 60 * 24 * 30,
                    httponly=True, samesite="lax")
    return resp


# ---------------------------------------------------------------------------
# Per-page mutation serialization — handlers do read-modify-write on the page
# JSON, so two concurrent mutations of the same page (second browser tab,
# batch op overlapping a click) could each read the old file and the later
# write would silently erase the earlier change.  _write_json makes each
# write crash-atomic; this middleware makes the whole request atomic by
# queueing mutating requests per (folder, stem).  SSE/streaming runs release
# the lock when the stream starts — long OCR/LLM runs stay unserialized by
# design (they write once at the end and the client serializes its own ops).
# ---------------------------------------------------------------------------
_PAGE_MUT_LOCKS: dict = {}
_PAGE_MUT_GUARD = threading.Lock()

def _page_mut_lock(folder: str, stem: str):
    try:
        fkey = str(Path(folder).resolve()).lower()
    except OSError:
        fkey = folder.lower()
    key = (fkey, stem.lower())
    with _PAGE_MUT_GUARD:
        lock = _PAGE_MUT_LOCKS.get(key)
        if lock is None:
            lock = _PAGE_MUT_LOCKS[key] = asyncio.Lock()
    return lock

@app.middleware("http")
async def _serialize_page_mutations(request, call_next):
    if request.method in ("PATCH", "POST", "PUT", "DELETE") \
            and request.url.path.startswith("/api/page") \
            and not request.url.path.startswith("/api/page/shape/llm"):
        # LLM endpoints are exempt: they run in parallel by design and use
        # merge-safe per-shape writes (_merge_shape_fields) instead of
        # whole-document read-modify-write.
        folder = request.query_params.get("folder")
        stem   = request.query_params.get("stem")
        if folder and stem:
            async with _page_mut_lock(folder, stem):
                return await call_next(request)
    return await call_next(request)

# Serve the frontend from app/static/
STATIC_DIR = Path(__file__).parent / "static"
STATIC_DIR.mkdir(exist_ok=True)

class _NoCacheStaticFiles(StaticFiles):
    """Static files with Cache-Control: no-cache.

    The browser may keep a copy but must revalidate on every request
    (a cheap 304 when unchanged), so a plain reload always picks up
    edits to index.html & co — no more stale-build hard-reload ritual."""
    def file_response(self, *args, **kwargs):
        resp = super().file_response(*args, **kwargs)
        resp.headers["Cache-Control"] = "no-cache"
        return resp

app.mount("/static", _NoCacheStaticFiles(directory=str(STATIC_DIR)), name="static")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _resolve_folder(folder: str) -> Path:
    """Resolve a folder path that may be absolute or relative to cwd.
    Authorized REMOTE requests (see _remote_guard) are caged to the repo's
    projects/ root — locally the historic any-path behavior is unchanged."""
    p = Path(folder)
    if not p.is_absolute():
        p = Path.cwd() / p
    if _REMOTE_REQ.get():
        try:
            rp = p.resolve()
            rp.relative_to(PROJECTS_ROOT)
        except (ValueError, OSError):
            raise HTTPException(status_code=403,
                detail="Remote sessions may only access folders under projects/")
        p = rp
    if not p.exists():
        raise HTTPException(status_code=404, detail=f"Folder not found: {p}")
    if not p.is_dir():
        raise HTTPException(status_code=400, detail=f"Not a directory: {p}")
    return p


def _find_image(folder: Path, stem: str) -> Optional[Path]:
    for ext in (".jpg", ".jpeg", ".png", ".tif", ".tiff"):
        candidate = folder / (stem + ext)
        if candidate.exists():
            return candidate
    return None


def _resolve_pdf_source(ann_dir: Path, pdf_source: str) -> Optional[Path]:
    """Locate the source PDF for a page on THIS machine.

    `pdf_source` in older JSONs is an absolute path baked in at import time
    (machine-specific). The PDF itself is copied into <project>/sources/ at
    import, so we resolve robustly: relative-to-project first, then a basename
    fallback in sources/ — which repairs the legacy absolute-path JSONs too."""
    if not pdf_source:
        return None
    proj = ann_dir.parent                 # annotations/ -> project root
    base = Path(pdf_source).name
    cands = []
    p = Path(pdf_source)
    if p.is_absolute():
        cands.append(p)                   # same machine / unchanged path
    else:
        cands += [proj / pdf_source, ann_dir / pdf_source]
    cands += [proj / "sources" / base, ann_dir / "sources" / base,
              proj / base, ann_dir / base]
    for c in cands:
        try:
            if c.exists():
                return c
        except OSError:
            pass
    return None


def _page_sort_key(name: str) -> tuple:
    """Sort page_1 < page_2 < ... < page_10 (natural sort)."""
    parts = re.split(r"(\d+)", name)
    return tuple(int(p) if p.isdigit() else p.lower() for p in parts)


# ---------------------------------------------------------------------------
# Routes — page listing
# ---------------------------------------------------------------------------

@app.get("/api/pages")
def list_pages(folder: str = Query(..., description="Absolute or relative path to the annotations folder")):
    """Return a sorted list of page stems that have both a JSON and an image."""
    d = _resolve_folder(folder)
    pages = []
    for jf in sorted(d.glob("*.json"), key=lambda p: _page_sort_key(p.stem)):
        img = _find_image(d, jf.stem)
        if img:
            pages.append({
                "stem": jf.stem,
                "json": jf.name,
                "image": img.name,
            })
    return {"folder": str(d), "pages": pages}


@app.get("/api/json-stems")
def list_json_stems(folder: str = Query(...)):
    """Return every JSON stem in the folder sorted naturally — no image required."""
    d = _resolve_folder(folder)
    stems = [jf.stem for jf in sorted(d.glob("*.json"),
                                       key=lambda p: _page_sort_key(p.stem))]
    return {"stems": stems}


# ---------------------------------------------------------------------------
# Routes — serving files
# ---------------------------------------------------------------------------

@app.get("/api/image")
def get_image(folder: str = Query(...), stem: str = Query(...)):
    """Serve the image file for a given page stem."""
    d = _resolve_folder(folder)
    img = _find_image(d, stem)
    if img is None:
        raise HTTPException(status_code=404, detail=f"No image found for '{stem}'")
    return FileResponse(str(img), media_type="image/jpeg",
                        headers={"Cache-Control": "no-store"})


@app.get("/api/page")
def get_page(folder: str = Query(...), stem: str = Query(...)):
    """Return the LabelMe JSON for a page, with shape indices added."""
    d = _resolve_folder(folder)
    jf = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")

    data = json.loads(jf.read_text(encoding="utf-8"))

    # Inject a stable index into each shape so the frontend can reference them
    for i, shape in enumerate(data.get("shapes", [])):
        shape["_idx"] = i

    return data


# ---------------------------------------------------------------------------
# Routes — saving corrections
# ---------------------------------------------------------------------------

class ShapeUpdate(BaseModel):
    human_corrected_text: Optional[str] = None
    points: Optional[list] = None          # [[x1,y1],[x2,y2]] for box edits
    label: Optional[str] = None            # label rename


@app.patch("/api/page/shape")
def update_shape(
    folder: str = Query(...),
    stem:   str = Query(...),
    idx:    int = Query(..., description="Shape index (_idx from GET /api/page)"),
    body:   ShapeUpdate = ...,
):
    """Patch one shape in the JSON and write the file back."""
    d  = _resolve_folder(folder)
    jf = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")

    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])

    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail=f"Shape index {idx} out of range (0–{len(shapes)-1})")

    shape = shapes[idx]

    if body.human_corrected_text is not None:
        if "human_output" not in shape:
            shape["human_output"] = {}
        shape["human_output"]["human_corrected_text"] = body.human_corrected_text
        _distribute_flat_to_rows(shape, "human", body.human_corrected_text)

    if body.points is not None:
        _rescale_row_struct(shape, shape.get("points"), body.points)
        shape["points"] = body.points

    if body.label is not None:
        shape["label"] = body.label

    _write_json(jf, data)
    return {"ok": True, "idx": idx}


# ---------------------------------------------------------------------------
# Routes — add / delete shapes
# ---------------------------------------------------------------------------

class NewShape(BaseModel):
    label: str
    points: list
    shape_type: str = "rectangle"


@app.post("/api/page/shape")
def add_shape(
    folder: str = Query(...),
    stem:   str = Query(...),
    body:   NewShape = ...,
):
    """Append a new shape to the JSON and return its index."""
    d  = _resolve_folder(folder)
    jf = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")

    data = json.loads(jf.read_text(encoding="utf-8"))
    new_shape = {
        "label":      body.label,
        "points":     body.points,
        "group_id":   None,
        "shape_type": body.shape_type,
        "flags":      {},
    }
    data["shapes"].append(new_shape)
    new_idx = len(data["shapes"]) - 1
    _write_json(jf, data)
    return {"ok": True, "idx": new_idx}


class ShapesReplace(BaseModel):
    shapes: list


@app.put("/api/page/shapes")
def replace_shapes(
    folder: str = Query(...),
    stem:   str = Query(...),
    body:   ShapesReplace = ...,
):
    """Replace the entire shapes array (used by undo)."""
    d  = _resolve_folder(folder)
    jf = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")
    data = json.loads(jf.read_text(encoding="utf-8"))
    data["shapes"] = body.shapes
    _write_json(jf, data)
    return {"ok": True, "count": len(body.shapes)}


# ---------------------------------------------------------------------------
# Row rules — arithmetic checks between lattice columns (e.g. "1+2=4"),
# stored as JSON in the annotation folder.  Deliberately NOT a .json file so
# the page-listing globs (*.json) never mistake it for a page.
# ---------------------------------------------------------------------------

_RULES_FILENAME = "econai_rules.cfg"


@app.get("/api/rules")
def api_get_rules(folder: str = Query(...)):
    d  = _resolve_folder(folder)
    rf = d / _RULES_FILENAME
    if not rf.exists():
        return {"rules": []}
    try:
        return {"rules": json.loads(rf.read_text(encoding="utf-8")).get("rules", [])}
    except Exception:
        return {"rules": []}


class RulesBody(BaseModel):
    rules: list    # [{"expr": "1+2=4", "name": "male + female = all"}, ...]


@app.put("/api/rules")
def api_put_rules(folder: str = Query(...), body: RulesBody = ...):
    d = _resolve_folder(folder)
    _write_json(d / _RULES_FILENAME, {"rules": body.rules})
    return {"ok": True, "count": len(body.rules)}


# ---------------------------------------------------------------------------
# Clips — a shared integer id stamped on annotations to link the same data
# unit across pages (shape["clip"]).  This endpoint scans every page JSON and
# returns the document-wide index clip -> [{stem, idx}] plus the max id, so
# the editor can show "dangling" clips and mint fresh numbers.
# ---------------------------------------------------------------------------

@app.get("/api/clips")
def api_get_clips(folder: str = Query(...)):
    d = _resolve_folder(folder)
    index: dict = {}
    mx = 0
    for jf in sorted(d.glob("*.json")):
        try:
            shapes = json.loads(jf.read_text(encoding="utf-8")).get("shapes", [])
        except Exception:
            continue
        for i, s in enumerate(shapes):
            c = s.get("clip")
            if c is None:
                continue
            try:
                c = int(c)
            except (TypeError, ValueError):
                continue
            index.setdefault(str(c), []).append({"stem": jf.stem, "idx": i})
            mx = max(mx, c)
    return {"clips": index, "max": mx}


class RuleLlmBody(BaseModel):
    prompt: str
    idxs:   Optional[list] = None    # legacy: whole-shape crops
    labels: Optional[list] = None    # parallel text labels for the images
    crops:  Optional[list] = None    # [{idx, label, y0, y1}] — y in abs page px;
                                     # band crop of one internal row (preferred)


@app.post("/api/page/rule-llm")
def api_rule_llm(
    folder: str = Query(...),
    stem:   str = Query(...),
    model:  str = Query("gpt-4o"),
    body:   RuleLlmBody = ...,
):
    """Rule-fix helper: send one prompt + several labeled cell crops to the
    LLM in a single message and return its raw reply."""
    import base64, io as _io
    from PIL import Image as PILImage

    d        = _resolve_folder(folder)
    jf       = d / f"{stem}.json"
    img_path = _find_image(d, stem)
    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])
    img    = PILImage.open(str(img_path)).convert("RGB")
    w, h   = img.size

    # Normalize the two request shapes into one crop-spec list
    specs = []
    if body.crops:
        specs = body.crops
    elif body.idxs:
        labels = body.labels or [f"Image {i}" for i in body.idxs]
        specs  = [{"idx": i, "label": l} for i, l in zip(body.idxs, labels)]

    content = [{"type": "text", "text": body.prompt}]
    for spec in specs:
        idx = int(spec.get("idx", -1))
        if idx < 0 or idx >= len(shapes):
            continue
        x1, y1, x2, y2 = _shape_bbox(shapes[idx])
        pad = 4
        if spec.get("y0") is not None and spec.get("y1") is not None:
            cy0 = max(0, int(spec["y0"]) - 2)
            cy1 = min(h, int(spec["y1"]) + 2)
        else:
            cy0 = max(0, int(y1) - pad)
            cy1 = min(h, int(y2) + pad)
        crop = img.crop((max(0, int(x1) - pad), cy0,
                         min(w, int(x2) + pad), max(cy0 + 1, cy1)))
        # Upscale small band crops so digits stay readable for the LLM
        if crop.height < 48:
            scale = 48 / max(1, crop.height)
            crop  = crop.resize((max(1, int(crop.width * scale)), 48),
                                PILImage.LANCZOS)
        buf = _io.BytesIO()
        crop.save(buf, format="JPEG", quality=92)
        b64 = base64.b64encode(buf.getvalue()).decode()
        content.append({"type": "text", "text": str(spec.get("label", ""))})
        content.append({"type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{b64}",
                                      "detail": "high"}})

    try:
        client = _make_llm_client(model)
        resp   = _llm_complete(client, model,
                               [{"role": "user", "content": content}], 2048)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc))
    return {"response": (resp.choices[0].message.content or "").strip()}


# ---------------------------------------------------------------------------
# Route — LLM layout detection (experimental: whole page -> annotation boxes)
# ---------------------------------------------------------------------------

class LlmLayoutBody(BaseModel):
    prompt: str


@app.post("/api/page/llm-layout")
def api_llm_layout(
    folder: str = Query(...),
    stem:   str = Query(...),
    model:  str = Query("gpt-4o"),
    body:   LlmLayoutBody = ...,
):
    """Send the whole page image to the LLM and parse its returned bounding
    boxes into LabelMe shapes. Coordinates are expected as fractions of the
    image (0..1); a 0..1000 convention is auto-detected. Returns the parsed
    shapes (image-pixel points) + the raw reply — the client applies them so
    undo works."""
    import base64, io, json as _json, re
    from PIL import Image as PILImage

    d        = _resolve_folder(folder)
    img_path = _find_image(d, stem)
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    img = PILImage.open(str(img_path)).convert("RGB")
    W, H = img.size
    # Downscale for the LLM (coords are normalized, so the original W,H still
    # map correctly); keep aspect ratio, long side <= 2048
    send = img
    longest = max(W, H)
    if longest > 2048:
        scale = 2048 / longest
        send = img.resize((max(1, int(W * scale)), max(1, int(H * scale))), PILImage.LANCZOS)
    buf = io.BytesIO(); send.save(buf, format="JPEG", quality=90)
    b64 = base64.b64encode(buf.getvalue()).decode()

    content = [
        {"type": "text", "text": body.prompt},
        {"type": "image_url",
         "image_url": {"url": f"data:image/jpeg;base64,{b64}", "detail": "high"}},
    ]
    try:
        client = _make_llm_client(model)
        resp   = _llm_complete(client, model, [{"role": "user", "content": content}], 4096)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc))
    raw = (resp.choices[0].message.content or "").strip()

    # Parse the first JSON array in the reply
    a, b = raw.find("["), raw.rfind("]")
    items = []
    if 0 <= a < b:
        try:
            items = _json.loads(raw[a:b + 1])
        except Exception:
            items = []

    # Auto-detect coordinate scale: fractions (<=1) vs 0..1000
    allnums = []
    for it in items:
        box = it.get("box") or it.get("bbox") or it.get("points")
        if isinstance(box, list) and len(box) >= 4:
            allnums += [float(x) for x in box[:4] if isinstance(x, (int, float))]
    scale = 1000.0 if (allnums and max(allnums) > 1.5) else 1.0

    shapes = []
    for it in items:
        if not isinstance(it, dict):
            continue
        label = str(it.get("label") or it.get("name") or "region").strip()
        box = it.get("box") or it.get("bbox") or it.get("points")
        if not (isinstance(box, list) and len(box) >= 4):
            continue
        try:
            x1, y1, x2, y2 = (float(box[0]), float(box[1]), float(box[2]), float(box[3]))
        except (TypeError, ValueError):
            continue
        x1, x2 = sorted((x1 / scale * W, x2 / scale * W))
        y1, y2 = sorted((y1 / scale * H, y2 / scale * H))
        x1 = max(0, min(W, x1)); x2 = max(0, min(W, x2))
        y1 = max(0, min(H, y1)); y2 = max(0, min(H, y2))
        if x2 - x1 < 1 or y2 - y1 < 1:
            continue
        shapes.append({"label": label,
                       "points": [[round(x1, 1), round(y1, 1)], [round(x2, 1), round(y2, 1)]],
                       "group_id": None, "shape_type": "rectangle", "flags": {},
                       "llm_layout": True})

    return {"response": raw, "shapes": shapes, "count": len(shapes),
            "image_w": W, "image_h": H}


# ---------------------------------------------------------------------------
# Internal row structure (row_struct)
#
# Each shape may carry exactly one row_struct describing its internal rows:
#   {"version": 1, "origin": "linebyline|anchored|converted|manual",
#    "rows": [{"n": 1, "y0": <abs page Y>, "y1": <abs page Y>,
#              "ocr": "...", "llm": "...", "human": "..."}, ...]}
# Coordinates are absolute page pixels; X is implicit (the shape bbox).
# The legacy flat text fields are re-derived from row_struct on every write,
# so all existing consumers (Excel export, diagnostics, batch conditions,
# anchoring source lookups) keep working unchanged.
# ---------------------------------------------------------------------------

_ROW_LAYERS = ("ocr", "llm", "human")


def _shape_bbox(shape):
    pts = shape["points"]
    xs = [p[0] for p in pts]; ys = [p[1] for p in pts]
    return min(xs), min(ys), max(xs), max(ys)


def _sync_flat_from_rows(shape, layers=("human",)):
    """Re-derive flat text fields from row_struct for the given layers.

    Default = HUMAN ONLY. The flat OCR/LLM fields are the models' original
    outputs — overwriting them with the row-join on every row save (the old
    behavior) silently destroyed the originals: after inserting a row the
    flat text gained a phantom empty line, which made the panel's ⟳
    "re-distribute" a circular no-op forever. Only a line-by-line run that
    rewrites a layer should sync that layer's flat (pass it in `layers`).
    A layer whose rows are all empty is left untouched."""
    rs = shape.get("row_struct")
    if not rs or not rs.get("rows"):
        return
    rows = rs["rows"]

    def joined(layer):
        vals = [(r.get(layer) or "") for r in rows]
        return "\n".join(vals) if any(v.strip() for v in vals) else None

    if "ocr" in layers:
        t = joined("ocr")
        if t is not None:
            if not isinstance(shape.get("tesseract_output"), dict):
                shape["tesseract_output"] = {}
            shape["tesseract_output"]["ocr_text"] = t
    if "llm" in layers:
        t = joined("llm")
        if t is not None:
            if not isinstance(shape.get("openai_output"), dict):
                shape["openai_output"] = {}
            shape["openai_output"]["response"] = t
    if "human" in layers:
        t = joined("human")
        if t is not None:
            if not isinstance(shape.get("human_output"), dict):
                shape["human_output"] = {}
            shape["human_output"]["human_corrected_text"] = t


def _apply_layer_rows(shape, bands_abs, layer, texts, origin, force_boxes=False):
    """Write one layer's per-row texts into the shape's row_struct.
    If a structure with the same row count already exists, its boxes are kept
    and only this layer's values change; otherwise the structure is rebuilt
    from bands_abs (other layers survive when the row count matches).
    force_boxes=True always adopts bands_abs as the stored boxes — used by
    row-structure anchoring, where the projected bands are authoritative."""
    rs = shape.get("row_struct")
    if not force_boxes and rs and len(rs.get("rows", [])) == len(texts):
        for r, t in zip(rs["rows"], texts):
            r[layer] = t
            if layer == "llm":
                r.pop("llm_fixed", None)   # fresh LLM read replaces a rule-fix
    else:
        old = (rs or {}).get("rows", [])
        new_rows = []
        for i, ((b0, b1), t) in enumerate(zip(bands_abs, texts)):
            row = {"n": i + 1, "y0": float(b0), "y1": float(b1),
                   "ocr": "", "llm": "", "human": ""}
            row[layer] = t
            if len(old) == len(texts):
                for lay in _ROW_LAYERS:
                    if lay != layer and old[i].get(lay):
                        row[lay] = old[i][lay]
                        if lay == "llm" and old[i].get("llm_fixed"):
                            row["llm_fixed"] = True
                if old[i].get("pdf") is not None:
                    row["pdf"] = old[i]["pdf"]
            new_rows.append(row)
        shape["row_struct"] = {"version": 1, "origin": origin, "rows": new_rows}

    # A flat Human correction must never vanish from the table: if the rows'
    # human column is empty but flat human text exists (e.g. the structure
    # was just created by anchoring), pull it in top-aligned — extra lines
    # beyond the row count are truncated.
    rows_now = shape["row_struct"]["rows"]
    if layer != "human" and not any((r.get("human") or "").strip() for r in rows_now):
        flat_h = (shape.get("human_output") or {}).get("human_corrected_text") or ""
        if flat_h.strip():
            for r, t in zip(rows_now, _split_lines(flat_h)):
                r["human"] = t

    # a line-by-line run REWROTE this layer — its flat mirrors the new rows
    _sync_flat_from_rows(shape, layers=(layer, "human"))


def _split_lines(text):
    """Line split that PRESERVES row positions: leading/interior empty lines
    are kept — in a row-join they mark empty rows, and stripping them (the old
    behavior) shifted every value up so imports landed in the first N rows.
    Only CRs and trailing blank lines (a stray final newline) are dropped."""
    lines = [l.rstrip("\r") for l in (text or "").split("\n")]
    while lines and not lines[-1].strip():
        lines.pop()
    return lines


def _distribute_flat_to_rows(shape, layer, text):
    """Whole-cell OCR/LLM/human writes: push the flat text into the existing
    row_struct when the line counts match, so the table view stays in sync."""
    rs = shape.get("row_struct")
    if not rs or not rs.get("rows"):
        return
    lines = _split_lines(text)
    # fewer lines than rows = the missing tail rows are empty (a row-join
    # loses trailing empty rows on split); more lines than rows = a real
    # mismatch, leave the rows alone
    if lines and len(lines) <= len(rs["rows"]):
        lines += [""] * (len(rs["rows"]) - len(lines))
        for r, t in zip(rs["rows"], lines):
            r[layer] = t
            if layer == "llm":
                r.pop("llm_fixed", None)


def _existing_row_bands_rel(shape, crop_top, crop_h):
    """Return the existing row_struct bands converted to crop-relative
    coordinates (clamped), or None.  Used so re-running a layer reuses the
    established structure instead of re-detecting rows."""
    rs = shape.get("row_struct")
    if not rs or not rs.get("rows"):
        return None
    bands = []
    for r in rs["rows"]:
        t = max(0, int(round(r["y0"] - crop_top)))
        b = min(crop_h, int(round(r["y1"] - crop_top)))
        if b <= t:
            b = min(crop_h, t + 1)
        bands.append((t, b))
    return bands


def _rows_for_source(shape, crop, crop_top, cell_height, rows_source):
    """Row bands for a per-internal-row run, honoring the scope choice:
    'existing' = the cell's stored bands (400 if none), 'detect' = always
    re-detect from pixels, 'auto' = existing else detect."""
    existing = _existing_row_bands_rel(shape, crop_top, crop.height)
    if rows_source == "existing":
        if not existing:
            raise HTTPException(status_code=400,
                detail="This cell has no internal row structure — build one "
                       "first, or use 're-detect rows'")
        return existing
    if rows_source == "detect":
        return _detect_text_rows(crop, cell_height)
    return existing or _detect_text_rows(crop, cell_height)


def _project_ref_bands(ref_shape, crop_top, crop_h, ty1, ty2):
    """Project a reference shape's row_struct bands onto a target cell:
    linear map from the reference bbox y-range to the target bbox y-range,
    returned crop-relative and clamped.  None if the reference has no rows."""
    rows = ((ref_shape or {}).get("row_struct") or {}).get("rows") or []
    if not rows:
        return None
    _, ry1, _, ry2 = _shape_bbox(ref_shape)
    rh = (ry2 - ry1) or 1.0
    th = ty2 - ty1
    bands = []
    for r in rows:
        a = ty1 + (r["y0"] - ry1) * th / rh
        b = ty1 + (r["y1"] - ry1) * th / rh
        t  = max(0, int(round(a - crop_top)))
        bb = min(crop_h, int(round(b - crop_top)))
        if bb <= t:
            bb = min(crop_h, t + 1)
        bands.append((t, bb))
    return bands


def _rescale_row_struct(shape, old_pts, new_pts):
    """Linearly remap row bands when the shape bbox changes (move / resize)."""
    rs = shape.get("row_struct")
    if not rs or not rs.get("rows") or not old_pts or not new_pts:
        return
    oy = [p[1] for p in old_pts]; ny = [p[1] for p in new_pts]
    oy1, oy2 = min(oy), max(oy)
    ny1, ny2 = min(ny), max(ny)
    oh, nh = oy2 - oy1, ny2 - ny1
    if oh <= 0 or nh <= 0 or (oy1 == ny1 and oy2 == ny2):
        return
    for r in rs["rows"]:
        r["y0"] = ny1 + (r["y0"] - oy1) * nh / oh
        r["y1"] = ny1 + (r["y1"] - oy1) * nh / oh


class RowStructBody(BaseModel):
    rows:   list                    # [{y0, y1, ocr, llm, human}, ...] abs page Y
    origin: Optional[str] = None


@app.patch("/api/page/shape/rows")
def update_row_struct(
    folder: str = Query(...),
    stem:   str = Query(...),
    idx:    int = Query(...),
    body:   RowStructBody = ...,
):
    """Replace a shape's internal row structure (divider edits, human cell
    edits, removal).  An empty rows list deletes the structure."""
    d  = _resolve_folder(folder)
    jf = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")
    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")
    shape = shapes[idx]

    if not body.rows:
        shape.pop("row_struct", None)
    else:
        rs = shape.get("row_struct") or {"version": 1, "origin": body.origin or "manual"}
        if body.origin:
            rs["origin"] = body.origin
        rows = []
        for r in body.rows:
            row = {"n": 0, "y0": float(r["y0"]), "y1": float(r["y1"]),
                   "ocr": r.get("ocr") or "", "llm": r.get("llm") or "",
                   "human": r.get("human") or ""}
            if r.get("pdf") is not None:
                row["pdf"] = r["pdf"]
            if r.get("llm_fixed"):
                row["llm_fixed"] = True
            if r.get("authority"):          # resolved gazetteer entity (per row)
                row["authority"] = r["authority"]
            if r.get("blank"):              # structural blank (ink scan)
                row["blank"] = True
            rows.append(row)
        rows.sort(key=lambda r: r["y0"])
        for i, r in enumerate(rows):
            r["n"] = i + 1
        rs["rows"] = rows
        shape["row_struct"] = rs
        _sync_flat_from_rows(shape)

    _write_json(jf, data)
    return {"ok": True, "row_struct": shape.get("row_struct")}


@app.post("/api/page/shape/rows/convert")
def api_rows_convert(
    folder: str  = Query(...),
    stem:   str  = Query(...),
    idx:    int  = Query(...),
    force:  bool = Query(False, description="Rebuild even if row_struct exists"),
):
    """Build row_struct for a legacy annotation: the line count of the best
    text layer (Human > LLM > OCR) fixes N, then the same histogram-valley
    algorithm used by anchoring finds the band boundaries.  Layers whose line
    count matches N are distributed into the rows; mismatched layers stay
    empty in the table (their flat text is preserved)."""
    d        = _resolve_folder(folder)
    jf       = d / f"{stem}.json"
    img_path = _find_image(d, stem)
    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")
    shape = shapes[idx]

    if shape.get("row_struct") and not force:
        raise HTTPException(status_code=400, detail="row_struct already exists (use force=true)")

    human = (shape.get("human_output")    or {}).get("human_corrected_text") or ""
    llm   = (shape.get("openai_output")   or {}).get("response") or ""
    ocr   = ((shape.get("tesseract_output") or {}).get("ocr_text")
             or (shape.get("easyocr_output") or {}).get("ocr_text") or "")
    best  = human or llm or ocr
    if not best.strip():
        raise HTTPException(status_code=400, detail="Shape has no text in any layer")
    n = len(_split_lines(best))

    x1, y1, x2, y2 = _shape_bbox(shape)
    shadow = _get_shadow_page(folder, stem, img_path)
    iw, ih = shadow.size
    pad    = 4
    cy1    = max(0, int(y1) - pad)
    crop   = shadow.crop((
        max(0, int(x1) - pad), cy1,
        min(iw, int(x2) + pad), min(ih, int(y2) + pad),
    ))
    bands     = _split_into_n_rows(crop, n)
    bands_abs = [(t + cy1, b + cy1) for t, b in bands]

    layers = {"human": _split_lines(human) if human.strip() else None,
              "llm":   _split_lines(llm)   if llm.strip()   else None,
              "ocr":   _split_lines(ocr)   if ocr.strip()   else None}
    rows = []
    for i, (b0, b1) in enumerate(bands_abs):
        row = {"n": i + 1, "y0": float(b0), "y1": float(b1),
               "ocr": "", "llm": "", "human": ""}
        for lay, lines in layers.items():
            if lines and len(lines) == n:
                row[lay] = lines[i]
        rows.append(row)

    shape["row_struct"] = {"version": 1, "origin": "converted", "rows": rows}
    _sync_flat_from_rows(shape)
    _write_json(jf, data)
    return {"ok": True, "rows": len(rows), "row_struct": shape["row_struct"]}


@app.post("/api/page/shape/rows/pdf-refresh")
def api_rows_pdf_refresh(
    folder: str = Query(...),
    stem:   str = Query(...),
    idx:    int = Query(...),
):
    """Re-extract the PDF text layer for one shape, clipped per internal row
    band — each row gets exactly the PDF words inside its own band, stored as
    row['pdf'] (display-only, not part of the OCR/LLM/Human layer triad)."""
    import fitz

    d  = _resolve_folder(folder)
    jf = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")
    shape = shapes[idx]
    rs    = shape.get("row_struct")
    if not rs or not rs.get("rows"):
        raise HTTPException(status_code=400, detail="Shape has no internal row structure")

    pdf_page  = data.get("pdf_page")
    pdf_scale = data.get("pdf_scale", 2.0)
    if not data.get("pdf_source") or pdf_page is None:
        raise HTTPException(status_code=400, detail="No PDF source recorded for this page")
    pdf_path = _resolve_pdf_source(d, data.get("pdf_source"))
    if pdf_path is None:
        raise HTTPException(status_code=404,
            detail=f"Source PDF not found (looked for '{Path(data['pdf_source']).name}' "
                   f"in the project's sources/ folder)")

    doc  = fitz.open(str(pdf_path))
    page = doc[int(pdf_page)]
    x1, _, x2, _ = _shape_bbox(shape)
    for r in rs["rows"]:
        rect = fitz.Rect(x1 / pdf_scale, r["y0"] / pdf_scale,
                         x2 / pdf_scale, r["y1"] / pdf_scale)
        r["pdf"] = " ".join(page.get_text("text", clip=rect).split())
    doc.close()

    _write_json(jf, data)
    return {"ok": True, "row_struct": shape["row_struct"]}


@app.delete("/api/page/shape")
def delete_shape(
    folder: str = Query(...),
    stem:   str = Query(...),
    idx:    int = Query(...),
):
    """Remove one shape from the JSON by index."""
    d  = _resolve_folder(folder)
    jf = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")

    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail=f"Shape index {idx} out of range")

    shapes.pop(idx)
    _write_json(jf, data)
    return {"ok": True, "remaining": len(shapes)}


# ---------------------------------------------------------------------------
# Routes — PDF text layer extraction
# ---------------------------------------------------------------------------

@app.post("/api/page/pdf-text-layer")
def api_pdf_text_layer(
    folder: str = Query(...),
    stem:   str = Query(...),
):
    """Extract text from the PDF text layer for every shape on a page.

    Requires the page JSON to have pdf_source / pdf_page / pdf_scale fields,
    which are written automatically when importing from a PDF.
    Saves pdf_text on each shape and returns the updated shapes list.
    """
    import fitz

    d   = _resolve_folder(folder)
    jf  = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")

    data = json.loads(jf.read_text(encoding="utf-8"))

    pdf_page  = data.get("pdf_page")
    pdf_scale = data.get("pdf_scale", 2.0)

    if not data.get("pdf_source") or pdf_page is None:
        raise HTTPException(
            status_code=400,
            detail="No PDF source recorded for this page. "
                   "Only pages imported from a PDF with the current version "
                   "of the app have this information."
        )

    pdf_path = _resolve_pdf_source(d, data.get("pdf_source"))
    if pdf_path is None:
        raise HTTPException(
            status_code=404,
            detail=f"Source PDF '{Path(data['pdf_source']).name}' not found in the "
                   "project's sources/ folder. Re-import the PDF or place it there."
        )

    doc  = fitz.open(str(pdf_path))
    page = doc[int(pdf_page)]

    shapes  = data.get("shapes", [])
    updated = 0

    for shape in shapes:
        pts = shape.get("points", [])
        if len(pts) < 2:
            continue
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        # Convert image pixel coords → PDF point coords
        x0, y0 = min(xs) / pdf_scale, min(ys) / pdf_scale
        x1, y1 = max(xs) / pdf_scale, max(ys) / pdf_scale
        rect = fitz.Rect(x0, y0, x1, y1)
        text = page.get_text("text", clip=rect).strip()
        shape["pdf_text"] = text
        updated += 1

    doc.close()

    _write_json(jf, data)
    return {"ok": True, "updated": updated, "shapes": shapes}


# ---------------------------------------------------------------------------
# Routes — cell image crop
# ---------------------------------------------------------------------------

@app.get("/api/cell")
def get_cell(
    folder: str  = Query(...),
    stem:   str  = Query(...),
    idx:    int  = Query(...),
    pad:    int  = Query(4, description="Padding in pixels"),
    shadow: bool = Query(False, description="Return shadow (line-erased) version"),
    y0:     Optional[float] = Query(None, description="Crop only this y-band (abs page px) — for one internal row"),
    y1:     Optional[float] = Query(None, description="Band bottom (abs page px)"),
):
    """Return a cropped image of a single cell (for the right panel zoom).
    With y0/y1, crops just that vertical band — a single internal row."""
    from PIL import Image
    import io
    from fastapi.responses import StreamingResponse

    d   = _resolve_folder(folder)
    jf  = d / f"{stem}.json"
    img_path = _find_image(d, stem)

    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")

    pts = shapes[idx]["points"]
    xs  = [p[0] for p in pts]
    ys  = [p[1] for p in pts]
    x1, y1b, x2, y2 = min(xs), min(ys), max(xs), max(ys)
    top = y0 if y0 is not None else y1b
    bot = y1 if y1 is not None else y2

    img  = _get_shadow_page(folder, stem, img_path) if shadow else Image.open(str(img_path))
    w, h = img.size
    crop = img.crop((
        max(0, int(x1) - pad),
        max(0, int(top) - pad),
        min(w, int(x2) + pad),
        min(h, int(bot) + pad),
    ))

    buf = io.BytesIO()
    crop.save(buf, format="JPEG", quality=90)
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/jpeg")


@app.post("/api/page/shape/ocr")
def api_ocr_cell(
    folder: str = Query(...),
    stem:   str = Query(...),
    idx:    int = Query(...),
    lang:   str = Query("hun", description="Tesseract language code"),
):
    """Run Tesseract OCR on a single cell and store the result in the page JSON."""
    import pytesseract
    from PIL import Image
    # Ensure Tesseract binary is found on Windows even if not on PATH
    import shutil
    if not shutil.which("tesseract"):
        pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

    d        = _resolve_folder(folder)
    jf       = d / f"{stem}.json"
    img_path = _find_image(d, stem)

    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")

    pts = shapes[idx]["points"]
    xs  = [p[0] for p in pts]
    ys  = [p[1] for p in pts]
    x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)

    shadow     = _get_shadow_page(folder, stem, img_path)
    sw, sh     = shadow.size
    pad        = 4
    crop_grey  = shadow.crop((
        max(0, int(x1) - pad), max(0, int(y1) - pad),
        min(sw, int(x2) + pad), min(sh, int(y2) + pad),
    )).convert("L")

    tess = pytesseract.image_to_data(
        crop_grey, lang=lang,
        config="--psm 4",          # single column, preserves line breaks
        output_type=pytesseract.Output.DICT,
    )

    # Group confident words by (block, paragraph, line) to preserve layout.
    # Also track the minimum top-coordinate of each line so we can sort by
    # actual pixel position (Tesseract block_num order ≠ visual top-to-bottom order).
    from collections import defaultdict
    line_words: dict = defaultdict(list)
    line_top:   dict = defaultdict(lambda: float("inf"))
    conf_values: list = []
    for txt, conf, blk, par, ln, top in zip(
        tess["text"], tess["conf"],
        tess["block_num"], tess["par_num"], tess["line_num"],
        tess["top"],
    ):
        c = int(conf)
        if txt.strip() and c > 0:
            key = (blk, par, ln)
            line_words[key].append(txt)
            line_top[key] = min(line_top[key], int(top))
            conf_values.append(c)

    # Sort lines by their top pixel coordinate (visual reading order)
    sorted_keys = sorted(line_words.keys(), key=lambda k: line_top[k])
    lines = [" ".join(line_words[k]) for k in sorted_keys]
    ocr_text  = "\n".join(lines).strip()
    mean_conf = round(sum(conf_values) / len(conf_values), 1) if conf_values else 0.0

    shapes[idx]["tesseract_output"] = {
        "ocr_text":  ocr_text,
        "mean_conf": mean_conf,
        "lang":      lang,
    }
    _distribute_flat_to_rows(shapes[idx], "ocr", ocr_text)
    _write_json(jf, data)

    return {"ocr_text": ocr_text, "mean_conf": mean_conf}


# ---------------------------------------------------------------------------
# Shadow page cache — long table-line removal for cleaner OCR input
# ---------------------------------------------------------------------------
# OCR settings — tunable preprocessing parameters, persisted to disk
# ---------------------------------------------------------------------------
_OCR_SETTINGS_PATH = Path(__file__).parent / "ocr_settings.json"

_OCR_SETTINGS_DEFAULTS: dict = {
    "line_removal_fraction":  0.22,
    "v_line_fraction":        0.0,   # vertical line min-length override (0 = use line_removal_fraction)
    "use_adaptive_threshold": True,
    "adaptive_block_size":    15,
    "adaptive_c":             10,
    "morph_open_kernel":      2,
    # Line detection tuning
    "line_close_kernel":      0,     # horizontal close before detection (bridges dashes, 0=off)
    "line_dilate_thickness":  3,     # dilation after detection to cover line thickness
    # Post-line-removal output processing
    "blur_sigma":             0,     # Gaussian blur σ applied to output (0 = off)
    "output_binarize":        False, # binarize the final output image
    "output_open_kernel":     0,     # morphological opening on output (0 = off)
}


def _load_ocr_settings() -> dict:
    if _OCR_SETTINGS_PATH.exists():
        try:
            return {**_OCR_SETTINGS_DEFAULTS,
                    **json.loads(_OCR_SETTINGS_PATH.read_text(encoding="utf-8"))}
        except Exception:
            pass
    return dict(_OCR_SETTINGS_DEFAULTS)


_ocr_settings: dict = _load_ocr_settings()


class OcrSettingsBody(BaseModel):
    line_removal_fraction:  Optional[float] = None
    use_adaptive_threshold: Optional[bool]  = None
    adaptive_block_size:    Optional[int]   = None
    adaptive_c:             Optional[int]   = None
    morph_open_kernel:      Optional[int]   = None
    v_line_fraction:        Optional[float] = None
    line_close_kernel:      Optional[int]   = None
    line_dilate_thickness:  Optional[int]   = None
    blur_sigma:             Optional[int]   = None
    output_binarize:        Optional[bool]  = None
    output_open_kernel:     Optional[int]   = None


@app.get("/api/ocr-settings")
def api_get_ocr_settings():
    return _ocr_settings


@app.post("/api/ocr-settings")
def api_set_ocr_settings(body: OcrSettingsBody, save: bool = Query(False)):
    global _ocr_settings
    if body.line_removal_fraction is not None:
        _ocr_settings["line_removal_fraction"] = max(0.01, min(0.9, body.line_removal_fraction))
    if body.use_adaptive_threshold is not None:
        _ocr_settings["use_adaptive_threshold"] = body.use_adaptive_threshold
    if body.adaptive_block_size is not None:
        bs = max(3, body.adaptive_block_size)
        if bs % 2 == 0: bs += 1
        _ocr_settings["adaptive_block_size"] = bs
    if body.adaptive_c is not None:
        _ocr_settings["adaptive_c"] = max(0, min(50, body.adaptive_c))
    if body.morph_open_kernel is not None:
        _ocr_settings["morph_open_kernel"] = max(0, min(10, body.morph_open_kernel))
    if body.v_line_fraction is not None:
        _ocr_settings["v_line_fraction"] = max(0.0, min(0.9, body.v_line_fraction))
    if body.line_close_kernel is not None:
        _ocr_settings["line_close_kernel"] = max(0, min(30, body.line_close_kernel))
    if body.line_dilate_thickness is not None:
        _ocr_settings["line_dilate_thickness"] = max(1, min(15, body.line_dilate_thickness))
    if body.blur_sigma is not None:
        _ocr_settings["blur_sigma"] = max(0, min(10, body.blur_sigma))
    if body.output_binarize is not None:
        _ocr_settings["output_binarize"] = body.output_binarize
    if body.output_open_kernel is not None:
        _ocr_settings["output_open_kernel"] = max(0, min(10, body.output_open_kernel))
    _shadow_page_cache.clear()
    if save:
        _OCR_SETTINGS_PATH.write_text(
            json.dumps(_ocr_settings, indent=2), encoding="utf-8"
        )
    return _ocr_settings


# ---------------------------------------------------------------------------
# Shadow page cache — table line removal for cleaner OCR input
# ---------------------------------------------------------------------------
_shadow_page_cache: dict = {}


def _build_shadow_page(pil_image, settings: dict):
    """
    Line removal only — the original grayscale is preserved intact.

    Binarisation (adaptive threshold or Otsu) is used solely to make line
    detection reliable.  After the line mask is found, we revert to the
    original grayscale and paint only the detected line pixels white.
    EasyOCR therefore sees the natural scan, minus the table rules.
    """
    import cv2
    import numpy as np
    from PIL import Image as _PIL

    img = np.array(pil_image.convert("L"))   # original grayscale — final output base
    h, w = img.shape

    # ── Step 1: binarise for line detection only ──────────────────────────────
    if settings.get("use_adaptive_threshold", True):
        block = settings.get("adaptive_block_size", 15)
        if block % 2 == 0:
            block += 1
        binary = cv2.adaptiveThreshold(
            img, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,
            block, settings.get("adaptive_c", 10),
        )
        k_sz = settings.get("morph_open_kernel", 0)
        if k_sz > 0:
            _k     = cv2.getStructuringElement(cv2.MORPH_RECT, (k_sz, k_sz))
            binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN, _k)
        morph_in = cv2.bitwise_not(binary)   # text/lines = 255, bg = 0
    else:
        # Otsu for line detection only
        _, morph_in = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    # ── Step 2: detect long lines in the binarised image ─────────────────────
    frac   = settings.get("line_removal_fraction", 0.22)
    v_frac = settings.get("v_line_fraction", 0.0) or frac  # 0 means "use same as horizontal"

    # Optional: close gaps in dashed/dotted lines before detection
    close_k = settings.get("line_close_kernel", 0)
    detect_in = morph_in
    if close_k > 0:
        hc = cv2.getStructuringElement(cv2.MORPH_RECT, (close_k, 1))
        vc = cv2.getStructuringElement(cv2.MORPH_RECT, (1, close_k))
        detect_in = cv2.morphologyEx(detect_in, cv2.MORPH_CLOSE, hc)
        detect_in = cv2.morphologyEx(detect_in, cv2.MORPH_CLOSE, vc)

    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(1, int(w * frac)),  1))
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(1, int(h * v_frac))))

    dilate_t = settings.get("line_dilate_thickness", 3)
    mask = cv2.dilate(
        cv2.bitwise_or(
            cv2.morphologyEx(detect_in, cv2.MORPH_OPEN, h_kernel),
            cv2.morphologyEx(detect_in, cv2.MORPH_OPEN, v_kernel),
        ),
        np.ones((dilate_t, dilate_t), np.uint8),
    )

    # ── Step 3: revert to original grayscale, paint line pixels white ─────────
    result = img.copy()
    result[mask > 0] = 255

    # ── Step 4 (optional): Gaussian blur to smooth noise ─────────────────────
    blur_sigma = settings.get("blur_sigma", 0)
    if blur_sigma > 0:
        result = cv2.GaussianBlur(result, (0, 0), blur_sigma)

    # ── Step 5 (optional): binarize the output image ──────────────────────────
    if settings.get("output_binarize", False):
        block = settings.get("adaptive_block_size", 15)
        if block % 2 == 0:
            block += 1
        result = cv2.adaptiveThreshold(
            result, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,
            block, settings.get("adaptive_c", 10),
        )

    # ── Step 6 (optional): morphological opening on output (speckle removal) ─
    out_open = settings.get("output_open_kernel", 0)
    if out_open > 0:
        _k = cv2.getStructuringElement(cv2.MORPH_RECT, (out_open, out_open))
        result = cv2.morphologyEx(result, cv2.MORPH_OPEN, _k)

    return _PIL.fromarray(result).convert("RGB")


def _get_shadow_page(folder: str, stem: str, img_path):
    """Return (and cache) the fully-preprocessed shadow page.
    Cache key includes every setting that affects the output."""
    from PIL import Image as _PIL
    s   = _ocr_settings
    key = (
        folder, stem,
        s["line_removal_fraction"],
        s.get("v_line_fraction", 0.0),
        s["use_adaptive_threshold"],
        s["adaptive_block_size"],
        s["adaptive_c"],
        s["morph_open_kernel"],
        s.get("line_close_kernel", 0),
        s.get("line_dilate_thickness", 3),
        s.get("blur_sigma", 0),
        s.get("output_binarize", False),
        s.get("output_open_kernel", 0),
    )
    if key not in _shadow_page_cache:
        if len(_shadow_page_cache) >= 10:
            _shadow_page_cache.clear()
        full = _PIL.open(str(img_path)).convert("RGB")
        _shadow_page_cache[key] = _build_shadow_page(full, dict(s))
    return _shadow_page_cache[key]


@app.get("/api/shadow-preview")
def api_shadow_preview(folder: str = Query(...), stem: str = Query(...)):
    """Return the fully preprocessed shadow page as JPEG (OCR view preview)."""
    import io as _io

    d        = _resolve_folder(folder)
    img_path = _find_image(d, stem)
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    shadow = _get_shadow_page(folder, stem, img_path)

    buf = _io.BytesIO()
    shadow.save(buf, format="JPEG", quality=85)
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/jpeg",
                             headers={"Cache-Control": "no-cache, no-store"})


# ---------------------------------------------------------------------------
# EasyOCR — cached reader (initialisation takes a few seconds the first time)
# ---------------------------------------------------------------------------
_easyocr_reader = None

def _get_easyocr_reader(langs: list[str]):
    global _easyocr_reader
    import easyocr, io, sys
    key = tuple(sorted(langs))
    if _easyocr_reader is None or _easyocr_reader[0] != key:
        # Suppress stdout/stderr during init to avoid cp1250 UnicodeEncodeError
        # from EasyOCR's progress-bar block characters on Windows
        old_out, old_err = sys.stdout, sys.stderr
        sys.stdout = sys.stderr = io.StringIO()
        try:
            reader = easyocr.Reader(langs, gpu=True)
        finally:
            sys.stdout, sys.stderr = old_out, old_err
        _easyocr_reader = (key, reader)
    return _easyocr_reader[1]


@app.post("/api/page/shape/ocr/easyocr")
def api_ocr_easyocr(
    folder: str = Query(...),
    stem:   str = Query(...),
    idx:    int = Query(...),
    langs:  str = Query("en,hu", description="Comma-separated EasyOCR language codes"),
):
    """Run EasyOCR on a single cell and store the result in the page JSON."""
    import numpy as np
    from PIL import Image as PILImage

    d        = _resolve_folder(folder)
    jf       = d / f"{stem}.json"
    img_path = _find_image(d, stem)

    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    data_doc = json.loads(jf.read_text(encoding="utf-8"))
    shapes   = data_doc.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")

    pts = shapes[idx]["points"]
    xs  = [p[0] for p in pts]; ys = [p[1] for p in pts]
    x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)

    shadow = _get_shadow_page(folder, stem, img_path)
    iw, ih = shadow.size
    pad    = 4
    crop   = shadow.crop((
        max(0, int(x1) - pad), max(0, int(y1) - pad),
        min(iw, int(x2) + pad), min(ih, int(y2) + pad),
    )).convert("RGB")

    lang_list = [l.strip() for l in langs.split(",") if l.strip()]
    reader    = _get_easyocr_reader(lang_list)

    results = reader.readtext(
        np.array(crop),
        allowlist="0123456789-",
        min_size=6,          # catch single-digit cells (default 20 misses them)
        text_threshold=0.5,  # slightly more permissive for sparse content
        low_text=0.3,
    )
    # results: [([[x1,y1],[x2,y2],[x3,y3],[x4,y4]], text, confidence), ...]
    # Sort top-to-bottom by the minimum y of each bounding box
    results.sort(key=lambda r: min(pt[1] for pt in r[0]))

    lines      = [text for _, text, _ in results]
    confs      = [conf for _, _, conf in results]
    ocr_text   = "\n".join(lines).strip()
    mean_conf  = round(sum(confs) / len(confs) * 100, 1) if confs else 0.0

    shapes[idx]["tesseract_output"] = {
        "ocr_text":  ocr_text,
        "mean_conf": mean_conf,
        "engine":    "easyocr",
        "langs":     lang_list,
    }
    _distribute_flat_to_rows(shapes[idx], "ocr", ocr_text)
    _write_json(jf, data_doc)

    return {"ocr_text": ocr_text, "mean_conf": mean_conf}


@app.post("/api/page/shape/ocr/linebyline")
async def api_ocr_linebyline(
    folder:      str = Query(...),
    stem:        str = Query(...),
    idx:         int = Query(...),
    cell_height: int = Query(26),
    lang:        str = Query("hun"),
    rows_source: str = Query("auto", description="existing (fail if none) | detect | auto"),
):
    """
    Row-by-row Tesseract OCR with digit/dash whitelist.
    Uses the same row-detection and SSE streaming as the LLM line-by-line endpoint.
    Each row is read with --psm 7 and tessedit_char_whitelist=0123456789-
    Results are saved to shape['tesseract_output'] and streamed as SSE.
    """
    import asyncio, concurrent.futures, shutil
    import json as _json
    import pytesseract
    from PIL import Image as PILImage, ImageOps

    if not shutil.which("tesseract"):
        pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

    d        = _resolve_folder(folder)
    jf       = d / f"{stem}.json"
    img_path = _find_image(d, stem)

    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    data_doc = json.loads(jf.read_text(encoding="utf-8"))
    shapes   = data_doc.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")

    shape = shapes[idx]
    pts   = shape["points"]
    xs    = [p[0] for p in pts]; ys = [p[1] for p in pts]
    x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)

    shadow   = _get_shadow_page(folder, stem, img_path)
    iw, ih   = shadow.size
    pad      = 4
    crop     = shadow.crop((
        max(0, int(x1) - pad), max(0, int(y1) - pad),
        min(iw, int(x2) + pad), min(ih, int(y2) + pad),
    ))

    crop_top = max(0, int(y1) - pad)
    rows = _rows_for_source(shape, crop, crop_top, cell_height, rows_source)

    tess_config = "--psm 7 -c tessedit_char_whitelist=0123456789-"

    def gen():
        yield _json.dumps({"type": "lines_detected", "count": len(rows),
                           "lines": [list(r) for r in rows]})

        line_texts: list[str] = []
        conf_values: list[float] = []

        for i, (top, bottom) in enumerate(rows):
            row_pad = max(4, cell_height // 6)
            rt = max(0, top - row_pad)
            rb = min(crop.height, bottom + row_pad)
            row_img = crop.crop((0, rt, crop.width, rb))

            # Upscale small rows
            if row_img.height < 48:
                scale   = 48 / row_img.height
                row_img = row_img.resize(
                    (int(row_img.width * scale), 48), PILImage.LANCZOS
                )

            grey = ImageOps.autocontrast(row_img.convert("L"))

            try:
                tess = pytesseract.image_to_data(
                    grey, lang=lang, config=tess_config,
                    output_type=pytesseract.Output.DICT,
                )
                words = [
                    (txt, int(conf))
                    for txt, conf in zip(tess["text"], tess["conf"])
                    if txt.strip() and int(conf) > 0
                ]
                text = " ".join(t for t, _ in words).strip() or "-"
                conf = round(sum(c for _, c in words) / len(words), 1) if words else 0.0
            except Exception as exc:
                text = f"[err: {exc}]"
                conf = 0.0

            line_texts.append(text)
            conf_values.append(conf)
            yield _json.dumps({"type": "row_result", "row": i, "text": text,
                               "top": top, "bottom": bottom})

        combined  = "\n".join(line_texts)
        mean_conf = round(sum(conf_values) / len(conf_values), 1) if conf_values else 0.0

        _apply_layer_rows(shape, [(t + crop_top, b + crop_top) for t, b in rows],
                          "ocr", line_texts, "linebyline")
        shape["tesseract_output"] = {
            "ocr_text":  combined,
            "mean_conf": mean_conf,
            "lang":      lang,
            "mode":      "linebyline",
        }
        _write_json(jf, data_doc)

        yield _json.dumps({"type": "done", "ocr_text": combined, "mean_conf": mean_conf})

    async def event_gen():
        loop = asyncio.get_event_loop()
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            it = iter(gen())
            while True:
                item = await loop.run_in_executor(pool, _safe_next, it)
                if item is _SSE_DONE:
                    break
                yield f"data: {item}\n\n"

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/page/shape/ocr/easyocr/linebyline")
async def api_ocr_easyocr_linebyline(
    folder:      str = Query(...),
    stem:        str = Query(...),
    idx:         int = Query(...),
    cell_height: int = Query(26),
    langs:       str = Query("en,hu"),
    rows_source: str = Query("auto", description="existing (fail if none) | detect | auto"),
):
    """Row-by-row EasyOCR using the comb-filter row detector. Streams SSE."""
    import asyncio, concurrent.futures
    import json as _json
    import numpy as np
    from PIL import Image as PILImage, ImageOps

    d        = _resolve_folder(folder)
    jf       = d / f"{stem}.json"
    img_path = _find_image(d, stem)

    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    data_doc = json.loads(jf.read_text(encoding="utf-8"))
    shapes   = data_doc.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")

    shape = shapes[idx]
    pts   = shape["points"]
    xs    = [p[0] for p in pts]; ys = [p[1] for p in pts]
    x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)

    shadow   = _get_shadow_page(folder, stem, img_path)
    iw, ih   = shadow.size
    pad      = 4
    crop     = shadow.crop((
        max(0, int(x1) - pad), max(0, int(y1) - pad),
        min(iw, int(x2) + pad), min(ih, int(y2) + pad),
    ))

    crop_top  = max(0, int(y1) - pad)
    rows      = _rows_for_source(shape, crop, crop_top, cell_height, rows_source)
    lang_list = [l.strip() for l in langs.split(",") if l.strip()]
    reader    = _get_easyocr_reader(lang_list)

    # crop is already fully preprocessed (threshold + line removal) by _get_shadow_page
    crop_bin = crop.convert("L")

    def gen():
        yield _json.dumps({"type": "lines_detected", "count": len(rows),
                           "lines": [list(r) for r in rows]})

        line_texts:  list[str]   = []
        conf_values: list[float] = []

        for i, (top, bottom) in enumerate(rows):
            row_pad = max(4, cell_height // 6)
            # Clamp padding to the midpoint between adjacent rows so we never
            # bleed content from the row above or below into this slice
            prev_bottom = rows[i - 1][1] if i > 0 else 0
            next_top    = rows[i + 1][0] if i < len(rows) - 1 else crop_bin.height
            rt = max(top    - row_pad, (prev_bottom + top)    // 2)
            rb = min(bottom + row_pad, (bottom      + next_top) // 2)
            rt = max(0, rt)
            rb = min(crop_bin.height, rb)
            row_img = crop_bin.crop((0, rt, crop_bin.width, rb))

            # Upscale with LANCZOS — soft edges match EasyOCR's training distribution
            # better than blocky NEAREST, which can make digit strokes look like
            # separate characters to the CRAFT detector
            target_h = 128
            scale    = target_h / row_img.height
            row_img  = row_img.resize(
                (int(row_img.width * scale), target_h), PILImage.LANCZOS
            ).convert("RGB")

            # Encode preprocessed image for the frontend preview panel
            import io as _io2, base64 as _b64
            _buf = _io2.BytesIO()
            row_img.save(_buf, format="PNG")
            img_b64 = _b64.b64encode(_buf.getvalue()).decode("ascii")

            try:
                results = reader.readtext(
                    np.array(row_img),
                    allowlist="0123456789-",
                    min_size=4,
                    text_threshold=0.4,
                    low_text=0.3,
                )
                results.sort(key=lambda r: min(pt[1] for pt in r[0]))
                texts = [t for _, t, _ in results]
                confs = [c for _, _, c in results]
                text  = " ".join(texts).strip()
                conf  = round(sum(confs) / len(confs) * 100, 1) if confs else 0.0
            except Exception as exc:
                text = ""
                conf = 0.0

            text = text or "-"

            line_texts.append(text)
            conf_values.append(conf)
            yield _json.dumps({"type": "row_result", "row": i, "text": text,
                               "img_b64": img_b64,
                               "top": top, "bottom": bottom})

        combined  = "\n".join(line_texts)
        mean_conf = round(sum(conf_values) / len(conf_values), 1) if conf_values else 0.0

        _apply_layer_rows(shape, [(t + crop_top, b + crop_top) for t, b in rows],
                          "ocr", line_texts, "linebyline")
        shape["tesseract_output"] = {
            "ocr_text":  combined,
            "mean_conf": mean_conf,
            "engine":    "easyocr",
            "mode":      "linebyline",
        }
        _write_json(jf, data_doc)

        yield _json.dumps({"type": "done", "ocr_text": combined, "mean_conf": mean_conf})

    async def event_gen():
        loop = asyncio.get_event_loop()
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            it = iter(gen())
            while True:
                item = await loop.run_in_executor(pool, _safe_next, it)
                if item is _SSE_DONE:
                    break
                yield f"data: {item}\n\n"

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
# EasyOCR — anchored (forced-N-row split guided by reference column)
# ---------------------------------------------------------------------------

def _split_into_n_rows(crop_image, n_rows: int) -> list[tuple[int, int]]:
    """
    Cut crop_image into exactly n_rows bands using histogram valley search.
    For each expected cut position (i * h // n_rows) we search a small window
    for the brightest (whitest = least ink = inter-row gap) pixel row and use
    that as the actual cut.  Returns a list of n_rows (top, bottom) tuples.
    """
    import numpy as np
    gray = np.array(crop_image.convert("L"))
    h = gray.shape[0]
    if n_rows <= 1 or h == 0:
        return [(0, h)]
    hist    = gray.mean(axis=1)          # per-row brightness (higher = whiter)
    window  = max(3, h // (n_rows * 3)) # search ±window around each nominal cut
    cuts    = [0]
    for i in range(1, n_rows):
        center = i * h // n_rows
        lo     = max(0, center - window)
        hi     = min(h - 1, center + window)
        best_y = lo + int(np.argmax(hist[lo : hi + 1]))
        cuts.append(best_y)
    cuts.append(h)
    return [(cuts[i], cuts[i + 1]) for i in range(n_rows)]


@app.post("/api/page/shape/ocr/easyocr/anchored")
async def api_ocr_easyocr_anchored(
    folder:  str = Query(...),
    stem:    str = Query(...),
    idx:     int = Query(...),
    n_rows:  int = Query(...),
    langs:   str = Query("en,hu"),
    ref_idx: int = Query(-1, description="Project this shape's row_struct bands instead of the histogram split"),
):
    """
    EasyOCR with a forced n_rows split derived from a reference column.
    Streams SSE in the same format as /easyocr/linebyline.
    """
    import asyncio, concurrent.futures
    import json as _json
    import numpy as np
    from PIL import Image as PILImage

    d        = _resolve_folder(folder)
    jf       = d / f"{stem}.json"
    img_path = _find_image(d, stem)

    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    data_doc = json.loads(jf.read_text(encoding="utf-8"))
    shapes   = data_doc.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")

    shape = shapes[idx]
    pts   = shape["points"]
    xs    = [p[0] for p in pts]; ys = [p[1] for p in pts]
    x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)

    shadow  = _get_shadow_page(folder, stem, img_path)
    iw, ih  = shadow.size
    pad     = 4
    crop    = shadow.crop((
        max(0, int(x1) - pad), max(0, int(y1) - pad),
        min(iw, int(x2) + pad), min(ih, int(y2) + pad),
    ))

    crop_top = max(0, int(y1) - pad)
    # Anchor at row structure: project the reference shape's bands directly
    projected = (_project_ref_bands(shapes[ref_idx], crop_top, crop.height, y1, y2)
                 if 0 <= ref_idx < len(shapes) else None)
    existing  = _existing_row_bands_rel(shape, crop_top, crop.height)
    # Anchoring is an explicit structural directive: keep the existing
    # structure only when its row count already matches the anchor
    rows      = projected or (existing if (existing and len(existing) == n_rows)
                              else _split_into_n_rows(crop, n_rows))
    lang_list = [l.strip() for l in langs.split(",") if l.strip()]
    reader    = _get_easyocr_reader(lang_list)
    crop_bin  = crop.convert("L")

    def gen():
        yield _json.dumps({"type": "lines_detected", "count": len(rows),
                           "lines": [list(r) for r in rows]})

        line_texts:  list[str]   = []
        conf_values: list[float] = []

        for i, (top, bottom) in enumerate(rows):
            # Small symmetric padding, clamped by neighbours
            row_pad     = max(2, (bottom - top) // 6)
            prev_bottom = rows[i - 1][1] if i > 0 else 0
            next_top    = rows[i + 1][0] if i < len(rows) - 1 else crop_bin.height
            rt = max(top    - row_pad, (prev_bottom + top)    // 2)
            rb = min(bottom + row_pad, (bottom      + next_top) // 2)
            rt = max(0, rt);  rb = min(crop_bin.height, rb)
            row_img = crop_bin.crop((0, rt, crop_bin.width, rb))

            target_h = 128
            scale    = target_h / max(1, row_img.height)
            row_img  = row_img.resize(
                (max(1, int(row_img.width * scale)), target_h), PILImage.LANCZOS
            ).convert("RGB")

            import io as _io2, base64 as _b64
            _buf = _io2.BytesIO()
            row_img.save(_buf, format="PNG")
            img_b64 = _b64.b64encode(_buf.getvalue()).decode("ascii")

            try:
                results = reader.readtext(
                    np.array(row_img),
                    allowlist="0123456789-",
                    min_size=4,
                    text_threshold=0.4,
                    low_text=0.3,
                )
                results.sort(key=lambda r: min(pt[1] for pt in r[0]))
                texts = [t for _, t, _ in results]
                confs = [c for _, _, c in results]
                text  = " ".join(texts).strip()
                conf  = round(sum(confs) / len(confs) * 100, 1) if confs else 0.0
            except Exception:
                text, conf = "", 0.0

            text = text or "-"
            line_texts.append(text)
            conf_values.append(conf)
            yield _json.dumps({"type": "row_result", "row": i, "text": text,
                               "img_b64": img_b64, "top": top, "bottom": bottom})

        combined  = "\n".join(line_texts)
        mean_conf = round(sum(conf_values) / len(conf_values), 1) if conf_values else 0.0
        _apply_layer_rows(shape, [(t + crop_top, b + crop_top) for t, b in rows],
                          "ocr", line_texts, "anchored",
                          force_boxes=projected is not None)
        shape["tesseract_output"] = {
            "ocr_text": combined, "mean_conf": mean_conf,
            "engine": "easyocr", "mode": "anchored",
        }
        _write_json(jf, data_doc)
        yield _json.dumps({"type": "done", "ocr_text": combined, "mean_conf": mean_conf})

    async def event_gen():
        loop = asyncio.get_event_loop()
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            it = iter(gen())
            while True:
                item = await loop.run_in_executor(pool, _safe_next, it)
                if item is _SSE_DONE:
                    break
                yield f"data: {item}\n\n"

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
# Routes — LLM cleaner
# ---------------------------------------------------------------------------

def _detect_text_rows(crop_image, cell_height: int = 28) -> list[tuple[int, int]]:
    """
    Comb-filter phase search + sequential gap snap.

    Pass 1 — comb filter finds the best global starting offset (phase).
    Pass 2 — each subsequent boundary is placed at prev + cell_height, then
    nudged ±snap_r pixels to the nearest local ink minimum (inter-row gap).
    Because each snap is relative to the previous boundary (not the ideal
    grid), corrections compound and any amount of cumulative drift is tracked,
    while the small snap_r prevents jumping over sparse rows (dashes/blanks).
    """
    import numpy as np

    gray       = np.array(crop_image.convert("L"))
    binary     = (gray < 180).astype(np.float32)
    projection = binary.sum(axis=1)

    roi_h = len(projection)
    if roi_h < cell_height:
        return [(0, roi_h)] if roi_h > 0 else []

    # Smooth projection — damps isolated ink blobs without washing out gaps
    k        = max(3, cell_height // 6)
    smoothed = np.convolve(projection, np.ones(k) / k, mode='same')

    cum = np.concatenate(([0.0], np.cumsum(projection)))

    def window_sum(t: int, b: int) -> float:
        t = max(0, t); b = min(roi_h, b)
        return float(cum[b] - cum[t]) if b > t else 0.0

    # Pass 1: find best phase via comb filter
    best_score, best_start = -1.0, 0
    for start in range(cell_height):
        score = 0.0; top = start
        while top + cell_height <= roi_h:
            score += window_sum(top, top + cell_height); top += cell_height
        if score > best_score:
            best_score = score; best_start = start

    # Pass 2: sequential snap — small radius, relative to previous boundary
    snap_r   = max(2, cell_height // 6)   # ~4 px at default 26 px height
    snapped  = [best_start]               # first boundary: trust comb filter

    while True:
        expected = snapped[-1] + cell_height
        if expected > roi_h:
            break
        lo  = max(0,        expected - snap_r)
        hi  = min(roi_h - 1, expected + snap_r)
        new = (lo + int(np.argmin(smoothed[lo: hi + 1]))) if lo < hi else expected
        new = max(new, snapped[-1] + 1)   # strict monotonicity
        snapped.append(new)

    # Partial last row only if ≥ 60 % of cell_height remains
    if snapped and roi_h - snapped[-1] >= cell_height * 0.6:
        snapped.append(roi_h)

    rows = [
        (snapped[i], min(snapped[i + 1], roi_h))
        for i in range(len(snapped) - 1)
        if snapped[i + 1] > snapped[i] and snapped[i] < roi_h
    ]
    return rows


class LlmRequest(BaseModel):
    prompt: str
    json_schema: Optional[dict] = None   # when set → structured-output (JSON) mode
    schema_name: Optional[str] = None    # which project schema was used (for provenance)


# Models served locally via ollama (OpenAI-compatible endpoint)
_LOCAL_MODELS: set[str] = {"qwen2.5vl:7b"}

# Azure model prefix — "azure:gpt-5-mini" routes to Azure OpenAI, strips prefix before the API call
_AZURE_PREFIX = "azure:"
# Second Azure resource ("digitization-US") — its own endpoint + key env vars.
# On Azure the bare name is the DEPLOYMENT name (e.g. "gpt-5.4-mini-batch").
_AZURE_US_PREFIX = "azure-us:"

# TK (institutional GPU server) prefix — "tk:vllm/..." routes to the OpenAI-compatible vllm endpoint
_TK_PREFIX = "tk:"
_TK_BASE_URL = "http://193.224.38.28:9000/v1"

def _bare_model(model: str) -> str:
    """Strip any routing prefix, returning the raw model name for the API call."""
    return (model.removeprefix(_AZURE_US_PREFIX)
                 .removeprefix(_AZURE_PREFIX)
                 .removeprefix(_TK_PREFIX))

# Always appended to every LLM prompt to suppress hallucinations on empty/dash cells
_EMPTY_CELL_GUARD = (
    "\nIf the cell is empty, contains only a dash, or the image shows no readable content, "
    "return exactly -."
)

def _make_llm_client(model: str):
    """Return an OpenAI-compatible client for the given model.
    Local models (ollama) → ollama server.
    azure:-prefixed models → Azure OpenAI (AZURE_OPENAI_ENDPOINT + AZURE_OPENAI_API_KEY).
    All others → OpenAI (OPENAI_API_KEY)."""
    import os
    from openai import OpenAI
    if model in _LOCAL_MODELS:
        host = os.environ.get("OLLAMA_HOST", "http://gpu.koren.work:11434")
        return OpenAI(api_key="ollama", base_url=f"{host}/v1")
    if model.startswith(_AZURE_US_PREFIX) or model.startswith(_AZURE_PREFIX):
        suffix = "_US" if model.startswith(_AZURE_US_PREFIX) else ""
        endpoint = os.environ.get(f"AZURE_OPENAI_ENDPOINT{suffix}")
        api_key  = os.environ.get(f"AZURE_OPENAI_API_KEY{suffix}")
        if not endpoint or not api_key:
            raise HTTPException(status_code=500,
                detail=f"AZURE_OPENAI_ENDPOINT{suffix} and AZURE_OPENAI_API_KEY{suffix} "
                       f"must be set for {model.split(':', 1)[0]}: models")
        # Normalize to scheme://host only — the portal's "Target URI" includes
        # a full path + api-version query (…/openai/deployments/…/chat/
        # completions?api-version=…); appending /openai/v1 to THAT yields a
        # 404. Keep just the resource host so files/batches resolve.
        from urllib.parse import urlparse
        u = urlparse(endpoint if "://" in endpoint else "https://" + endpoint)
        base = f"{u.scheme or 'https'}://{u.netloc}/openai/v1"
        return OpenAI(api_key=api_key, base_url=base + "/")
    if model.startswith(_TK_PREFIX):
        return OpenAI(api_key="none", base_url=_TK_BASE_URL)
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="OPENAI_API_KEY environment variable is not set")
    return OpenAI(api_key=api_key)


def _is_reasoning_model(model: str) -> bool:
    """GPT-5 family and the o-series are reasoning models: via chat completions
    they reject `max_tokens` (need `max_completion_tokens`), reject any
    temperature other than the default 1, and burn tokens on hidden reasoning."""
    m = _bare_model(model or "").lower()
    return m.startswith(("gpt-5", "o1", "o1-", "o3", "o3-", "o4", "o4-"))


# ── LLM response cache ───────────────────────────────────────────────────────
# Answers already paid for are reused: keyed on the full request (model,
# messages incl. image bytes, token budget, temperature, response_format).
# Opt-in per call (use_cache=True) — bulk/batch endpoints use it so re-runs
# are instant and free; interactive "give me a fresh attempt" paths (rule fix,
# LLM test modal) don't.
_LLM_CACHE_PATH = Path(os.environ.get("ECONAI_LLM_CACHE")
                       or Path(__file__).parent.parent / ".llm_cache.sqlite")
_LLM_CACHE_LOCK = threading.Lock()


def _llm_cache_conn():
    import sqlite3
    conn = sqlite3.connect(str(_LLM_CACHE_PATH))
    conn.execute("CREATE TABLE IF NOT EXISTS cache ("
                 "key TEXT PRIMARY KEY, content TEXT, model TEXT,"
                 "tokens_in INT, tokens_out INT, ts TEXT)")
    return conn


def _llm_cache_key(model, messages, max_out, temperature, response_format):
    import hashlib
    blob = json.dumps([_bare_model(model), messages, max_out, temperature,
                       response_format], sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def _llm_cache_get(key):
    from types import SimpleNamespace
    try:
        with _LLM_CACHE_LOCK, _llm_cache_conn() as conn:
            row = conn.execute("SELECT content, tokens_in, tokens_out "
                               "FROM cache WHERE key=?", (key,)).fetchone()
    except Exception:
        return None
    if row is None:
        return None
    return SimpleNamespace(
        cached=True,
        choices=[SimpleNamespace(message=SimpleNamespace(content=row[0]))],
        usage=SimpleNamespace(prompt_tokens=row[1] or 0,
                              completion_tokens=row[2] or 0))


def _llm_cache_put(key, model, content, tokens_in, tokens_out):
    import datetime as _dt
    try:
        with _LLM_CACHE_LOCK, _llm_cache_conn() as conn:
            conn.execute("INSERT OR REPLACE INTO cache VALUES (?,?,?,?,?,?)",
                         (key, content, _bare_model(model), tokens_in, tokens_out,
                          _dt.datetime.utcnow().isoformat() + "Z"))
    except Exception as e:
        print(f"[llm-cache] write failed: {e}", flush=True)


def _llm_complete_raw(client, model, messages, max_out, temperature=0, response_format=None):
    model = _bare_model(model)
    rf = {"response_format": response_format} if response_format else {}
    if _is_reasoning_model(model):
        kwargs = dict(model=model, messages=messages,
                      max_completion_tokens=max(max_out, 2000), **rf)
        try:
            return client.chat.completions.create(reasoning_effort="low", **kwargs)
        except Exception:
            # SDK / model that doesn't accept reasoning_effort — retry without it
            return client.chat.completions.create(**kwargs)
    return client.chat.completions.create(
        model=model, messages=messages, max_tokens=max_out,
        temperature=temperature, **rf)


def _llm_complete(client, model, messages, max_out, temperature=0,
                  response_format=None, use_cache=False):
    """Model-family-aware chat completion so reasoning models work alongside
    the classic chat models with one call signature.  Reasoning models get a
    token floor (reasoning tokens count against the budget) and low effort to
    stay fast/cheap on these simple digit-transcription tasks.
    `response_format` (e.g. a json_schema spec) is passed through when given.
    With use_cache=True an identical past request is answered from the local
    cache (free, instant); the response then has .cached == True."""
    key = None
    if use_cache:
        key = _llm_cache_key(model, messages, max_out, temperature, response_format)
        hit = _llm_cache_get(key)
        if hit is not None:
            return hit
    resp = _llm_complete_raw(client, model, messages, max_out, temperature, response_format)
    if key is not None:
        content = None
        try:
            content = resp.choices[0].message.content
        except Exception:
            pass
        if content:
            u = getattr(resp, "usage", None)
            _llm_cache_put(key, model, content,
                           getattr(u, "prompt_tokens", 0) if u else 0,
                           getattr(u, "completion_tokens", 0) if u else 0)
    return resp


def _json_schema_response_format(schema: dict, name: str = "record", strict: bool = True):
    """Wrap a JSON Schema for OpenAI Structured Outputs."""
    return {"type": "json_schema",
            "json_schema": {"name": (name or "record")[:60] or "record",
                            "strict": strict, "schema": schema}}


def _llm_complete_json(client, model, messages, max_out, schema, name, use_cache=False):
    """Structured-output completion: try strict json_schema, then non-strict,
    then plain json_object (for models/schemas that reject strict mode)."""
    try:
        return _llm_complete(client, model, messages, max_out, temperature=0,
                             response_format=_json_schema_response_format(schema, name, True),
                             use_cache=use_cache)
    except Exception:
        try:
            return _llm_complete(client, model, messages, max_out, temperature=0,
                                 response_format=_json_schema_response_format(schema, name, False),
                                 use_cache=use_cache)
        except Exception:
            return _llm_complete(client, model, messages, max_out, temperature=0,
                                 response_format={"type": "json_object"},
                                 use_cache=use_cache)


# ── Merge-safe shape writes ──────────────────────────────────────────────────
# Parallel LLM runs on the SAME page each hold their own copy of the page
# JSON; writing the whole document back would clobber sibling results. This
# re-reads the file and writes only the given fields of one shape.
_SHAPE_MERGE_LOCK = threading.Lock()


def _merge_shape_fields(jf, idx: int, fields: dict) -> bool:
    with _SHAPE_MERGE_LOCK:
        try:
            data = json.loads(Path(jf).read_text(encoding="utf-8"))
        except Exception:
            return False
        shapes = data.get("shapes", [])
        if idx < 0 or idx >= len(shapes):
            return False
        shapes[idx].update(fields)
        _write_json(jf, data)
        return True


# ---------------------------------------------------------------------------
# Overnight LLM batches (OpenAI / Azure Batch API — 50% price, ≤24h)
# A separate lane from the live batch: requests are packaged into one file,
# submitted to the provider's batch service, and applied to the pages later.
# Job manifests live per project in intermediate/llm_batch_jobs.json.
# ---------------------------------------------------------------------------

_LLM_JOBS_LOCK = threading.Lock()


def _llm_jobs_path(folder: str) -> Path:
    p = _resolve_folder(folder).parent / "intermediate"
    p.mkdir(exist_ok=True)
    return p / "llm_batch_jobs.json"


def _llm_jobs_load(folder: str) -> list:
    p = _llm_jobs_path(folder)
    if not p.exists():
        return []
    try:
        return json.loads(p.read_text(encoding="utf-8")).get("jobs", [])
    except Exception:
        return []


def _llm_jobs_save(folder: str, jobs: list):
    _write_json(_llm_jobs_path(folder), {"jobs": jobs})


def _llm_batch_supported(model: str) -> bool:
    m = (model or "").lower()
    return not (m.startswith(_TK_PREFIX) or _bare_model(model) in _LOCAL_MODELS)


def _llm_batch_line(custom_id: str, model: str, messages, max_out: int,
                    temperature: float, response_format=None) -> dict:
    """One JSONL request mirroring _llm_complete_raw's model-family handling."""
    body = {"model": _bare_model(model), "messages": messages}
    if response_format:
        body["response_format"] = response_format
    if _is_reasoning_model(model):
        body["max_completion_tokens"] = max(max_out, 2000)
        body["reasoning_effort"] = "low"
    else:
        body["max_tokens"] = max_out
        body["temperature"] = temperature
    return {"custom_id": custom_id, "method": "POST",
            "url": "/v1/chat/completions", "body": body}


def _img_part(crop) -> dict:
    import base64, io as _io
    buf = _io.BytesIO()
    crop.save(buf, format="JPEG", quality=92)
    b64 = base64.b64encode(buf.getvalue()).decode()
    return {"type": "image_url",
            "image_url": {"url": f"data:image/jpeg;base64,{b64}", "detail": "high"}}


class LlmBatchSubmit(BaseModel):
    targets:     List[dict]                 # [{stem, idx}]
    model:       str
    mode:        str = "image"              # image | image+ocr | ocr | linebyline
    prompt:      str
    cell_height: int = 26
    use_shadow:  bool = False
    json_schema: Optional[dict] = None
    schema_name: Optional[str] = None
    payload:     str = "image"              # per-row content: image | ocr | image+ocr
    rows_source: str = "auto"               # existing | detect | auto


# Providers cap the batch input file (Azure: 200 MB). Flush a chunk into its
# own job when it nears the limit so a big submission becomes several jobs.
_BATCH_CHUNK_BYTES = 180 * 1024 * 1024
_BATCH_CHUNK_MAX_REQ = 40000


def _gen_batch_requests(d, folder, body, json_mode, rf):
    """Yield (jsonl_line_dict, cell_key, meta_val_or_None) per request.
    meta_val (row bands) is attached only to the FIRST line of a linebyline
    cell; other lines carry None."""
    from collections import defaultdict
    from PIL import Image as PILImage
    by_stem = defaultdict(list)
    for t in body.targets:
        by_stem[t["stem"]].append(int(t["idx"]))
    for stem, idxs in by_stem.items():
        jf = d / f"{stem}.json"
        img_path = _find_image(d, stem)
        if not jf.exists():
            continue
        shapes = json.loads(jf.read_text(encoding="utf-8")).get("shapes", [])
        img = None
        if body.mode in ("image", "image+ocr", "linebyline") and img_path is not None:
            img = (_get_shadow_page(folder, stem, img_path).convert("RGB")
                   if body.use_shadow else PILImage.open(str(img_path)).convert("RGB"))
        for idx in idxs:
            if idx < 0 or idx >= len(shapes):
                continue
            shape = shapes[idx]
            pts = shape.get("points") or []
            if not pts:
                continue
            xs = [p[0] for p in pts]; ys = [p[1] for p in pts]
            x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
            pad = 4
            ckey = f"{stem}|{idx}"

            if body.mode == "linebyline":
                if img is None:
                    continue
                crop_top = max(0, int(y1) - pad)
                crop = img.crop((max(0, int(x1) - pad), crop_top,
                                 min(img.width, int(x2) + pad), min(img.height, int(y2) + pad)))
                existing = _existing_row_bands_rel(shape, crop_top, crop.height)
                if body.rows_source == "existing":
                    rows = existing
                elif body.rows_source == "detect":
                    rows = _detect_text_rows(crop, body.cell_height)
                else:
                    rows = existing or _detect_text_rows(crop, body.cell_height)
                if not rows:
                    continue
                row_ocr = ([(r.get("ocr") or "") for r in
                            (shape.get("row_struct") or {}).get("rows") or []]
                           if (existing and rows is existing) else [])
                mval = {"stem": stem, "idx": idx,
                        "bands": [[t0 + crop_top, b0 + crop_top] for t0, b0 in rows]}
                row_pad = max(4, body.cell_height // 6)
                for i, (top, bottom) in enumerate(rows):
                    txt = body.prompt
                    if body.payload in ("ocr", "image+ocr"):
                        txt = f"{body.prompt}\n\nOCR text:\n{row_ocr[i] if i < len(row_ocr) else ''}"
                    content = [{"type": "text", "text": txt}]
                    if body.payload in ("image", "image+ocr"):
                        rt = max(0, top - row_pad)
                        rb = min(crop.height, bottom + row_pad)
                        row_img = crop.crop((0, rt, crop.width, rb))
                        if row_img.height < 48:
                            scale = 48 / row_img.height
                            row_img = row_img.resize((int(row_img.width * scale), 48),
                                                     PILImage.LANCZOS)
                        content.append(_img_part(row_img))
                    messages = ([{"role": "user", "content": txt}]
                                if body.payload == "ocr"
                                else [{"role": "user", "content": content}])
                    yield (_llm_batch_line(f"{stem}|{idx}|{i}", body.model, messages, 64, 0),
                           ckey, mval if i == 0 else None)
                continue

            # whole-cell modes (mirrors api_llm_cell)
            content = []
            if body.mode in ("image", "image+ocr"):
                if img is None:
                    continue
                crop = img.crop((max(0, int(x1) - pad), max(0, int(y1) - pad),
                                 min(img.width, int(x2) + pad), min(img.height, int(y2) + pad)))
                content.append(_img_part(crop))
            prompt_text = body.prompt
            ocr_text = (shape.get("tesseract_output") or {}).get("ocr_text", "")
            if body.mode == "image+ocr" and ocr_text:
                prompt_text = f"{body.prompt}\n\nOCR text:\n{ocr_text}"
            elif body.mode == "ocr":
                prompt_text = f"{body.prompt}\n\n{ocr_text}"
            content.append({"type": "text", "text": prompt_text})
            messages = ([{"role": "user", "content": prompt_text}] if body.mode == "ocr"
                        else [{"role": "user", "content": content}])
            yield (_llm_batch_line(f"{stem}|{idx}|-1", body.model, messages,
                                   2048 if json_mode else 1024,
                                   0 if json_mode else 0.2, rf),
                   ckey, None)


@app.post("/api/llm_batch/submit")
def api_llm_batch_submit(folder: str = Query(...), body: LlmBatchSubmit = ...):
    """Build the batch request file(s) and submit them. Streams SSE progress
    (phase / requests built / running MB / per-chunk upload) and auto-splits
    into multiple provider jobs when the file nears the 200 MB limit."""
    import asyncio, concurrent.futures, datetime as _dt
    import json as _json

    if not _llm_batch_supported(body.model):
        raise HTTPException(status_code=400,
            detail="Overnight batches need an OpenAI-hosted (or Azure) model — "
                   "TK/local models have no batch service.")
    if not body.targets:
        raise HTTPException(status_code=400, detail="No target cells")
    d = _resolve_folder(folder)
    json_mode = bool(body.json_schema)
    rf = (_json_schema_response_format(body.json_schema, body.schema_name or "record", True)
          if json_mode else None)
    client = _make_llm_client(body.model)   # fail early if unconfigured
    provider = ("azure-us" if body.model.startswith(_AZURE_US_PREFIX)
                else "azure" if body.model.startswith(_AZURE_PREFIX) else "openai")

    def gen():
        chunk, chunk_bytes, chunk_meta, chunk_cells = [], 0, {}, set()
        made, req_total, cell_total = [], 0, 0

        def flush():
            nonlocal chunk, chunk_bytes, chunk_meta, chunk_cells
            if not chunk:
                return None
            jsonl = "\n".join(_json.dumps(l, ensure_ascii=False) for l in chunk)
            mb = len(jsonl.encode("utf-8")) / 1048576
            fobj = client.files.create(file=("econai_batch.jsonl", jsonl.encode("utf-8")),
                                       purpose="batch")
            remote = client.batches.create(input_file_id=fobj.id,
                                            endpoint="/v1/chat/completions",
                                            completion_window="24h")
            with _LLM_JOBS_LOCK:
                jobs = _llm_jobs_load(folder)
                job = {
                    "id": f"job-{len(jobs) + 1}", "remote_id": remote.id,
                    "provider": provider, "model": body.model, "mode": body.mode,
                    "prompt": body.prompt, "payload": body.payload,
                    "rows_source": body.rows_source, "json": json_mode,
                    "schema_name": body.schema_name,
                    "n_requests": len(chunk), "n_cells": len(chunk_cells),
                    "submitted": _dt.datetime.utcnow().isoformat() + "Z",
                    "status": getattr(remote, "status", "validating"),
                    "meta": chunk_meta,
                }
                jobs.append(job); _llm_jobs_save(folder, jobs)
            info = {"id": job["id"], "requests": len(chunk),
                    "cells": len(chunk_cells), "mb": round(mb, 1)}
            made.append(info)
            chunk, chunk_bytes, chunk_meta, chunk_cells = [], 0, {}, set()
            return info

        try:
            for line, ckey, mval in _gen_batch_requests(d, folder, body, json_mode, rf):
                b = len(_json.dumps(line, ensure_ascii=False).encode("utf-8")) + 1
                if chunk and (chunk_bytes + b > _BATCH_CHUNK_BYTES
                              or len(chunk) >= _BATCH_CHUNK_MAX_REQ):
                    yield _json.dumps({"type": "uploading",
                                       "mb": round(chunk_bytes / 1048576, 1),
                                       "requests": len(chunk)})
                    info = flush()
                    yield _json.dumps({"type": "chunk_done", **info})
                chunk.append(line); chunk_bytes += b
                chunk_cells.add(ckey)
                if mval is not None:
                    chunk_meta[ckey] = mval
                req_total += 1
                if req_total % 400 == 0:
                    yield _json.dumps({"type": "building", "requests": req_total,
                                       "cells": cell_total + len(chunk_cells),
                                       "mb": round((chunk_bytes) / 1048576, 1)})
            if chunk:
                yield _json.dumps({"type": "uploading",
                                   "mb": round(chunk_bytes / 1048576, 1),
                                   "requests": len(chunk)})
                info = flush()
                yield _json.dumps({"type": "chunk_done", **info})
            if not made:
                yield _json.dumps({"type": "error", "error": "No usable requests could be built"})
                return
            yield _json.dumps({"type": "done", "jobs": made,
                               "requests": sum(m["requests"] for m in made),
                               "cells": sum(m["cells"] for m in made)})
        except Exception as exc:
            yield _json.dumps({"type": "error", "error": f"Batch submit failed: {exc}"})

    async def event_gen():
        loop = asyncio.get_event_loop()
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            it = iter(gen())
            while True:
                item = await loop.run_in_executor(pool, _safe_next, it)
                if item is _SSE_DONE:
                    break
                yield f"data: {item}\n\n"

    return StreamingResponse(event_gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


def _llm_job_public(j: dict) -> dict:
    return {k: v for k, v in j.items() if k not in ("meta", "prompt")}


@app.get("/api/llm_batch/jobs")
def api_llm_batch_jobs(folder: str = Query(...), refresh: bool = Query(True)):
    """List this project's overnight jobs; live statuses re-checked with the provider."""
    with _LLM_JOBS_LOCK:
        jobs = _llm_jobs_load(folder)
        changed = False
        for j in jobs:
            if not refresh:
                continue
            terminal = j.get("status") in ("completed", "failed", "cancelled",
                                           "expired", "applied")
            # failed jobs get ONE more lookup to fetch the failure reason
            if terminal and not (j.get("status") == "failed"
                                 and "status_note" not in j):
                continue
            try:
                client = _make_llm_client(j["model"])
                remote = client.batches.retrieve(j["remote_id"])
                j["status"] = remote.status
                rc = getattr(remote, "request_counts", None)
                if rc:
                    j["counts"] = {"completed": getattr(rc, "completed", 0),
                                   "failed": getattr(rc, "failed", 0),
                                   "total": getattr(rc, "total", 0)}
                # surface the provider's failure reason (e.g. missing Global
                # Batch deployment) instead of a bare 'failed'
                errs = getattr(remote, "errors", None)
                data = getattr(errs, "data", None) if errs else None
                if data:
                    j["status_note"] = "; ".join(
                        f"{getattr(e, 'code', '')}: {getattr(e, 'message', '')}".strip(": ")
                        for e in data[:3])
                elif remote.status == "failed" and "status_note" not in j:
                    j["status_note"] = "no further detail from the provider"
                changed = True
            except Exception as e:
                j["status_note"] = f"status check failed: {e}"
        if changed:
            _llm_jobs_save(folder, jobs)
    return {"jobs": [_llm_job_public(j) for j in jobs]}


@app.post("/api/llm_batch/cancel")
def api_llm_batch_cancel(folder: str = Query(...), job: str = Query(...)):
    with _LLM_JOBS_LOCK:
        jobs = _llm_jobs_load(folder)
        j = next((x for x in jobs if x["id"] == job), None)
        if j is None:
            raise HTTPException(status_code=404, detail=f"Job not found: {job}")
        try:
            client = _make_llm_client(j["model"])
            remote = client.batches.cancel(j["remote_id"])
            j["status"] = getattr(remote, "status", "cancelling")
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"Cancel failed: {exc}")
        _llm_jobs_save(folder, jobs)
    return {"ok": True, "status": j["status"]}


@app.post("/api/llm_batch/remove")
def api_llm_batch_remove(folder: str = Query(...), job: Optional[str] = Query(None),
                         finished: bool = Query(False)):
    """Forget job(s) from the local list. Does NOT cancel/delete anything on the
    provider — just prunes the manifest. `finished=1` drops all terminal jobs;
    otherwise `job` removes one (any status)."""
    _TERMINAL = {"applied", "failed", "cancelled", "expired", "completed"}
    with _LLM_JOBS_LOCK:
        jobs = _llm_jobs_load(folder)
        before = len(jobs)
        if finished:
            jobs = [j for j in jobs if j.get("status") not in _TERMINAL]
        elif job:
            jobs = [j for j in jobs if j.get("id") != job]
        else:
            raise HTTPException(status_code=400, detail="Pass job=<id> or finished=1")
        _llm_jobs_save(folder, jobs)
    return {"ok": True, "removed": before - len(jobs)}


@app.post("/api/llm_batch/apply")
def api_llm_batch_apply(folder: str = Query(...), job: str = Query(...)):
    """Download a completed job's answers and write them into the pages exactly
    like the live path (openai_output / row_struct / structured)."""
    import datetime as _dt
    from collections import defaultdict

    d = _resolve_folder(folder)
    with _LLM_JOBS_LOCK:
        jobs = _llm_jobs_load(folder)
    j = next((x for x in jobs if x["id"] == job), None)
    if j is None:
        raise HTTPException(status_code=404, detail=f"Job not found: {job}")

    try:
        client = _make_llm_client(j["model"])
        remote = client.batches.retrieve(j["remote_id"])
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Job lookup failed: {exc}")
    if remote.status != "completed":
        raise HTTPException(status_code=400, detail=f"Job is {remote.status}, not completed")
    if not getattr(remote, "output_file_id", None):
        raise HTTPException(status_code=502, detail="Completed job has no output file")

    raw = client.files.content(remote.output_file_id).text
    answers, failed = {}, []
    for line in raw.splitlines():
        if not line.strip():
            continue
        try:
            rec = json.loads(line)
        except Exception:
            continue
        cid = rec.get("custom_id", "")
        resp = rec.get("response") or {}
        if rec.get("error") or resp.get("status_code") != 200:
            failed.append(cid)
            continue
        try:
            answers[cid] = (resp["body"]["choices"][0]["message"]["content"] or "").strip()
        except Exception:
            failed.append(cid)

    timestamp = _dt.datetime.utcnow().isoformat() + "Z"
    meta = j.get("meta") or {}
    mode, json_mode = j.get("mode"), bool(j.get("json"))

    # group per page, write each page once under the merge lock
    per_stem = defaultdict(dict)          # stem -> idx -> {row -> text}
    for cid, text in answers.items():
        stem, idx, row = cid.rsplit("|", 2)
        per_stem[stem].setdefault(int(idx), {})[int(row)] = text

    applied, bad_json = 0, 0
    for stem, shapes_res in per_stem.items():
        jf = d / f"{stem}.json"
        if not jf.exists():
            continue
        with _SHAPE_MERGE_LOCK:
            data = json.loads(jf.read_text(encoding="utf-8"))
            shapes = data.get("shapes", [])
            for idx, rows_res in shapes_res.items():
                if idx < 0 or idx >= len(shapes):
                    continue
                shape = shapes[idx]
                if mode == "linebyline":
                    bands = (meta.get(f"{stem}|{idx}") or {}).get("bands") or []
                    texts = [rows_res.get(i, "") for i in range(len(bands))]
                    if not bands:
                        continue
                    _apply_layer_rows(shape, [tuple(b) for b in bands],
                                      "llm", texts, "linebyline")
                    shape["openai_output"] = {
                        "response": "\n".join(texts), "model": j["model"],
                        "mode": "linebyline", "timestamp": timestamp,
                        "lines_detected": len(bands), "batch_job": j["id"],
                    }
                elif json_mode:
                    try:
                        parsed = json.loads(rows_res.get(-1, ""))
                    except Exception:
                        bad_json += 1
                        continue
                    prev = shape.get("structured") or {}
                    shape["structured"] = {
                        "schema_name": j.get("schema_name"), "llm": parsed,
                        "data": prev.get("data") if prev.get("edited") else parsed,
                        "edited": bool(prev.get("edited")),
                        "model": j["model"], "ts": timestamp, "batch_job": j["id"],
                    }
                else:
                    result = rows_res.get(-1, "")
                    shape["openai_output"] = {
                        "response": result, "model": j["model"], "mode": mode,
                        "timestamp": timestamp, "batch_job": j["id"],
                    }
                    _distribute_flat_to_rows(shape, "llm", result)
                applied += 1
            _write_json(jf, data)

    with _LLM_JOBS_LOCK:
        jobs = _llm_jobs_load(folder)
        j2 = next((x for x in jobs if x["id"] == job), None)
        if j2 is not None:
            j2["status"] = "applied"
            j2["applied_ts"] = timestamp
            j2["applied_cells"] = applied
            j2["failed_requests"] = len(failed)
            _llm_jobs_save(folder, jobs)

    return {"ok": True, "applied_cells": applied, "failed_requests": len(failed),
            "bad_json": bad_json, "failed_ids": failed[:50]}


# ---------------------------------------------------------------------------
# Batch undo — one-generation snapshot of the page JSONs a batch is about to
# touch (a zip in the project's intermediate/), restorable with one click.
# ---------------------------------------------------------------------------

_BATCH_UNDO_LOCK = threading.Lock()


def _batch_undo_path(folder: str) -> Path:
    p = _resolve_folder(folder).parent / "intermediate"
    p.mkdir(exist_ok=True)
    return p / "batch_undo.zip"


class BatchSnapshotBody(BaseModel):
    stems: List[str]
    op:    str = ""


@app.post("/api/batch_snapshot")
def api_batch_snapshot(folder: str = Query(...), body: BatchSnapshotBody = ...):
    """Snapshot the given pages' JSONs before a batch runs (replaces the
    previous snapshot — one undo generation)."""
    import datetime as _dt
    import zipfile
    d = _resolve_folder(folder)
    stems = [s for s in body.stems if s][:2000]
    with _BATCH_UNDO_LOCK:
        zp = _batch_undo_path(folder)
        n = 0
        with zipfile.ZipFile(zp, "w", zipfile.ZIP_DEFLATED) as z:
            for stem in stems:
                jf = d / f"{stem}.json"
                if jf.exists():
                    z.writestr(f"{stem}.json", jf.read_text(encoding="utf-8"))
                    n += 1
            z.writestr("_meta.json", json.dumps({
                "ts": _dt.datetime.utcnow().isoformat() + "Z",
                "op": body.op, "pages": n}))
    return {"ok": True, "pages": n}


@app.get("/api/batch_snapshot")
def api_batch_snapshot_meta(folder: str = Query(...)):
    import zipfile
    zp = _batch_undo_path(folder)
    if not zp.exists():
        return {"snapshot": None}
    try:
        with zipfile.ZipFile(zp) as z:
            return {"snapshot": json.loads(z.read("_meta.json").decode("utf-8"))}
    except Exception:
        return {"snapshot": None}


@app.post("/api/batch_snapshot/restore")
def api_batch_snapshot_restore(folder: str = Query(...)):
    """Restore every page in the snapshot to its pre-batch state."""
    import zipfile
    d = _resolve_folder(folder)
    zp = _batch_undo_path(folder)
    if not zp.exists():
        raise HTTPException(status_code=404, detail="No batch snapshot to restore")
    restored = 0
    with _BATCH_UNDO_LOCK, zipfile.ZipFile(zp) as z:
        for name in z.namelist():
            if name == "_meta.json" or not name.endswith(".json"):
                continue
            try:
                data = json.loads(z.read(name).decode("utf-8"))
            except Exception:
                continue
            _write_json(d / name, data)      # atomic per page
            restored += 1
    return {"ok": True, "restored": restored}


# ---------------------------------------------------------------------------
# Structural-blank detection — a FREE, local ink scan (no API). Lattice grids
# generate cells that are blank BY DESIGN (e.g. a district-header row over
# settlement columns). OCR yields dashes/noise on them and LLM hallucinates +
# wastes money. This marks inkless cells/rows `blank:true` so batches skip
# them and export emits them as missing (not a dash, not a zero).
# Ink test ported verbatim from the client _classifyEmptyRowBands (Otsu on the
# shadow crop, capped at 180; trim 3px borders; count pixels below threshold).
# ---------------------------------------------------------------------------

_BLANK_BORDER_PAD = 4     # px trimmed on ALL FOUR sides — excludes residual
                          # table border lines the shadow didn't fully erase
                          # (vertical borders were the main false-"has ink")
_BLANK_MIN_INK    = 4     # ink pixels below Otsu → NOT blank. A dash ≈ 15-40,
                          # so 4 tolerates a few stray border/dust pixels while
                          # still keeping any real mark.


def _otsu_threshold_np(gray) -> int:
    import numpy as np
    hist = np.bincount(gray.astype(np.uint8).ravel(), minlength=256).astype(np.float64)
    n = gray.size
    total = float((np.arange(256) * hist).sum())
    sumB = wB = 0.0
    max_var, thr = 0.0, 128
    for i in range(256):
        wB += hist[i]
        if wB == 0:
            continue
        wF = n - wB
        if wF == 0:
            break
        sumB += i * hist[i]
        mB = sumB / wB
        mF = (total - sumB) / wF
        v = wB * wF * (mB - mF) ** 2
        if v > max_var:
            max_var, thr = v, i
    return thr


def _band_has_ink(gray, y0: int, y1: int, otsu: int) -> bool:
    p = _BLANK_BORDER_PAD
    y0 = max(0, y0 + p)
    y1 = min(gray.shape[0], y1 - p)
    x0 = min(p, gray.shape[1] // 4)              # trim left/right borders too
    x1 = max(x0 + 1, gray.shape[1] - p)
    if y1 <= y0:
        return True   # too thin to judge → treat as non-blank (safe: never auto-skip)
    region = gray[y0:y1, x0:x1]
    return int((region <= otsu).sum()) >= _BLANK_MIN_INK


class MarkBlanksBody(BaseModel):
    stems:      List[str] = []
    col_filter: Optional[str] = None


@app.post("/api/batch/mark_blanks")
def api_mark_blanks(folder: str = Query(...), body: MarkBlanksBody = ...):
    """Ink-scan lattice cells over the given pages; mark blank cells/rows so
    OCR/LLM batches skip them and export treats them as missing. Idempotent:
    re-running clears the flag where ink is now present."""
    import numpy as np
    from PIL import Image as PILImage

    d = _resolve_folder(folder)
    col_ranges = _parse_col_ranges(body.col_filter)

    def col_ok(sc):
        if col_ranges is None:
            return True
        return any(lo <= sc <= (hi if hi is not None else sc) for lo, hi in col_ranges)

    stems = [s for s in (body.stems or []) if s] or \
            [jf.stem for jf in sorted(d.glob("*.json"), key=lambda p: _page_sort_key(p.stem))]
    totals = dict(cells_blank=0, cells_inked=0, rows_blank=0, scanned=0)
    per_page = []

    for stem in stems:
        jf = d / f"{stem}.json"
        img_path = _find_image(d, stem)
        if not jf.exists() or img_path is None:
            continue
        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except Exception:
            continue
        try:
            page_img = _get_shadow_page(folder, stem, img_path).convert("L")
        except Exception:
            page_img = PILImage.open(str(img_path)).convert("L")
        page = np.asarray(page_img)
        iw = page.shape[1]; ih = page.shape[0]
        changed = False
        pc = dict(cells_blank=0, cells_inked=0, rows_blank=0, scanned=0)

        for sh in data.get("shapes", []):
            sr, sc = sh.get("super_row"), sh.get("super_column")
            if sr is None or sc is None or not col_ok(int(sc)):
                continue
            x1, y1, x2, y2 = _shape_bbox(sh)
            cx1, cy1 = max(0, int(x1)), max(0, int(y1))
            cx2, cy2 = min(iw, int(x2)), min(ih, int(y2))
            if cx2 <= cx1 + 2 or cy2 <= cy1 + 2:
                continue
            crop = page[cy1:cy2, cx1:cx2]
            otsu = min(_otsu_threshold_np(crop), 180)
            pc["scanned"] += 1

            rows = (sh.get("row_struct") or {}).get("rows") or []
            if rows:
                all_blank = True
                for r in rows:
                    ry0 = int(round(r["y0"] - cy1)); ry1 = int(round(r["y1"] - cy1))
                    blank = not _band_has_ink(crop, ry0, ry1, otsu)
                    if blank and not r.get("blank"):
                        r["blank"] = True; changed = True; pc["rows_blank"] += 1
                    elif not blank and r.get("blank"):
                        r.pop("blank", None); changed = True
                    all_blank = all_blank and blank
                if all_blank and not sh.get("blank"):
                    sh["blank"] = True; changed = True; pc["cells_blank"] += 1
                elif not all_blank and sh.get("blank"):
                    sh.pop("blank", None); changed = True; pc["cells_inked"] += 1
            else:
                blank = not _band_has_ink(crop, 0, crop.shape[0], otsu)
                if blank and not sh.get("blank"):
                    sh["blank"] = True; changed = True; pc["cells_blank"] += 1
                elif not blank and sh.get("blank"):
                    sh.pop("blank", None); changed = True; pc["cells_inked"] += 1

        if changed:
            _write_json(jf, data)
        for k in totals:
            totals[k] += pc[k]
        per_page.append({"stem": stem, **pc})

    return {"ok": True, "pages": len(per_page), "totals": totals, "per_page": per_page}


# ---------------------------------------------------------------------------
# Row-structure builder — the SINGLE authoritative structure step (free,
# local, no API). Detects internal rows per cell, OR projects one anchor
# column's structure across the lattice row. Row count comes from the image
# (auto) or from a chosen content layer's line count. Content extraction is
# then a separate OCR/LLM pass with Scope = "keep structure". This replaces
# the anchored-OCR / anchored-LLM batch ops (which conflated the two).
# ---------------------------------------------------------------------------

def _cell_layer_text(shape, source):
    human = (shape.get("human_output") or {}).get("human_corrected_text") or ""
    llm   = (shape.get("openai_output") or {}).get("response") or ""
    ocr   = ((shape.get("tesseract_output") or {}).get("ocr_text")
             or (shape.get("easyocr_output") or {}).get("ocr_text") or "")
    pdf   = shape.get("pdf_text") or ""
    return {"pdf": pdf, "ocr": ocr, "llm": llm, "human": human,
            "best": human or llm or ocr or pdf}.get(source, "")


def _bands_for_cell(shadow, shape, source, cell_height):
    """Return absolute-y (y0, y1) bands for one cell, or None if undeterminable."""
    # "existing" → use the cell's own hand-made row_struct bands verbatim
    if source == "existing":
        rows = (shape.get("row_struct") or {}).get("rows") or []
        return [(float(r["y0"]), float(r["y1"])) for r in rows] if rows else None
    x1, y1, x2, y2 = _shape_bbox(shape)
    iw, ih = shadow.size
    pad = 4
    cy1 = max(0, int(y1) - pad)
    crop = shadow.crop((max(0, int(x1) - pad), cy1,
                        min(iw, int(x2) + pad), min(ih, int(y2) + pad)))
    if source == "image":
        bands = _detect_text_rows(crop, cell_height)
    else:
        txt = _cell_layer_text(shape, source)
        if not txt.strip():
            return None
        bands = _split_into_n_rows(crop, len(_split_lines(txt)))
    return [(t + cy1, b + cy1) for t, b in bands] if bands else None


def _project_abs_bands(bands_abs, ry1, ry2, ty1, ty2):
    """Linear-map absolute bands from a reference bbox y-range to a target's."""
    rh = (ry2 - ry1) or 1.0
    th = ty2 - ty1
    return [(ty1 + (b0 - ry1) * th / rh, ty1 + (b1 - ry1) * th / rh)
            for b0, b1 in bands_abs]


def _set_row_struct(shape, bands_abs, origin):
    """Write row_struct from absolute bands; distribute each flat layer whose
    line count matches the band count (mirrors the Convert op)."""
    n = len(bands_abs)
    layers = {}
    for lay, txt in (("human", (shape.get("human_output") or {}).get("human_corrected_text") or ""),
                     ("llm",   (shape.get("openai_output") or {}).get("response") or ""),
                     ("ocr",   ((shape.get("tesseract_output") or {}).get("ocr_text")
                                or (shape.get("easyocr_output") or {}).get("ocr_text") or ""))):
        lines = _split_lines(txt) if txt.strip() else None
        layers[lay] = lines if (lines and len(lines) == n) else None
    rows = []
    for i, (b0, b1) in enumerate(bands_abs):
        row = {"n": i + 1, "y0": float(b0), "y1": float(b1),
               "ocr": "", "llm": "", "human": ""}
        for lay, lines in layers.items():
            if lines:
                row[lay] = lines[i]
        rows.append(row)
    shape["row_struct"] = {"version": 1, "origin": origin, "rows": rows}
    _sync_flat_from_rows(shape)


class RowsBuildBody(BaseModel):
    stems:          List[str] = []
    anchor_pattern: Optional[str] = None   # cyclic per page; slot = anchor super_column,
                                           # empty slot = skip page; whole thing blank = no anchoring
    col_filter:     Optional[str] = None
    source:         str = "image"          # image | pdf | ocr | llm | human | best
    cell_height:    int = 26
    overwrite:      bool = False


@app.post("/api/rows/build")
def api_rows_build(folder: str = Query(...), body: RowsBuildBody = ...):
    """Build internal-row structure across pages — the one place row geometry
    is decided. Free, local, no API."""
    from collections import defaultdict
    d = _resolve_folder(folder)
    stems = [s for s in (body.stems or []) if s] or \
            [jf.stem for jf in sorted(d.glob("*.json"), key=lambda p: _page_sort_key(p.stem))]
    col_ok = _auth_col_pred(body.col_filter)

    # anchor pattern: cyclic over stems; empty slot skips the page; whole-blank = no anchoring
    pat = [p.strip() for p in (body.anchor_pattern or "").split(",")]
    pat = pat if any(pat) else []
    anchoring = bool(pat)

    totals = dict(built=0, projected=0, skipped=0, no_rows=0, pages=0)
    per_page = []

    for pi, stem in enumerate(stems):
        jf = d / f"{stem}.json"
        img_path = _find_image(d, stem)
        if not jf.exists() or img_path is None:
            continue
        anchor_col = None
        if anchoring:
            slot = pat[pi % len(pat)]
            if not slot:
                continue                    # empty slot → skip this page
            try:
                anchor_col = int(slot)
            except ValueError:
                continue
        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except Exception:
            continue
        try:
            shadow = _get_shadow_page(folder, stem, img_path)
        except Exception:
            from PIL import Image as _PIL
            shadow = _PIL.open(str(img_path))
        shapes = data.get("shapes", [])
        pc = dict(built=0, projected=0, skipped=0, no_rows=0)
        changed = False

        def eligible(sh):
            sc = sh.get("super_column")
            if sc is None or sh.get("super_row") is None:
                return False
            if not col_ok(int(sc)):
                return False
            if sh.get("blank"):
                return False
            if sh.get("row_struct") and not body.overwrite:
                return False
            return True

        if not anchoring:
            for sh in shapes:
                if not eligible(sh):
                    if sh.get("super_column") is not None:
                        pc["skipped"] += 1
                    continue
                bands = _bands_for_cell(shadow, sh, body.source, body.cell_height)
                if not bands:
                    pc["no_rows"] += 1; continue
                _set_row_struct(sh, bands, "detected")
                pc["built"] += 1; changed = True
        else:
            # group lattice cells by (table, super_row); project the anchor col
            groups = defaultdict(dict)   # (table, sr) -> {sc: shape}
            for sh in shapes:
                sr, sc = sh.get("super_row"), sh.get("super_column")
                if sr is not None and sc is not None:
                    groups[(sh.get("table") or 0, int(sr))][int(sc)] = sh
            for (_tbl, _sr), cells in groups.items():
                anchor = cells.get(anchor_col)
                if anchor is None:
                    continue
                abands = _bands_for_cell(shadow, anchor, body.source, body.cell_height)
                if not abands:
                    pc["no_rows"] += 1; continue
                _, ay1, _, ay2 = _shape_bbox(anchor)
                for sc, sh in cells.items():
                    if not eligible(sh):
                        if not (sh.get("row_struct") and not body.overwrite):
                            pc["skipped"] += 1
                        continue
                    if sc == anchor_col:
                        _set_row_struct(sh, abands, "anchor")
                        pc["built"] += 1; changed = True
                    else:
                        _, ty1, _, ty2 = _shape_bbox(sh)
                        proj = _project_abs_bands(abands, ay1, ay2, ty1, ty2)
                        _set_row_struct(sh, proj, "projected")
                        pc["projected"] += 1; changed = True

        if changed:
            _write_json(jf, data)
        for k in ("built", "projected", "skipped", "no_rows"):
            totals[k] += pc[k]
        totals["pages"] += 1
        per_page.append({"stem": stem, **pc})

    return {"ok": True, "totals": totals, "per_page": per_page}


# ---------------------------------------------------------------------------
# P3 — page status scoreboard. Status lives in the page's flags.status
# (predicted | corrected | verified | problem); set via PATCH /api/page/flags.
# ---------------------------------------------------------------------------

@app.get("/api/project/status")
def api_project_status(folder: str = Query(...)):
    """Per-status page counts for a project's annotation folder — the dashboard
    progress board and the editor's 'next unreviewed page' both read this."""
    d = _resolve_folder(folder)
    counts = {"predicted": 0, "corrected": 0, "verified": 0, "problem": 0}
    pages = []
    for jf in sorted(d.glob("*.json"), key=lambda p: _page_sort_key(p.stem)):
        if jf.name.endswith(_RULES_FILENAME):
            continue
        try:
            flags = (json.loads(jf.read_text(encoding="utf-8")).get("flags") or {})
        except Exception:
            flags = {}
        st = flags.get("status") or "predicted"
        if st not in counts:
            st = "predicted"
        counts[st] += 1
        pages.append({"stem": jf.stem, "status": st,
                      "assignee": flags.get("assignee")})
    return {"counts": counts, "total": sum(counts.values()), "pages": pages}


# ---------------------------------------------------------------------------
# P1 — review queue. A project-wide, frequency/severity-ranked list of suspect
# UNITS (a cell, or one internal row) so review is "fix the flagged, ignore the
# rest". Signals are all computable from the page JSON (no API): OCR≠LLM
# disagreement, numeric column outliers, unresolved authority, unverified.
# ---------------------------------------------------------------------------

def _norm_txt(t: str) -> str:
    return re.sub(r"\s+", " ", (t or "")).strip().lower()


def _parse_num(t: str):
    """Parse a Hungarian-style number ('1.446' = 1446, '1 234' = 1234)."""
    s = re.sub(r"[.\s ]", "", (t or "").strip())
    if re.fullmatch(r"-?\d+", s):
        try:
            return int(s)
        except ValueError:
            return None
    return None


class ReviewQueueBody(BaseModel):
    stems:      List[str] = []
    pattern:    Optional[str] = None
    col_filter: Optional[str] = None
    signals:    List[str] = ["disagree", "outlier", "unverified"]
    layer:      str = "best_llm"
    limit:      int = 50000
    exclude_verified: bool = True     # skip pages already marked verified


@app.post("/api/review/queue")
def api_review_queue(folder: str = Query(...), body: ReviewQueueBody = ...):
    from collections import defaultdict
    d = _resolve_folder(folder)
    pages = _auth_select_pages(d, body.stems, body.pattern)
    col_allowed = _auth_col_pred(body.col_filter)
    sig = set(body.signals or [])
    items = []

    # severity weight per signal → ranks the queue (higher = worse, shown first)
    W = {"rule": 5, "disagree": 4, "outlier": 3, "unresolved": 2, "unverified": 1}

    def best_of(human, ocr, llm, pdf):
        h, o, l, p = (human or "").strip(), (ocr or "").strip(), \
                     (llm or "").strip(), (pdf or "").strip()
        if body.layer == "ocr":  return o or h or l
        if body.layer == "llm":  return l or h or o
        return h or l or o or p

    for stem in pages:
        jf = d / f"{stem}.json"
        if not jf.exists():
            continue
        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except Exception:
            continue
        if body.exclude_verified and (data.get("flags") or {}).get("status") == "verified":
            continue
        shapes = data.get("shapes", [])

        # gather numeric values per (table, col) for outlier detection
        colvals = defaultdict(list)     # (table, col) -> [(value, ref)]

        def consider(i, sh, row_n, human, ocr, llm, pdf, y0=None, y1=None):
            # A human value = resolved. Accepting/typing in the review strip
            # writes Human, so a reviewed unit must never come back — even if
            # OCR≠LLM still differ or it's a numeric outlier.
            if (human or "").strip():
                return
            o, l = _norm_txt(ocr), _norm_txt(llm)
            best = best_of(human, ocr, llm, pdf)
            reasons = []
            if "disagree" in sig and o and l and o != l:
                reasons.append("OCR≠LLM")
            if "unverified" in sig and best:
                reasons.append("unverified")
            unit = {"stem": stem, "idx": i, "row": row_n, "best": best,
                    "y0": y0, "y1": y1,
                    "col": sh.get("super_column"), "table": sh.get("table") or 0}
            if reasons:
                sev = max(W.get(_reason_key(r), 1) for r in reasons)
                items.append({**unit, "why": ", ".join(reasons), "sev": sev})
            # collect numbers for the outlier pass (unverified cells only)
            if "outlier" in sig and sh.get("super_column") is not None:
                v = _parse_num(best)
                if v is not None:
                    colvals[(sh.get("table") or 0, int(sh["super_column"]))].append(
                        (v, {**unit, "why": "numeric outlier"}))

        for i, sh in enumerate(shapes):
            sc = sh.get("super_column")
            if sc is not None and not col_allowed(int(sc)):
                continue
            # unresolved-authority: in a column that HAS an authority mapping
            rows = (sh.get("row_struct") or {}).get("rows") or []
            if rows:
                for r in rows:
                    if r.get("blank"):
                        continue      # structural blanks never enter the queue
                    consider(i, sh, r.get("n"), r.get("human"), r.get("ocr"),
                             r.get("llm"), r.get("pdf"), r.get("y0"), r.get("y1"))
            elif not sh.get("blank"):
                consider(i, sh, None,
                         (sh.get("human_output") or {}).get("human_corrected_text"),
                         ((sh.get("tesseract_output") or {}).get("ocr_text") or
                          (sh.get("easyocr_output") or {}).get("ocr_text")),
                         (sh.get("openai_output") or {}).get("response"),
                         sh.get("pdf_text"))

        # outliers: values > 3 MADs from the column median
        for (_t, _c), vals in colvals.items():
            if len(vals) < 5:
                continue
            nums = sorted(v for v, _ in vals)
            med = nums[len(nums) // 2]
            devs = sorted(abs(v - med) for v in nums)
            mad = devs[len(devs) // 2] or 1
            for v, ref in vals:
                if abs(v - med) > 6 * mad and abs(v - med) > 10:
                    items.append({**ref, "sev": W["outlier"],
                                  "extra": f"{v} vs median {med}"})

    # de-dup (a unit can trip several signals), keep the highest severity
    best_by_key = {}
    for it in items:
        k = (it["stem"], it["idx"], it["row"])
        if k not in best_by_key or it["sev"] > best_by_key[k]["sev"]:
            best_by_key[k] = it
    ranked = sorted(best_by_key.values(), key=lambda x: -x["sev"])[:body.limit]
    return {"queue": ranked, "total": len(best_by_key), "pages": len(pages)}


def _reason_key(reason: str) -> str:
    return {"OCR≠LLM": "disagree", "unverified": "unverified",
            "numeric outlier": "outlier"}.get(reason, "unverified")


class ReviewAcceptBody(BaseModel):
    stem:  str
    idx:   int
    row:   Optional[int] = None      # internal row number (1-based), None = whole cell
    value: str = ""
    restore_blank: Optional[bool] = None   # set → explicit write (undo path),
                                           # None → empty value marks blank


@app.post("/api/review/accept")
def api_review_accept(folder: str = Query(...), body: ReviewAcceptBody = ...):
    """The single source of truth for a review decision (desktop strip and the
    mobile page both use it). Writes the value to the Human layer; an EMPTY
    value marks the unit a structural blank so it never re-enters the queue.
    Returns the previous state so clients can implement undo (re-post it with
    restore_blank set)."""
    d  = _resolve_folder(folder)
    jf = d / f"{body.stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")
    with _SHAPE_MERGE_LOCK:
        data = json.loads(jf.read_text(encoding="utf-8"))
        shapes = data.get("shapes", [])
        if body.idx < 0 or body.idx >= len(shapes):
            raise HTTPException(status_code=400, detail="Shape index out of range")
        shape = shapes[body.idx]
        is_empty = not body.value.strip()
        make_blank = body.restore_blank if body.restore_blank is not None else is_empty

        if body.row is not None:
            rows = (shape.get("row_struct") or {}).get("rows") or []
            r = next((x for x in rows if x.get("n") == body.row), None)
            if r is None:
                raise HTTPException(status_code=400, detail=f"Row {body.row} not found")
            prev = {"human": r.get("human") or "", "blank": bool(r.get("blank"))}
            r["human"] = body.value
            if make_blank:
                r["blank"] = True
            else:
                r.pop("blank", None)
            _sync_flat_from_rows(shape)
        else:
            prev = {"human": (shape.get("human_output") or {}).get("human_corrected_text") or "",
                    "blank": bool(shape.get("blank"))}
            shape.setdefault("human_output", {})["human_corrected_text"] = body.value
            if make_blank:
                shape["blank"] = True
            else:
                shape.pop("blank", None)
            _distribute_flat_to_rows(shape, "human", body.value)
        _write_json(jf, data)
    return {"ok": True, "blank": bool(make_blank), "prev": prev}


@app.post("/api/page/shape/llm")
def api_llm_cell(
    folder:     str  = Query(...),
    stem:       str  = Query(...),
    idx:        int  = Query(...),
    model:      str  = Query("gpt-4o-mini"),
    mode:       str  = Query("image", description="image | image+ocr | ocr | linebyline"),
    use_shadow: bool = Query(False, description="Use OCR shadow (line-erased) image instead of original"),
    dry_run:    bool = Query(False, description="Return result without writing to JSON (for testing)"),
    cached:     bool = Query(True, description="Reuse an identical past answer from the local cache"),
    body:       LlmRequest = ...,
):
    """Send a cell to an LLM and store the result in the page JSON."""
    import os, base64, io, datetime
    from PIL import Image as PILImage

    d        = _resolve_folder(folder)
    jf       = d / f"{stem}.json"
    img_path = _find_image(d, stem)

    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None and mode in ("image", "image+ocr"):
        raise HTTPException(status_code=404, detail="Image not found")

    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")

    shape = shapes[idx]
    pts   = shape["points"]
    xs    = [p[0] for p in pts]; ys = [p[1] for p in pts]
    x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)

    # Build the message content
    content: list = []

    if mode in ("image", "image+ocr"):
        if use_shadow:
            img = _get_shadow_page(folder, stem, img_path).convert("RGB")
        else:
            img = PILImage.open(str(img_path)).convert("RGB")
        w, h = img.size
        pad  = 4
        crop = img.crop((
            max(0, int(x1) - pad), max(0, int(y1) - pad),
            min(w, int(x2) + pad), min(h, int(y2) + pad),
        ))
        buf = io.BytesIO()
        crop.save(buf, format="JPEG", quality=92)
        b64 = base64.b64encode(buf.getvalue()).decode()
        content.append({
            "type":      "image_url",
            "image_url": {"url": f"data:image/jpeg;base64,{b64}", "detail": "high"},
        })

    prompt_text = body.prompt
    if mode == "image+ocr":
        ocr_text = shape.get("tesseract_output", {}).get("ocr_text", "")
        if ocr_text:
            prompt_text = f"{body.prompt}\n\nOCR text:\n{ocr_text}"
    elif mode == "ocr":
        ocr_text = shape.get("tesseract_output", {}).get("ocr_text", "")
        prompt_text = f"{body.prompt}\n\n{ocr_text}"

    content.append({"type": "text", "text": prompt_text})

    # For text-only mode, skip the content list and use a plain string
    if mode == "ocr":
        messages = [{"role": "user", "content": prompt_text}]
    else:
        messages = [{"role": "user", "content": content}]

    print(f"[LLM] model={model} mode={mode} dry_run={dry_run} "
          f"prompt={prompt_text!r}", flush=True)

    json_mode = bool(body.json_schema)
    use_cache = cached and not dry_run
    try:
        client = _make_llm_client(model)
        if json_mode:
            response = _llm_complete_json(client, model, messages, 2048,
                                          body.json_schema, body.schema_name or "record",
                                          use_cache=use_cache)
        else:
            # temperature 0.2 for classic models — a tiny bit of freedom to
            # self-correct; reasoning models ignore temperature (see _llm_complete)
            response = _llm_complete(client, model, messages, 1024, temperature=0.2,
                                     use_cache=use_cache)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc))

    from_cache  = bool(getattr(response, "cached", False))
    raw_content = response.choices[0].message.content
    result      = (raw_content or "").strip()
    tokens_in   = response.usage.prompt_tokens     if response.usage else 0
    tokens_out  = response.usage.completion_tokens if response.usage else 0
    print(f"[LLM] result={result!r}  tokens={tokens_in}→{tokens_out}"
          + ("  (cache hit)" if from_cache else ""), flush=True)
    timestamp = datetime.datetime.utcnow().isoformat() + "Z"

    # ── JSON / structured mode: parse and store on shape["structured"] ───────
    if json_mode:
        try:
            parsed = json.loads(result)
        except Exception as exc:
            raise HTTPException(status_code=502,
                detail=f"LLM did not return valid JSON: {exc}. Raw: {result[:200]}")
        if not dry_run:
            prev = shape.get("structured") or {}
            shape["structured"] = {
                "schema_name": body.schema_name,
                "llm":   parsed,
                # keep an existing human-edited `data` only if the user had one;
                # otherwise seed it with the fresh LLM output
                "data":  prev.get("data") if prev.get("edited") else parsed,
                "edited": bool(prev.get("edited")),
                "model": model, "ts": timestamp,
            }
            # merge-write: parallel runs on the same page must not clobber
            _merge_shape_fields(jf, idx, {"structured": shape["structured"]})
        return {"structured": parsed, "schema_name": body.schema_name,
                "model": model, "mode": mode, "timestamp": timestamp,
                "tokens_in": tokens_in, "tokens_out": tokens_out,
                "cached": from_cache}

    if not dry_run:
        shape["openai_output"] = {
            "response":  result,
            "model":     model,
            "mode":      mode,
            "timestamp": timestamp,
        }
        _distribute_flat_to_rows(shape, "llm", result)
        fields = {"openai_output": shape["openai_output"]}
        if shape.get("row_struct"):
            fields["row_struct"] = shape["row_struct"]
        _merge_shape_fields(jf, idx, fields)

    return {"response": result, "model": model, "mode": mode,
            "timestamp": timestamp, "prompt_sent": prompt_text,
            "tokens_in": tokens_in, "tokens_out": tokens_out,
            "cached": from_cache}


@app.post("/api/page/shape/llm/linebyline")
async def api_llm_linebyline(
    folder:      str  = Query(...),
    stem:        str  = Query(...),
    idx:         int  = Query(...),
    model:       str  = Query("gpt-4o-mini"),
    cell_height: int  = Query(28, description="Expected height of one text row in pixels"),
    use_shadow:  bool = Query(False, description="Use OCR shadow (line-erased) image instead of original"),
    dry_run:     bool = Query(False, description="Return result without writing to JSON (for testing)"),
    cached:      bool = Query(True, description="Reuse identical past answers from the local cache"),
    payload:     str  = Query("image", description="What the model sees per row: image | ocr | image+ocr"),
    rows_source: str  = Query("auto", description="Row bands: existing (fail if none) | detect (always re-detect) | auto (existing else detect)"),
    body:        LlmRequest = ...,
):
    """
    Per-internal-row LLM: slice the cell into rows and send each row to the
    LLM individually, streaming results as SSE.  `payload` picks what the
    model sees (row image, that row's OCR text, or both); `rows_source` picks
    where the bands come from (the cell's stored structure or fresh pixel
    detection).

    SSE message types:
      {"type": "lines_detected", "count": N, "lines": [[top,bottom], ...]}
      {"type": "row_result",     "row": i,  "text": "...", "top": t, "bottom": b}
      {"type": "done",           "response": "...", "model": "...", "timestamp": "..."}
    """
    import os, asyncio, base64, io as _io, datetime, concurrent.futures
    import json as _json
    from PIL import Image as PILImage

    d        = _resolve_folder(folder)
    jf       = d / f"{stem}.json"
    img_path = _find_image(d, stem)

    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    data_doc = json.loads(jf.read_text(encoding="utf-8"))
    shapes   = data_doc.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")

    shape = shapes[idx]
    pts   = shape["points"]
    xs    = [p[0] for p in pts]; ys = [p[1] for p in pts]
    x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)

    if use_shadow:
        full_img = _get_shadow_page(folder, stem, img_path).convert("RGB")
    else:
        full_img = PILImage.open(str(img_path)).convert("RGB")
    iw, ih   = full_img.size
    pad      = 4
    crop     = full_img.crop((
        max(0, int(x1) - pad), max(0, int(y1) - pad),
        min(iw, int(x2) + pad), min(ih, int(y2) + pad),
    ))

    crop_top = max(0, int(y1) - pad)
    existing = _existing_row_bands_rel(shape, crop_top, crop.height)
    if rows_source == "existing":
        if not existing:
            raise HTTPException(status_code=400,
                detail="This cell has no internal row structure — pick "
                       "'re-detect rows' or create the structure first")
        rows = existing
    elif rows_source == "detect":
        rows = _detect_text_rows(crop, cell_height)
    else:                                   # auto: existing, else detect
        rows = existing or _detect_text_rows(crop, cell_height)
    # per-row OCR text is only meaningful when the bands ARE the structure's
    # rows (index-aligned); freshly detected bands have no per-row OCR
    row_ocr = ([(r.get("ocr") or "") for r in
                (shape.get("row_struct") or {}).get("rows") or []]
               if (existing and rows is existing) else [])
    prompt_text = body.prompt
    # Create the client BEFORE streaming starts: a missing API key / endpoint
    # becomes a clean HTTP 500 the editor can display, not a dead SSE stream.
    client = _make_llm_client(model)

    def gen():
        yield _json.dumps({"type": "lines_detected", "count": len(rows),
                           "lines": [list(r) for r in rows]})

        print(f"[LLM/linebyline] model={model} rows={len(rows)} payload={payload} "
              f"rows_source={rows_source} dry_run={dry_run} "
              f"prompt={prompt_text!r}", flush=True)

        line_responses: list[str] = []

        for i, (top, bottom) in enumerate(rows):
            txt = prompt_text
            if payload in ("ocr", "image+ocr"):
                txt = f"{prompt_text}\n\nOCR text:\n{row_ocr[i] if i < len(row_ocr) else ''}"
            content = [{"type": "text", "text": txt}]
            if payload in ("image", "image+ocr"):
                # Add a few pixels of vertical breathing room so ascenders /
                # descenders aren't clipped, then upscale very small rows so
                # the LLM can read digits reliably.
                row_pad = max(4, cell_height // 6)
                rt = max(0, top    - row_pad)
                rb = min(crop.height, bottom + row_pad)
                row_img = crop.crop((0, rt, crop.width, rb))
                if row_img.height < 48:
                    scale   = 48 / row_img.height
                    row_img = row_img.resize(
                        (int(row_img.width * scale), 48), PILImage.LANCZOS
                    )
                buf = _io.BytesIO()
                row_img.save(buf, format="JPEG", quality=92)
                b64 = base64.b64encode(buf.getvalue()).decode()
                content.append({"type": "image_url",
                                "image_url": {"url": f"data:image/jpeg;base64,{b64}",
                                              "detail": "high"}})

            messages = ([{"role": "user", "content": txt}] if payload == "ocr"
                        else [{"role": "user", "content": content}])
            try:
                resp = _llm_complete(client, model, messages, 64,
                                     use_cache=(cached and not dry_run))
                text = (resp.choices[0].message.content or "").strip()
            except Exception as exc:
                text = f"[error: {exc}]"

            line_responses.append(text)
            yield _json.dumps({"type": "row_result", "row": i, "text": text,
                               "top": top, "bottom": bottom})

        combined  = "\n".join(line_responses)
        timestamp = datetime.datetime.utcnow().isoformat() + "Z"

        if not dry_run:
            _apply_layer_rows(shape, [(t + crop_top, b + crop_top) for t, b in rows],
                              "llm", line_responses, "linebyline")
            shape["openai_output"] = {
                "response":       combined,
                "model":          model,
                "mode":           "linebyline",
                "timestamp":      timestamp,
                "lines_detected": len(rows),
            }
            fields = {"openai_output": shape["openai_output"]}
            if shape.get("row_struct"):
                fields["row_struct"] = shape["row_struct"]
            _merge_shape_fields(jf, idx, fields)   # parallel-safe per-shape write

        yield _json.dumps({"type": "done", "response": combined,
                           "model": model, "timestamp": timestamp,
                           "prompt_sent": prompt_text})

    # Stream the sync generator as SSE
    async def event_gen():
        loop = asyncio.get_event_loop()
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            it = iter(gen())
            while True:
                item = await loop.run_in_executor(pool, _safe_next, it)
                if item is _SSE_DONE:
                    break
                yield f"data: {item}\n\n"

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/page/shape/llm/anchored")
async def api_llm_anchored(
    folder:     str  = Query(...),
    stem:       str  = Query(...),
    idx:        int  = Query(...),
    n_rows:     int  = Query(...),
    model:      str  = Query("gpt-4o-mini"),
    use_shadow: bool = Query(False),
    ref_idx:    int  = Query(-1, description="Project this shape's row_struct bands instead of the histogram split"),
    cached:     bool = Query(True, description="Reuse identical past answers from the local cache"),
    body:       LlmRequest = ...,
):
    """
    LLM anchored: same SSE format as /llm/linebyline but uses forced n_rows
    split (histogram valley search) instead of auto-detecting row count.
    """
    import os, asyncio, base64, io as _io, datetime, concurrent.futures
    import json as _json
    from PIL import Image as PILImage

    d        = _resolve_folder(folder)
    jf       = d / f"{stem}.json"
    img_path = _find_image(d, stem)

    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")

    data_doc = json.loads(jf.read_text(encoding="utf-8"))
    shapes   = data_doc.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")

    shape = shapes[idx]
    pts   = shape["points"]
    xs    = [p[0] for p in pts]; ys = [p[1] for p in pts]
    x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)

    if use_shadow:
        full_img = _get_shadow_page(folder, stem, img_path).convert("RGB")
    else:
        full_img = PILImage.open(str(img_path)).convert("RGB")
    iw, ih = full_img.size
    pad    = 4
    crop   = full_img.crop((
        max(0, int(x1) - pad), max(0, int(y1) - pad),
        min(iw, int(x2) + pad), min(ih, int(y2) + pad),
    ))

    crop_top  = max(0, int(y1) - pad)
    projected = (_project_ref_bands(shapes[ref_idx], crop_top, crop.height, y1, y2)
                 if 0 <= ref_idx < len(shapes) else None)
    existing  = _existing_row_bands_rel(shape, crop_top, crop.height)
    rows        = projected or (existing if (existing and len(existing) == n_rows)
                                else _split_into_n_rows(crop, n_rows))
    prompt_text = body.prompt

    # Fail before the stream starts if the model's API isn't configured.
    client = _make_llm_client(model)

    def gen():
        yield _json.dumps({"type": "lines_detected", "count": len(rows),
                           "lines": [list(r) for r in rows]})

        line_responses: list[str] = []

        for i, (top, bottom) in enumerate(rows):
            row_pad = max(4, (bottom - top) // 6)
            rt = max(0, top    - row_pad)
            rb = min(crop.height, bottom + row_pad)
            row_img = crop.crop((0, rt, crop.width, rb))
            if row_img.height < 48:
                scale   = 48 / max(1, row_img.height)
                row_img = row_img.resize(
                    (max(1, int(row_img.width * scale)), 48), PILImage.LANCZOS
                )
            buf = _io.BytesIO()
            row_img.save(buf, format="JPEG", quality=92)
            b64 = base64.b64encode(buf.getvalue()).decode()

            messages = [{"role": "user", "content": [
                {"type": "text",      "text": prompt_text},
                {"type": "image_url", "image_url": {
                    "url": f"data:image/jpeg;base64,{b64}", "detail": "high"}},
            ]}]
            try:
                resp = _llm_complete(client, model, messages, 64, use_cache=cached)
                text = (resp.choices[0].message.content or "").strip()
            except Exception as exc:
                text = f"[error: {exc}]"

            line_responses.append(text)
            yield _json.dumps({"type": "row_result", "row": i, "text": text,
                               "top": top, "bottom": bottom})

        combined  = "\n".join(line_responses)
        timestamp = datetime.datetime.utcnow().isoformat() + "Z"
        _apply_layer_rows(shape, [(t + crop_top, b + crop_top) for t, b in rows],
                          "llm", line_responses, "anchored",
                          force_boxes=projected is not None)
        shape["openai_output"] = {
            "response":       combined,
            "model":          model,
            "mode":           "anchored",
            "timestamp":      timestamp,
            "lines_detected": len(rows),
        }
        fields = {"openai_output": shape["openai_output"]}
        if shape.get("row_struct"):
            fields["row_struct"] = shape["row_struct"]
        _merge_shape_fields(jf, idx, fields)   # parallel-safe per-shape write
        yield _json.dumps({"type": "done", "response": combined,
                           "model": model, "timestamp": timestamp})

    async def event_gen():
        loop = asyncio.get_event_loop()
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            it = iter(gen())
            while True:
                item = await loop.run_in_executor(pool, _safe_next, it)
                if item is _SSE_DONE:
                    break
                yield f"data: {item}\n\n"

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
# Routes — project management
# ---------------------------------------------------------------------------

@app.get("/api/projects")
def api_list_projects():
    return {"projects": list_projects()}


@app.get("/api/project/{name}")
def api_get_project(name: str):
    try:
        cfg   = load_config(name)
        state = load_pipeline(name)
        pdir  = project_dir(name)
        ann   = pdir / "annotations"
        n_json = len(list(ann.glob("*.json"))) if ann.exists() else 0
        n_img  = len(list(ann.glob("*.png")) + list(ann.glob("*.jpg"))) if ann.exists() else 0
        return {
            "config":  cfg,
            "pipeline": state,
            "stats":   {"json": n_json, "images": n_img},
            "stages":  stages_for(cfg["type"]),
            "annotations_path": str(ann),
        }
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


class ConfigUpdate(BaseModel):
    server: Optional[dict] = None
    llm:    Optional[dict] = None
    labels: Optional[list] = None
    server_profile: Optional[str] = None   # "" clears (back to inline/custom)


@app.patch("/api/project/{name}/config")
def api_update_config(name: str, body: ConfigUpdate):
    try:
        cfg  = load_config(name)
        pdir = project_dir(name)
        if body.server_profile is not None:
            if body.server_profile == "":
                cfg.pop("server_profile", None)
            else:
                cfg["server_profile"] = body.server_profile
        if body.server is not None:
            cfg["server"] = {**cfg.get("server", {}), **body.server}
        if body.llm is not None:
            cfg["llm"] = {**cfg.get("llm", {}), **body.llm}
        if body.labels is not None:
            cfg["labels"] = body.labels
        (pdir / "config.json").write_text(
            json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        return {"ok": True, "config": cfg}
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


class NewProject(BaseModel):
    name:   str
    type:   str = "A"
    labels: list = []

@app.post("/api/project/new")
def api_new_project(body: NewProject):
    from app.pipeline import create_project
    try:
        pdir = create_project(body.name, body.type, body.labels)
        return {"ok": True, "path": str(pdir)}
    except (FileExistsError, ValueError) as e:
        raise HTTPException(status_code=400, detail=str(e))


class CloneProject(BaseModel):
    new_name: str

@app.post("/api/project/{name}/clone")
def api_clone_project(name: str, body: CloneProject):
    from app.pipeline import clone_project
    try:
        pdir = clone_project(name, body.new_name)
        return {"ok": True, "path": str(pdir)}
    except (FileNotFoundError, FileExistsError) as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/project/{name}/advance")
def api_advance(name: str):
    try:
        new_stage = advance_stage(name)
        return {"ok": True, "stage": new_stage}
    except (FileNotFoundError, ValueError) as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/project/{name}/set-stage")
def api_set_stage(name: str, stage: str = Query(...)):
    try:
        set_stage(name, stage)
        return {"ok": True, "stage": stage}
    except (FileNotFoundError, ValueError) as e:
        raise HTTPException(status_code=400, detail=str(e))


# ---------------------------------------------------------------------------
# Docker container configuration
# ---------------------------------------------------------------------------
from app import docker_config as _docker_cfg_mod

def _docker_cfg() -> dict:
    return _docker_cfg_mod.load()

def _ns_suffix(host_path: str) -> str:
    """A stable, docker-safe suffix derived from the host workspace path a
    container will mount. Each distinct workspace gets its own container, so
    two users on the same GPU host (whose /workspace binds differ) never
    share — and can never clobber — one another's container. A short hash of
    the path (no directory names, to keep container names brand-clean).
    Empty path → no suffix (legacy single-user behavior)."""
    import hashlib
    host_path = (host_path or "").rstrip("/")
    if not host_path:
        return ""
    return hashlib.sha1(host_path.encode("utf-8")).hexdigest()[:10]

def _predict_workspace_root(srv: dict) -> str:
    """Host dir the predict container mounts as /workspace."""
    return (srv.get("predict_remote_path") or srv.get("remote_path") or "").rstrip("/")

def _predict_container(srv: dict | None = None) -> str:
    base = _docker_cfg()["predict_container"]
    sfx = _ns_suffix(_predict_workspace_root(srv)) if srv else ""
    return f"{base}_{sfx}" if sfx else base

def _train_container(srv: dict | None = None) -> str:
    base = _docker_cfg()["train_container"]
    sfx = _ns_suffix((srv.get("remote_path") or "").rstrip("/")) if srv else ""
    return f"{base}_{sfx}" if sfx else base

def _stop_container_gen(srv: dict, passphrase, container: str, tag: str, keep: bool):
    """Stop a Docker container after a job finishes, yielding SSE log lines.
    No-op when keep=True. Best-effort: failures are reported, not raised."""
    if keep:
        yield f"[{tag}] keep_container set — leaving '{container}' running."
        return
    from app import ssh_ops
    yield f"[{tag}] Job finished — stopping container '{container}'..."
    try:
        c = ssh_ops._client(srv["host"], srv["user"], srv["key_path"], passphrase)
        _, out, _ = c.exec_command(f"docker stop {container}")
        out.read()
        c.close()
        yield f"[{tag}] Container '{container}' stopped."
    except Exception as e:
        yield f"[{tag}] Could not stop container '{container}': {e}"


class DockerConfigBody(BaseModel):
    predict_container: Optional[str] = None
    train_container:   Optional[str] = None
    image_name:        Optional[str] = None


@app.get("/api/docker-config")
def api_get_docker_config():
    return _docker_cfg()


@app.post("/api/docker-config")
def api_set_docker_config(body: DockerConfigBody):
    update = {k: v for k, v in body.dict().items() if v is not None}
    return _docker_cfg_mod.save(update)


@app.get("/api/docker-config/status")
def api_docker_container_status(body: SshRequest = None,
                                passphrase: Optional[str] = Query(None)):
    """Check whether predict/train containers exist on the GPU server.
    Returns exists/running status for each. Requires a project with server cfg."""
    # Use the first server-configured project (list_projects() yields dicts).
    from app import ssh_ops
    names = _server_project_names()
    if not names:
        raise HTTPException(status_code=400, detail="Server not configured on any project")
    srv = _server_cfg(names[0])

    pw = passphrase or srv.get("passphrase") or ""
    result = {}
    for role, cname in [("predict", _predict_container(srv)),
                        ("train",   _train_container(srv))]:
        try:
            c = ssh_ops._client(srv["host"], srv["user"], srv["key_path"], pw)
            _, out, _ = c.exec_command(
                f"docker ps -a --filter name=^{cname}$ --format '{{{{.Status}}}}'"
            )
            status_str = out.read().decode().strip()
            c.close()
            if not status_str:
                result[role] = {"name": cname, "exists": False, "running": False}
            else:
                result[role] = {"name": cname, "exists": True,
                                "running": status_str.lower().startswith("up")}
        except Exception as e:
            result[role] = {"name": cname, "exists": None, "error": str(e)}
    return result


@app.post("/api/docker-config/build")
async def api_docker_build(passphrase: Optional[str] = Query(None),
                           role: str = Query("predict"),
                           project: Optional[str] = Query(None)):
    """Build the dedust-layout image on the GPU server and create the named
    container(s). role: 'predict' | 'train' | 'both'. If `project` is given the
    SSH connection uses that project's server config; otherwise the first
    server-configured project. Because container names/mounts are per-workspace,
    we create one container per DISTINCT workspace across all projects on the
    same host — so whichever project you later infer from, its container exists.
    Streams SSE log lines."""
    from app import ssh_ops
    import posixpath as pp

    names = _server_project_names()
    if not names:
        raise HTTPException(status_code=400, detail="Server not configured on any project")
    srv = _server_cfg(project) if project else _server_cfg(names[0])

    cfg      = _docker_cfg()
    image    = cfg["image_name"]
    pw       = passphrase or srv.get("passphrase") or ""

    # Read the local Dockerfile and upload it
    dockerfile_path = Path(__file__).parent.parent / "Dockerfile"
    if not dockerfile_path.exists():
        raise HTTPException(status_code=500, detail="Dockerfile not found in repo root")
    dockerfile_content = dockerfile_path.read_text(encoding="utf-8")

    # All server configs sharing this host — one container per distinct
    # workspace root so every project on the host is covered.
    host_srvs = []
    for nm in names:
        try:
            s = _server_cfg(nm)
        except HTTPException:
            continue
        if s.get("host") == srv["host"]:
            host_srvs.append(s)

    roles_to_build = []
    if role in ("predict", "both"):
        for r in sorted({_predict_workspace_root(s) for s in host_srvs
                         if _predict_workspace_root(s)}):
            roles_to_build.append(("predict", _predict_container({"predict_remote_path": r}), r))
    if role in ("train", "both"):
        for r in sorted({(s.get("remote_path") or "").rstrip("/") for s in host_srvs
                         if (s.get("remote_path") or "").strip()}):
            roles_to_build.append(("train", _train_container({"remote_path": r}), r))
    # Deduplicate — same name = only create once
    seen = set()
    roles_unique = []
    for r, n, ws in roles_to_build:
        if n not in seen:
            seen.add(n)
            roles_unique.append((r, n, ws))

    def build_gen():
        c = ssh_ops._client(srv["host"], srv["user"], srv["key_path"], pw)
        sftp = c.open_sftp()
        try:
            # Upload Dockerfile to server
            remote_df = "/tmp/dedust_Dockerfile"
            yield f"[docker] Uploading Dockerfile to {remote_df}..."
            with sftp.open(remote_df, "w") as f:
                f.write(dockerfile_content)
            sftp.close()

            # Build image
            yield f"[docker] Building image '{image}' (this can take 10-30 min)..."
            build_cmd = f"docker build -t {image} -f {remote_df} /tmp"
            yield from ssh_ops.stream_command(
                srv["host"], srv["user"], srv["key_path"], build_cmd, pw)

            # Create each new container, each bound to its own workspace root
            # (deterministic from config — no cross-container mount inheritance).
            for _role, cname, workspace_host in roles_unique:
                if not workspace_host:
                    yield (f"[docker] SKIP {_role}: no workspace path configured "
                           f"(set remote_path / predict_remote_path first).")
                    continue
                yield f"[docker] {_role} container '{cname}'  /workspace <- {workspace_host}"
                # Is it already there, and running?
                _, chk, _ = c.exec_command(
                    f"docker ps -a --filter name=^{cname}$ --format '{{{{.Status}}}}'"
                )
                existing = chk.read().decode().strip()
                if existing:
                    if existing.lower().startswith("up"):
                        yield f"[docker] Container '{cname}' is running — left as-is."
                        continue
                    # Stopped/exited: remove and recreate so it picks up the
                    # current mount + keep-alive config (self-healing rebuild).
                    yield f"[docker] Container '{cname}' exists but stopped — recreating."
                    _, rmo, rme = c.exec_command(f"docker rm {cname}")
                    rmo.read(); rme.read()
                yield f"[docker] Creating container '{cname}'..."
                # Keep-alive = a sentinel loop instead of `tail -f /dev/null`:
                # a job's last act can `touch /tmp/dedust_selfstop` and the
                # container stops ITSELF (a container cannot docker-stop itself
                # and PID 1 ignores in-namespace kill signals). This is what
                # frees the GPU even when the browser that launched the job is
                # long gone. The stale sentinel is cleared on every start.
                create_cmd = (
                    f"docker create --name {cname} --gpus all --shm-size=8g "
                    f"-v {workspace_host}:/workspace {image} "
                    f"bash -c 'rm -f /tmp/dedust_selfstop; "
                    f"while [ ! -f /tmp/dedust_selfstop ]; do sleep 5; done'"
                )
                _, co, ce = c.exec_command(create_cmd)
                out_txt = co.read().decode().strip()
                err_txt = ce.read().decode().strip()
                if out_txt:
                    yield f"[docker] Created: {out_txt[:64]}"
                if err_txt:
                    yield f"[docker] {err_txt}"

            yield "[docker] Done."
        finally:
            try:
                c.close()
            except Exception:
                pass

    return await _sse_stream(build_gen())


# ---------------------------------------------------------------------------
# Routes — SSH / server operations
# ---------------------------------------------------------------------------

def _server_project_names() -> list[str]:
    """Names of projects that have a usable server config. list_projects()
    returns summary DICTS, so callers must not pass those straight to
    _server_cfg (which expects a project name)."""
    names = []
    for p in list_projects():
        nm = p["name"] if isinstance(p, dict) else p
        try:
            _server_cfg(nm)
            names.append(nm)
        except HTTPException:
            continue
    return names


def _server_cfg(name: str) -> dict:
    """Resolve a project's GPU server settings.

    If the project references a named profile (config.json "server_profile"),
    the profile — defined per-instance in app/gpu_servers.json — wins entirely;
    the same logical server needs different key paths from different machines,
    so profiles are resolved on the instance that executes the job. Projects
    without a profile keep the legacy inline server block. A profile may carry
    a stored key passphrase; endpoints use it when the request has none."""
    from app import gpu_profiles
    cfg = load_config(name)
    profile_name = cfg.get("server_profile")
    if profile_name:
        srv = gpu_profiles.get(profile_name)
        if srv is None:
            raise HTTPException(
                status_code=400,
                detail=(f"GPU profile '{profile_name}' is not defined on this "
                        f"instance — add it in the dashboard's GPU Server card "
                        f"(or switch the project to different settings)."))
    else:
        srv = cfg.get("server", {})
    if not srv.get("host"):
        raise HTTPException(status_code=400, detail="Server host not configured")
    if not srv.get("user"):
        raise HTTPException(status_code=400, detail="Server user not configured")
    # key_path may be empty: password-auth servers (VPN + plain login) use the
    # passphrase field / stored profile passphrase as the password instead.
    srv.setdefault("key_path", "")
    return srv


# ── GPU server profiles (named backends: koren / azure-gpu / …) ──────────────

class GpuProfileBody(BaseModel):
    name:   str
    server: dict


@app.get("/api/gpu-profiles")
def api_gpu_profiles_list():
    """All profiles on this instance. Passphrases are masked — the UI only
    needs to know whether one is stored."""
    from app import gpu_profiles
    out = {}
    for pname, srv in gpu_profiles.load_all().items():
        out[pname] = {**{k: v for k, v in srv.items() if k != "passphrase"},
                      "has_passphrase": bool(srv.get("passphrase"))}
    return {"profiles": out}


@app.post("/api/gpu-profiles")
def api_gpu_profiles_save(body: GpuProfileBody):
    from app import gpu_profiles
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Profile name required")
    server = dict(body.server or {})
    # empty passphrase in the payload = keep the stored one (the UI never
    # echoes it back); to actually clear it the UI sends passphrase=None + clear
    if not server.get("passphrase") and not server.pop("clear_passphrase", False):
        old = gpu_profiles.get(name) or {}
        if old.get("passphrase"):
            server["passphrase"] = old["passphrase"]
    gpu_profiles.save(name, server)
    return {"ok": True}


@app.delete("/api/gpu-profiles/{name}")
def api_gpu_profiles_delete(name: str):
    from app import gpu_profiles
    gpu_profiles.delete(name)
    return {"ok": True}


class SshRequest(BaseModel):
    passphrase: Optional[str] = None

class SshPullRequest(BaseModel):
    passphrase: Optional[str] = None
    subfolder:  str = "annotations"

class JobSubmit(BaseModel):
    command:    str
    passphrase: Optional[str] = None


@app.post("/api/project/{name}/server/test")
def api_server_test(name: str, body: SshRequest = SshRequest()):
    from app import ssh_ops
    try:
        srv = _server_cfg(name)
        return ssh_ops.test_connection(srv["host"], srv["user"], srv["key_path"],
                                       body.passphrase or srv.get("passphrase"))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/project/{name}/server/push")
def api_server_push(name: str, body: SshRequest = SshRequest()):
    """Upload the project's annotations/ folder to the server."""
    from app import ssh_ops
    try:
        srv  = _server_cfg(name)
        pdir = project_dir(name)
        local_ann  = pdir / "annotations"
        remote_ann = posixpath.join(srv["remote_path"], name, "annotations")
        if not local_ann.exists():
            raise HTTPException(status_code=400, detail="annotations/ folder does not exist")
        return ssh_ops.push_folder(
            srv["host"], srv["user"], srv["key_path"],
            local_ann, remote_ann, body.passphrase or srv.get("passphrase"),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/project/{name}/server/pull")
def api_server_pull(name: str, body: SshPullRequest = SshPullRequest()):
    """Download a subfolder (default: annotations/) back from the server."""
    from app import ssh_ops
    try:
        srv  = _server_cfg(name)
        pdir = project_dir(name)
        remote = posixpath.join(srv["remote_path"], name, body.subfolder)
        local  = pdir / body.subfolder
        return ssh_ops.pull_folder(
            srv["host"], srv["user"], srv["key_path"],
            remote, local, body.passphrase or srv.get("passphrase"),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/project/{name}/job/submit")
def api_job_submit(name: str, body: JobSubmit):
    """Submit an arbitrary shell command as a background job on the server."""
    from app import ssh_ops
    import time as _time
    try:
        srv      = _server_cfg(name)
        log_path = posixpath.join(
            srv["remote_path"], name, f"job_{int(_time.time())}.log"
        )
        result = ssh_ops.submit_job(
            srv["host"], srv["user"], srv["key_path"],
            body.command, log_path, body.passphrase or srv.get("passphrase"),
        )
        if result["ok"]:
            state = load_pipeline(name)
            state["job"] = {"pid": result["pid"], "log_path": log_path,
                            "command": body.command}
            save_pipeline(name, state)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/project/{name}/job/status")
def api_job_status(name: str, passphrase: Optional[str] = Query(None)):
    """Poll the status of the last submitted background job."""
    from app import ssh_ops
    try:
        srv   = _server_cfg(name)
        state = load_pipeline(name)
        job   = state.get("job")
        if not job:
            return {"ok": True, "running": False, "log_tail": "(no job submitted yet)"}
        return ssh_ops.job_status(
            srv["host"], srv["user"], srv["key_path"],
            job.get("pid"), job["log_path"], passphrase or srv.get("passphrase"),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Routes — page import
# ---------------------------------------------------------------------------

class ImportRequest(BaseModel):
    source_path: str  # local folder or single file path

@app.post("/api/project/{name}/import")
def api_import_pages(name: str, body: ImportRequest):
    """Import PDFs/images from a local path directly — no upload needed."""
    from app.page_import import import_from_path
    try:
        pdir    = project_dir(name)
        ann_dir = pdir / "annotations"
        results = import_from_path(Path(body.source_path), ann_dir)
        return {"ok": True, "pages": results, "total": len(results)}
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=traceback.format_exc())


@app.post("/api/project/{name}/upload-pages")
async def api_upload_pages(name: str, files: List[UploadFile] = File(...)):
    """Accept uploaded PDF/image files, convert them and add to the project's annotations/ folder."""
    import tempfile
    from app.page_import import _import_pdf, _save_image_file, _sanitize, IMAGE_EXTS

    try:
        pdir    = project_dir(name)
        ann_dir = pdir / "annotations"
        ann_dir.mkdir(parents=True, exist_ok=True)

        src_dir = pdir / "sources"
        src_dir.mkdir(exist_ok=True)

        results = []
        for uf in files:
            filename  = uf.filename or "upload"
            orig_stem = Path(filename).stem
            ext       = Path(filename).suffix.lower()

            content  = await uf.read()
            tmp_path = None
            try:
                if ext == ".pdf":
                    # Save PDF permanently so pdf_source stays valid for text-layer extraction
                    saved_pdf = src_dir / filename
                    saved_pdf.write_bytes(content)
                    base  = _sanitize(orig_stem)
                    pages = _import_pdf(saved_pdf, ann_dir, base=base)
                    results.extend(pages)
                elif ext in IMAGE_EXTS:
                    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                        tmp.write(content)
                        tmp_path = Path(tmp.name)
                    stem = _sanitize(orig_stem)
                    info = _save_image_file(tmp_path, ann_dir, stem)
                    results.append(info)
                else:
                    results.append({"stem": filename, "error": f"Unsupported extension: {ext}"})
            except Exception as e:
                import traceback
                results.append({"stem": filename, "error": traceback.format_exc()})
            finally:
                if tmp_path and tmp_path.exists():
                    tmp_path.unlink()

        return {"ok": True, "pages": results, "total": len(results)}
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=traceback.format_exc())


# ---------------------------------------------------------------------------
# Routes — prepare training data (LabelMe → COCO)
# ---------------------------------------------------------------------------

BASE_YAML = Path(__file__).parent.parent / "samples" / "ertesito2" / "fast_rcnn_R_50_FPN_3x.yaml"

# Shared, project-independent authority files (gazetteers / controlled vocabs).
# Data files here are git-ignored; see authorities/README.md for the schema.
AUTHORITIES_DIR = Path(__file__).parent.parent / "authorities"
AUTHORITIES_DIR.mkdir(exist_ok=True)


# ---------------------------------------------------------------------------
# Authority resolver — match OCR'd place names to gazetteer entity IDs
# ---------------------------------------------------------------------------
# Index a *.authority.json (see authorities/README.md) into in-memory lookup
# structures and fuzzy-match query strings (place names) against every entity
# name + alias, optionally constrained to descendants of a parent (county /
# district). rapidfuzz preferred; degrades to difflib if unavailable.

_AUTH_DEFAULT_FILE = "places_hu.authority.json"
_AUTH_PARENT_BOOST = 12.0    # soft ranking bonus for candidates under the context parent
_AUTH_MIN_ACCEPT   = 70.0    # batch auto-accept floor: below this = no real string match → don't guess
_AUTH_LOCK = threading.Lock()


def _auth_strip_status(s: str, tokens: set) -> str:
    """Drop trailing status tokens declared by the authority's `query_strip`
    (e.g. for places_hu: 'Szombathely rtv' -> 'szombathely'). Authority-specific,
    so it's read from the file, not hardcoded."""
    if not tokens:
        return s
    toks = s.split()
    while len(toks) > 1 and toks[-1] in tokens:
        toks.pop()
    return " ".join(toks) if toks else s


def _auth_via_rank(via: str) -> int:
    """Tie-break preference: primary period name beats period aliases, which
    beat modern/foreign aliases. Lower = preferred."""
    if via == "name":
        return 0
    if via in ("modern", "english"):
        return 2
    return 1
_AUTH_CACHE: dict = {}   # filename -> {"mtime": float, "index": dict}


def _auth_norm(s: str) -> str:
    """Casefold, strip punctuation, collapse whitespace — accent-preserving key."""
    s = (s or "").strip().casefold()
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    return re.sub(r"\s+", " ", s).strip()


# Import ONCE at module scope. A failed import inside the per-call function is
# a performance landmine: Python does not cache import failures, so every call
# re-scans sys.path (with the 32k-entry places index that was ~40 s of nt.stat
# calls per index build — the "authority resolve freezes the server" bug).
try:
    from unidecode import unidecode as _unidecode
except ImportError:
    _unidecode = None


def _auth_fold(s: str) -> str:
    """Accent-folded key (Vágagyagos -> vagagyagos) — tolerant of OCR dropping diacritics."""
    if _unidecode is not None:
        s = _unidecode(s or "")
    return _auth_norm(s)


def _auth_entity_view(e: dict) -> dict:
    """Flatten one authority entity to the fields the editor cares about.
    Uses the first slice (period-specific name/type/parent + county/district)."""
    sl = (e.get("slices") or [{}])
    s0 = sl[0] if sl else {}
    sattrs = s0.get("attrs") or {}
    eattrs = e.get("attrs") or {}
    return {
        "id":            e.get("id"),
        "name":          s0.get("name") or e.get("name"),
        "type":          s0.get("type"),
        "parent":        s0.get("parent"),
        "county_name":   sattrs.get("county_name"),
        "district_name": sattrs.get("district_name"),
        "lat":           eattrs.get("lat"),
        "lon":           eattrs.get("lon"),
        "modern_name":   eattrs.get("modern_name"),
    }


def _build_auth_index(data: dict) -> dict:
    by_id: dict = {}
    children: dict = {}
    pools: dict = {"county": [], "district": [], "settlement": []}
    for e in data.get("entities", []):
        v = _auth_entity_view(e)
        if not v["id"]:
            continue
        by_id[v["id"]] = v
        if v["parent"]:
            children.setdefault(v["parent"], []).append(v["id"])
        t = v["type"] or "settlement"
        cands = [(v["name"], "name")]
        for a in (e.get("aliases") or []):
            an = a.get("name")
            if an:
                cands.append((an, a.get("source") or "alias"))
        seen = set()
        for raw, via in cands:
            key = _auth_norm(raw)
            if not key or (key, via) in seen:
                continue
            seen.add((key, via))
            pools.setdefault(t, []).append({
                "id": v["id"], "raw": raw, "norm": key,
                "fold": _auth_fold(raw), "via": via, "parent": v["parent"],
            })
    # Precompute, once per load, the structures the matcher would otherwise
    # rebuild on every request: folded-key lists (for rapidfuzz) and exact
    # normalized-key maps (O(1) exact lookup instead of scanning the pool).
    folds_by_type = {t: [c["fold"] for c in pool] for t, pool in pools.items()}
    exact_by_type = {}
    for t, pool in pools.items():
        em: dict = {}
        for c in pool:
            em.setdefault(c["norm"], []).append(c)
        exact_by_type[t] = em
    pool_all, folds_all = [], []
    for t in pools:
        pool_all.extend(pools[t]); folds_all.extend(folds_by_type[t])
    exact_all: dict = {}
    for c in pool_all:
        exact_all.setdefault(c["norm"], []).append(c)
    # Authority-declared query normalization: trailing tokens to strip from a
    # query before matching (e.g. admin-status suffixes "rtv"/"tjv" for places).
    strip = {(_auth_norm(t)) for t in (data.get("query_strip") or []) if str(t).strip()}
    strip.discard("")
    return {"by_id": by_id, "children": children, "pools": pools,
            "folds_by_type": folds_by_type, "exact_by_type": exact_by_type,
            "pool_all": pool_all, "folds_all": folds_all, "exact_all": exact_all,
            "strip": strip}


def _load_authority(name: Optional[str] = None) -> dict:
    fname = name or _AUTH_DEFAULT_FILE
    path = AUTHORITIES_DIR / fname
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Authority file not found: {fname}")
    mtime = path.stat().st_mtime
    # Build INSIDE the lock: concurrent requests after an mtime change (server
    # start, file update, Dropbox touch) must wait for ONE build — building
    # outside the lock let every in-flight request start its own multi-second
    # build (thundering herd = the editor freezing for minutes).
    with _AUTH_LOCK:
        cached = _AUTH_CACHE.get(fname)
        if cached and cached["mtime"] == mtime:
            return cached["index"]
        data = json.loads(path.read_text(encoding="utf-8"))
        index = _build_auth_index(data)
        _AUTH_CACHE[fname] = {"mtime": mtime, "index": index}
        return index


_AUTH_META_CACHE: dict = {}   # filename -> {"mtime", "meta"}


def _authority_meta(fname: str) -> dict:
    """Header metadata for one authority file (no entity indexing). Cached by mtime."""
    path = AUTHORITIES_DIR / fname
    mtime = path.stat().st_mtime
    cached = _AUTH_META_CACHE.get(fname)
    if cached and cached["mtime"] == mtime:
        return cached["meta"]
    data = json.loads(path.read_text(encoding="utf-8"))
    meta = {
        "file":           fname,
        "authority":      data.get("authority"),
        "version":        data.get("version"),
        "entity_types":   data.get("entity_types") or [],
        "slices_present": data.get("slices_present") or [],
        "counts":         data.get("counts") or {},
    }
    _AUTH_META_CACHE[fname] = {"mtime": mtime, "meta": meta}
    return meta


@app.get("/api/authorities")
def api_list_authorities():
    """List the *.authority.json files available in AUTHORITIES_DIR with header
    metadata (entity types, counts). Powers the in-editor authority picker."""
    out = []
    for p in sorted(AUTHORITIES_DIR.glob("*.authority.json")):
        try:
            out.append(_authority_meta(p.name))
        except Exception as e:
            out.append({"file": p.name, "error": str(e)})
    return {"authorities": out, "default": _AUTH_DEFAULT_FILE}


def _auth_is_descendant(by_id: dict, node_id: str, ancestor_id: str) -> bool:
    cur, seen = node_id, set()
    while cur and cur not in seen:
        if cur == ancestor_id:
            return True
        seen.add(cur)
        cur = (by_id.get(cur) or {}).get("parent")
    return False


def _authority_match(q: str, type: Optional[str] = None, parent: Optional[str] = None,
                     k: int = 8, name: Optional[str] = None) -> list:
    index = _load_authority(name)
    by_id = index["by_id"]
    strip = index.get("strip") or set()
    nq, fq = _auth_strip_status(_auth_norm(q), strip), _auth_strip_status(_auth_fold(q), strip)
    if not nq:
        return []
    single = type in index["pools"]
    # Always search the full (cached) pool — `parent` is a SOFT preference
    # applied at ranking time, not a hard filter (see below).
    pool  = index["pools"][type]        if single else index["pool_all"]
    folds = index["folds_by_type"][type] if single else index["folds_all"]
    emap  = index["exact_by_type"][type] if single else index["exact_all"]
    if not pool:
        return []

    best: dict = {}   # id -> {"score", "via", "matched"}  (score = raw similarity)

    def _add(c, score):
        cur = best.get(c["id"])
        if not cur or score > cur["score"]:
            best[c["id"]] = {"score": float(score), "via": c["via"], "matched": c["raw"]}

    for c in emap.get(nq, []):           # exact (accent-preserving), O(1)
        _add(c, 100.0)

    limit = max(k * 5, 50)
    try:
        from rapidfuzz import process, fuzz
        # folds are already normalized → processor=None skips rapidfuzz's
        # per-comparison preprocessing (meaningfully faster over a big pool).
        for _, score, i in process.extract(fq, folds, scorer=fuzz.WRatio,
                                            processor=None, limit=limit,
                                            score_cutoff=55):
            _add(pool[i], 100.0 if pool[i]["norm"] == nq else score)
    except ImportError:
        import difflib
        for c in pool:
            if c["norm"] == nq:
                continue
            r = difflib.SequenceMatcher(None, fq, c["fold"]).ratio() * 100.0
            if r >= 55:
                _add(c, r)

    # Parent context is a SOFT preference: a strong string match OUTSIDE the
    # context (e.g. a district-seat town that lives under the county, not the
    # district) still beats a weak in-context match, while same-named places are
    # tie-broken toward the context. `score` stays the raw similarity so callers
    # can reject low-similarity guesses on real similarity (not the boost).
    def _sort_key(kv):
        cid, m = kv
        boost = _AUTH_PARENT_BOOST if (parent and _auth_is_descendant(by_id, cid, parent)) else 0.0
        # primary tie-break: adjusted score; secondary: matched via primary name
        return (-(m["score"] + boost), _auth_via_rank(m["via"]))

    out = []
    for cid, m in sorted(best.items(), key=_sort_key)[:k]:
        v = by_id.get(cid, {})
        out.append({**v, "score": round(m["score"], 1),
                    "via": m["via"], "matched": m["matched"]})
    return out


@app.get("/api/authority/resolve")
def api_authority_resolve(q: str = Query(...), type: Optional[str] = Query(None),
                          parent: Optional[str] = Query(None), k: int = Query(8),
                          name: Optional[str] = Query(None)):
    """Fuzzy-match a place string against the authority; return top-k candidates."""
    return {"query": q, "candidates": _authority_match(q, type, parent, k, name)}


@app.get("/api/authority/children")
def api_authority_children(parent: Optional[str] = Query(None),
                           name: Optional[str] = Query(None)):
    """List entities directly under `parent`; omit `parent` for the roots
    (entities with no parent — counties for places, top groups for industries).
    Authority-agnostic, so it powers the context pickers for any authority."""
    index = _load_authority(name)
    by_id = index["by_id"]
    if parent:
        ids = index["children"].get(parent, [])
    else:
        ids = [i for i, v in by_id.items() if not v.get("parent")]
    items = sorted((by_id[i] for i in ids if i in by_id),
                   key=lambda v: (v.get("name") or ""))
    return {"parent": parent, "count": len(items), "items": items}


@app.get("/api/authority/entity")
def api_authority_entity(id: str = Query(...), name: Optional[str] = Query(None)):
    index = _load_authority(name)
    v = index["by_id"].get(id)
    if not v:
        raise HTTPException(status_code=404, detail=f"Entity not found: {id}")
    return v


class AuthorityAssign(BaseModel):
    id:     Optional[str] = None    # entity id, or null/omitted to clear
    source: Optional[str] = "human"
    name:   Optional[str] = None    # authority filename override


@app.post("/api/page/shape/authority")
def api_shape_authority(folder: str = Query(...), stem: str = Query(...),
                        idx: int = Query(...), body: AuthorityAssign = ...):
    """Assign (or clear) the resolved authority entity on one shape."""
    d  = _resolve_folder(folder)
    jf = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")
    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail=f"Shape index {idx} out of range")
    shape = shapes[idx]
    if not body.id:
        shape.pop("authority", None)
        _write_json(jf, data)
        return {"ok": True, "authority": None}
    v = _load_authority(body.name)["by_id"].get(body.id)
    if not v:
        raise HTTPException(status_code=404, detail=f"Entity not found: {body.id}")
    import datetime as _dt
    shape["authority"] = {
        "id":            v["id"],
        "name":          v["name"],
        "type":          v["type"],
        "parent":        v["parent"],
        "county_name":   v.get("county_name"),
        "district_name": v.get("district_name"),
        "lat":           v.get("lat"),
        "lon":           v.get("lon"),
        "source":        body.source or "human",
        "ts":            _dt.datetime.utcnow().isoformat() + "Z",
    }
    _write_json(jf, data)
    return {"ok": True, "authority": shape["authority"]}


class PageFlagsUpdate(BaseModel):
    flags: dict


@app.patch("/api/page/flags")
def api_page_flags(folder: str = Query(...), stem: str = Query(...),
                   body: PageFlagsUpdate = ...):
    """Merge top-level keys into a page's `flags` object (page-level metadata
    such as the per-table authority context). Last-writer-wins per key."""
    d  = _resolve_folder(folder)
    jf = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")
    data = json.loads(jf.read_text(encoding="utf-8"))
    flags = data.get("flags") or {}
    flags.update(body.flags or {})
    data["flags"] = flags
    _write_json(jf, data)
    return {"ok": True, "flags": flags}


# ---------------------------------------------------------------------------
# Structured extraction — per-project JSON schemas + per-shape structured data
# ---------------------------------------------------------------------------

def _schemas_dir(folder: str) -> Path:
    """Project-level schemas live next to annotations/ at <project>/schemas/."""
    return _resolve_folder(folder).parent / "schemas"


def _safe_schema_name(name: str) -> str:
    base = re.sub(r"[^A-Za-z0-9_\- ]+", "", (name or "")).strip().replace(" ", "_")
    return base or "schema"


@app.get("/api/schemas")
def api_list_schemas(folder: str = Query(...)):
    """List the project's JSON extraction schemas (<project>/schemas/*.json)."""
    sdir = _schemas_dir(folder)
    out = []
    if sdir.exists():
        for p in sorted(sdir.glob("*.json")):
            try:
                out.append({"name": p.stem, "schema": json.loads(p.read_text(encoding="utf-8"))})
            except Exception as e:
                out.append({"name": p.stem, "error": str(e)})
    return {"schemas": out}


@app.put("/api/schemas")
def api_save_schema(folder: str = Query(...), name: str = Query(...),
                    schema_obj: dict = ...):
    """Save a JSON schema (the request body IS the schema object)."""
    sdir = _schemas_dir(folder)
    sdir.mkdir(parents=True, exist_ok=True)
    fname = _safe_schema_name(name)
    (sdir / f"{fname}.json").write_text(
        json.dumps(schema_obj, indent=2, ensure_ascii=False), encoding="utf-8")
    return {"ok": True, "name": fname}


@app.delete("/api/schemas")
def api_delete_schema(folder: str = Query(...), name: str = Query(...)):
    p = _schemas_dir(folder) / f"{_safe_schema_name(name)}.json"
    if p.exists():
        p.unlink()
    return {"ok": True}


class StructuredSave(BaseModel):
    data: Optional[dict] = None          # null/omitted to clear
    schema_name: Optional[str] = None


@app.patch("/api/page/shape/structured")
def api_shape_structured(folder: str = Query(...), stem: str = Query(...),
                         idx: int = Query(...), body: StructuredSave = ...):
    """Save the human-edited structured record for a shape (or clear it)."""
    d  = _resolve_folder(folder)
    jf = d / f"{stem}.json"
    if not jf.exists():
        raise HTTPException(status_code=404, detail=f"JSON not found: {jf}")
    data   = json.loads(jf.read_text(encoding="utf-8"))
    shapes = data.get("shapes", [])
    if idx < 0 or idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")
    shape = shapes[idx]
    if body.data is None:
        shape.pop("structured", None)
        _write_json(jf, data)
        return {"ok": True, "structured": None}
    st = shape.get("structured") or {}
    st["data"]   = body.data
    st["edited"] = True
    if body.schema_name is not None:
        st["schema_name"] = body.schema_name
    import datetime as _dt
    st["ts_edited"] = _dt.datetime.utcnow().isoformat() + "Z"
    shape["structured"] = st
    _write_json(jf, data)
    return {"ok": True, "structured": st}


# ---------------------------------------------------------------------------
# Authority — batch resolution across pages / columns (server-side, in-process)
# ---------------------------------------------------------------------------

_DITTO_RE = re.compile(r'^[\s.,\-–—―_=~:;"\'’‘”“„«»<>]+$')
_DITTO_WORDS = {"do", "dto", "ditto", "detto", "idem", "id", "ua", "uaz",
                "ugyanaz", "ugyanott", "uo", "uott", "azelobbi"}


def _auth_is_ditto(s: str) -> bool:
    s = (s or "").strip()
    if not s:
        return False
    if _DITTO_RE.match(s):
        return True
    return re.sub(r"[.\s]", "", s.lower()) in _DITTO_WORDS


_AUTH_MAX_QLEN = 80   # a single place/industry name is short; longer = not a name


def _auth_resolvable(s: str) -> bool:
    """Worth a lookup: has ≥2 non-space chars and isn't an insanely long blob
    (a mis-segmented paragraph etc., which would never be a single entity)."""
    s = (s or "").strip()
    return len(re.sub(r"\s", "", s)) >= 2 and len(s) <= _AUTH_MAX_QLEN


def _auth_layer_text(human, ocr, llm, pdf, layer):
    h, o, l, p = (human or "").strip(), (ocr or "").strip(), (llm or "").strip(), (pdf or "").strip()
    if layer == "human":    return h
    if layer == "ocr":      return o
    if layer == "llm":      return l
    if layer == "pdf":      return p
    if layer == "best_ocr": return h or o or l
    if layer == "best_pdf": return h or l or o or p
    return h or l or o      # best_llm (default)


def _auth_obj_from_cand(cand: dict, source: str) -> dict:
    import datetime as _dt
    out = {k: cand.get(k) for k in ("id", "name", "type", "parent",
                                    "county_name", "district_name", "lat", "lon",
                                    "score", "via")}
    out["source"] = source
    out["ts"] = _dt.datetime.utcnow().isoformat() + "Z"
    return out


def _parse_col_ranges(col_filter: Optional[str]):
    """1-indexed column spec → predicate on super_column (1-indexed). None = all."""
    if not col_filter or not col_filter.strip():
        return None
    ranges = []
    for part in col_filter.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            if "-" in part:
                lo_s, _, hi_s = part.partition("-")
                lo = int(lo_s) if lo_s.strip() else 1
                hi = int(hi_s) if hi_s.strip() else None
                ranges.append((lo, hi))
            else:
                v = int(part); ranges.append((v, v))
        except ValueError:
            pass
    return ranges


class AuthorityBatch(BaseModel):
    stems:       List[str] = []            # ordered page stems to consider
    pattern:     Optional[str] = None      # 1/0 cyclic over stems; 1 = process
    col_filter:  Optional[str] = None      # 1-indexed lattice columns / ranges
    name:        Optional[str] = None      # authority filename
    type:        Optional[str] = None      # restrict to one entity type
    layer:       str = "best_llm"          # which text layer to read
    overwrite:   bool = True               # re-resolve auto/empty (human always kept)
    use_context: bool = True               # use each page's per-table county/district context


@app.post("/api/authority/batch")
def api_authority_batch(folder: str = Query(...), body: AuthorityBatch = ...):
    """Resolve authority entities across many pages/columns in one server-side
    pass. Ditto marks inherit the entity above (per column, reading order);
    empty/1-char cells are skipped; existing manual (human) picks are kept."""
    d = _resolve_folder(folder)
    from collections import defaultdict

    # Page selection: cyclic 1/0 pattern over the given ordered stems.
    stems = [s for s in (body.stems or []) if s]
    if not stems:
        stems = [jf.stem for jf in sorted(d.glob("*.json"), key=lambda p: _page_sort_key(p.stem))]
    bits = [1 if p.strip() == "1" else 0 for p in (body.pattern or "").split(",") if p.strip() != ""]
    if not bits:
        bits = [1]
    pages = [s for i, s in enumerate(stems) if bits[i % len(bits)] == 1]

    col_ranges = _parse_col_ranges(body.col_filter)

    def col_allowed(sc):
        if col_ranges is None:
            return True
        return any(lo <= sc <= (hi if hi is not None else sc) for lo, hi in col_ranges)

    layer = body.layer or "best_llm"
    totals = dict(resolved=0, ditto=0, kept=0, skipped=0, nomatch=0)
    per_page = []

    for stem in pages:
        jf = d / f"{stem}.json"
        if not jf.exists():
            continue
        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except Exception:
            continue
        shapes = data.get("shapes", [])
        ctx_map = (data.get("flags") or {}).get("authority_context") or {}
        pc = dict(resolved=0, ditto=0, kept=0, skipped=0, nomatch=0)

        # Group lattice cells by (table, super_column) within the column filter
        cols = defaultdict(list)
        for sh in shapes:
            sr, sc = sh.get("super_row"), sh.get("super_column")
            if sr is None or sc is None or not col_allowed(int(sc)):
                continue
            cols[(sh.get("table") or 0, int(sc))].append((int(sr), sh))

        changed = False
        for (tbl, sc), items in cols.items():
            items.sort(key=lambda x: x[0])                 # reading order, top→bottom
            parent = None
            if body.use_context:
                c = ctx_map.get(str(tbl)) or {}
                parent = c.get("district") or c.get("county")
            # `last` carries the entity for ditto inheritance, but ONLY within a
            # single lattice cell's internal rows — a ditto in a cell's top row
            # must not inherit from a different cell. Reset per cell below.
            last = None

            def handle(container, text):
                nonlocal last, changed
                cur = container.get("authority")
                if cur and cur.get("source") == "human":
                    pc["kept"] += 1; last = cur; return
                if _auth_is_ditto(text):
                    if last:
                        container["authority"] = {**last, "source": "auto", "via": "ditto",
                                                  "ts": _auth_obj_from_cand({}, "auto")["ts"]}
                        pc["ditto"] += 1; changed = True
                    else:
                        if container.pop("authority", None) is not None: changed = True
                        pc["skipped"] += 1
                    return
                if not _auth_resolvable(text):
                    if container.pop("authority", None) is not None: changed = True
                    pc["skipped"] += 1; return
                if not body.overwrite and cur:
                    pc["kept"] += 1; last = cur; return
                cands = _authority_match(text, body.type, parent, 1, body.name)
                # Reject low-similarity tops: no real string match → don't guess
                # (e.g. don't grab a random village just because of the context).
                if not cands or cands[0].get("score", 0) < _AUTH_MIN_ACCEPT:
                    if container.pop("authority", None) is not None: changed = True
                    pc["nomatch"] += 1; return
                container["authority"] = _auth_obj_from_cand(cands[0], "auto")
                last = container["authority"]; pc["resolved"] += 1; changed = True

            for _sr, sh in items:
                last = None                          # ditto never crosses cells
                rows = (sh.get("row_struct") or {}).get("rows") or []
                if rows:
                    for r in rows:
                        handle(r, _auth_layer_text(r.get("human"), r.get("ocr"),
                                                   r.get("llm"), r.get("pdf"), layer))
                    # keep flat fields / row_struct in sync after edits
                    sh["row_struct"]["rows"] = rows
                else:
                    handle(sh, _auth_layer_text(
                        (sh.get("human_output") or {}).get("human_corrected_text"),
                        ((sh.get("tesseract_output") or {}).get("ocr_text") or
                         (sh.get("easyocr_output") or {}).get("ocr_text")),
                        (sh.get("openai_output") or {}).get("response"),
                        sh.get("pdf_text"), layer))

        if changed:
            _write_json(jf, data)
        for k in totals:
            totals[k] += pc[k]
        per_page.append({"stem": stem, **pc})

    return {"ok": True, "pages": len(pages), "totals": totals, "per_page": per_page,
            "authority": body.name or _AUTH_DEFAULT_FILE}


# ---------------------------------------------------------------------------
# Authority — unresolved worklist, apply-by-string, alias promotion, LLM pick
# ---------------------------------------------------------------------------

def _auth_select_pages(d, stems, pattern):
    """Ordered stems filtered by the cyclic 1/0 pattern (same as the batch op)."""
    stems = [s for s in (stems or []) if s]
    if not stems:
        stems = [jf.stem for jf in sorted(d.glob("*.json"), key=lambda p: _page_sort_key(p.stem))]
    bits = [1 if p.strip() == "1" else 0 for p in (pattern or "").split(",") if p.strip() != ""]
    if not bits:
        bits = [1]
    return [s for i, s in enumerate(stems) if bits[i % len(bits)] == 1]


def _auth_iter_units(shapes, col_allowed, layer):
    """Yield (shape_idx, container, row_n|None, text) for every lattice cell /
    internal row within the column filter, in reading order per column."""
    from collections import defaultdict
    cols = defaultdict(list)
    for i, sh in enumerate(shapes):
        sr, sc = sh.get("super_row"), sh.get("super_column")
        if sr is None or sc is None or not col_allowed(int(sc)):
            continue
        cols[(sh.get("table") or 0, int(sc))].append((int(sr), i, sh))
    for key in sorted(cols):
        items = cols[key]
        items.sort(key=lambda x: x[0])
        for _sr, i, sh in items:
            rows = (sh.get("row_struct") or {}).get("rows") or []
            if rows:
                for r in rows:
                    yield i, r, r.get("n"), _auth_layer_text(
                        r.get("human"), r.get("ocr"), r.get("llm"), r.get("pdf"), layer)
            else:
                yield i, sh, None, _auth_layer_text(
                    (sh.get("human_output") or {}).get("human_corrected_text"),
                    ((sh.get("tesseract_output") or {}).get("ocr_text") or
                     (sh.get("easyocr_output") or {}).get("ocr_text")),
                    (sh.get("openai_output") or {}).get("response"),
                    sh.get("pdf_text"), layer)


def _auth_col_pred(col_filter):
    ranges = _parse_col_ranges(col_filter)
    if ranges is None:
        return lambda sc: True
    return lambda sc: any(lo <= sc <= (hi if hi is not None else sc) for lo, hi in ranges)


class AuthorityScanBody(BaseModel):
    stems:      List[str] = []
    pattern:    Optional[str] = None
    col_filter: Optional[str] = None
    name:       Optional[str] = None       # authority filename
    type:       Optional[str] = None       # restrict candidate lookup to one type
    layer:      str = "best_llm"


@app.post("/api/authority/worklist")
def api_authority_worklist(folder: str = Query(...), body: AuthorityScanBody = ...):
    """Distinct unresolved strings across the selected pages/columns, grouped by
    folded form and sorted by frequency — fix the most common ones first, once."""
    from collections import Counter
    d = _resolve_folder(folder)
    index = _load_authority(body.name)
    strip = index.get("strip") or set()
    pages = _auth_select_pages(d, body.stems, body.pattern)
    col_allowed = _auth_col_pred(body.col_filter)

    groups: dict = {}   # fold -> {"count", "texts": Counter, "locations": []}
    for stem in pages:
        jf = d / f"{stem}.json"
        if not jf.exists():
            continue
        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except Exception:
            continue
        for i, container, row_n, text in _auth_iter_units(data.get("shapes", []), col_allowed, body.layer):
            if container.get("authority"):
                continue
            if _auth_is_ditto(text) or not _auth_resolvable(text):
                continue
            fold = _auth_strip_status(_auth_fold(text), strip)
            if not fold:
                continue
            g = groups.setdefault(fold, {"count": 0, "texts": Counter(), "locations": []})
            g["count"] += 1
            g["texts"][text.strip()] += 1
            if len(g["locations"]) < 20:
                g["locations"].append({"stem": stem, "idx": i, "row": row_n})

    ordered = sorted(groups.items(), key=lambda kv: -kv[1]["count"])[:500]
    out = []
    for fold, g in ordered:
        display = g["texts"].most_common(1)[0][0]
        item = {"fold": fold, "text": display, "count": g["count"],
                "locations": g["locations"]}
        # Prefill candidates for the most frequent strings (one match per
        # distinct string — cheap thanks to the cached index).
        if len(out) < 200:
            item["candidates"] = _authority_match(display, body.type, None, 5, body.name)
        out.append(item)
    return {"groups": out, "distinct": len(groups),
            "total_unresolved": sum(g["count"] for g in groups.values()),
            "pages": len(pages), "authority": body.name or _AUTH_DEFAULT_FILE}


class AuthorityApplyString(BaseModel):
    stems:      List[str] = []
    pattern:    Optional[str] = None
    col_filter: Optional[str] = None
    name:       Optional[str] = None
    layer:      str = "best_llm"
    fold:       str                        # target folded string from the worklist
    entity_id:  str                        # chosen entity


@app.post("/api/authority/apply_string")
def api_authority_apply_string(folder: str = Query(...), body: AuthorityApplyString = ...):
    """Resolve EVERY unresolved cell/row whose folded text equals `fold` to the
    chosen entity — one human decision applied to all occurrences. Written with
    source='human' (it is a human decision), so batches keep it."""
    d = _resolve_folder(folder)
    index = _load_authority(body.name)
    strip = index.get("strip") or set()
    v = index["by_id"].get(body.entity_id)
    if not v:
        raise HTTPException(status_code=404, detail=f"Entity not found: {body.entity_id}")
    proto = _auth_obj_from_cand({**v, "score": 100.0, "via": "worklist"}, "human")

    pages = _auth_select_pages(d, body.stems, body.pattern)
    col_allowed = _auth_col_pred(body.col_filter)
    applied, pages_changed = 0, 0
    for stem in pages:
        jf = d / f"{stem}.json"
        if not jf.exists():
            continue
        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except Exception:
            continue
        changed = False
        for i, container, row_n, text in _auth_iter_units(data.get("shapes", []), col_allowed, body.layer):
            if container.get("authority"):
                continue
            if _auth_is_ditto(text) or not _auth_resolvable(text):
                continue
            if _auth_strip_status(_auth_fold(text), strip) != body.fold:
                continue
            container["authority"] = dict(proto)
            applied += 1
            changed = True
        if changed:
            _write_json(jf, data)
            pages_changed += 1
    return {"ok": True, "applied": applied, "pages_changed": pages_changed,
            "entity": {"id": v["id"], "name": v["name"]}}


@app.post("/api/authority/alias_candidates")
def api_authority_alias_candidates(folder: str = Query(...), body: AuthorityScanBody = ...):
    """Strings a human resolved to an entity that are NOT yet a name/alias of
    that entity in the authority file — candidates for alias promotion, so
    every future project matches them automatically."""
    from collections import Counter
    d = _resolve_folder(folder)
    index = _load_authority(body.name)
    strip = index.get("strip") or set()
    known: dict = {}                       # id -> set of known folds
    for c in index["pool_all"]:
        known.setdefault(c["id"], set()).add(c["fold"])

    pages = _auth_select_pages(d, body.stems, body.pattern)
    col_allowed = _auth_col_pred(body.col_filter)
    cands: dict = {}                       # (id, fold) -> {"count", "texts": Counter, "name"}
    for stem in pages:
        jf = d / f"{stem}.json"
        if not jf.exists():
            continue
        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except Exception:
            continue
        for i, container, row_n, text in _auth_iter_units(data.get("shapes", []), col_allowed, body.layer):
            a = container.get("authority")
            if not a or a.get("source") != "human" or a.get("via") == "ditto":
                continue
            if _auth_is_ditto(text) or not _auth_resolvable(text):
                continue
            eid = a.get("id")
            if eid not in known:           # belongs to a different authority
                continue
            fold = _auth_strip_status(_auth_fold(text), strip)
            if not fold or fold in known[eid]:
                continue
            c = cands.setdefault((eid, fold), {"count": 0, "texts": Counter(),
                                               "name": a.get("name")})
            c["count"] += 1
            c["texts"][text.strip()] += 1

    out = [{"id": eid, "entity_name": c["name"],
            "alias": c["texts"].most_common(1)[0][0], "count": c["count"]}
           for (eid, fold), c in sorted(cands.items(), key=lambda kv: -kv[1]["count"])[:300]]
    return {"candidates": out, "authority": body.name or _AUTH_DEFAULT_FILE}


class AuthorityPromote(BaseModel):
    name:    Optional[str] = None
    aliases: List[dict] = []               # [{id, alias}]


@app.post("/api/authority/promote_aliases")
def api_authority_promote_aliases(body: AuthorityPromote = ...):
    """Append human-confirmed aliases to the authority FILE (source
    'econai_confirmed'). The file is git-tracked — review the diff before
    committing. The matcher picks the change up automatically (mtime cache)."""
    fname = body.name or _AUTH_DEFAULT_FILE
    path = AUTHORITIES_DIR / fname
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Authority file not found: {fname}")
    data = json.loads(path.read_text(encoding="utf-8"))
    by_id = {e.get("id"): e for e in data.get("entities", [])}
    added = 0
    for a in body.aliases:
        e = by_id.get(a.get("id"))
        alias = (a.get("alias") or "").strip()
        if not e or not alias:
            continue
        existing = {_auth_norm(x.get("name", "")) for x in (e.get("aliases") or [])}
        existing.add(_auth_norm(e.get("name") or ""))
        for sl in (e.get("slices") or []):
            existing.add(_auth_norm(sl.get("name") or ""))
        if _auth_norm(alias) in existing:
            continue
        e.setdefault("aliases", []).append({"name": alias, "source": "econai_confirmed"})
        added += 1
    if added:
        counts = data.get("counts")
        if isinstance(counts, dict) and "aliases" in counts:
            counts["aliases"] = counts["aliases"] + added
        _write_json(path, data)
    return {"ok": True, "added": added, "file": fname}


class AuthorityLlmPick(BaseModel):
    stem:       str
    idx:        int
    text:       str
    model:      str = "gpt-4o-mini"
    candidates: List[dict] = []            # [{id, name, type, county_name, district_name}]


@app.post("/api/authority/llm_pick")
def api_authority_llm_pick(folder: str = Query(...), body: AuthorityLlmPick = ...):
    """Disambiguate a near-tie: send the cell image + the candidate entities to
    the LLM and return the id it picks (or none)."""
    import base64, io as _io
    from PIL import Image as PILImage

    if not body.candidates:
        raise HTTPException(status_code=400, detail="No candidates given")
    d = _resolve_folder(folder)
    img_path = _find_image(d, body.stem)
    jf = d / f"{body.stem}.json"
    if img_path is None or not jf.exists():
        raise HTTPException(status_code=404, detail="Page not found")
    shapes = json.loads(jf.read_text(encoding="utf-8")).get("shapes", [])
    if body.idx < 0 or body.idx >= len(shapes):
        raise HTTPException(status_code=400, detail="Shape index out of range")

    img = PILImage.open(str(img_path)).convert("RGB")
    x1, y1, x2, y2 = _shape_bbox(shapes[body.idx])
    pad = 6
    crop = img.crop((max(0, int(x1) - pad), max(0, int(y1) - pad),
                     min(img.width, int(x2) + pad), min(img.height, int(y2) + pad)))
    if crop.height < 48:
        scale = 48 / max(1, crop.height)
        crop = crop.resize((max(1, int(crop.width * scale)), 48), PILImage.LANCZOS)
    buf = _io.BytesIO(); crop.save(buf, format="JPEG", quality=92)
    b64 = base64.b64encode(buf.getvalue()).decode()

    lines = []
    for c in body.candidates:
        ctx = ", ".join(x for x in (c.get("type"), c.get("county_name"),
                                    c.get("district_name")) if x)
        lines.append(f"{c.get('id')}: {c.get('name')} ({ctx})")
    prompt = (
        "The image is a cell from a historical Hungarian document. Its OCR "
        f"reading is: \"{body.text}\".\n"
        "Which of these gazetteer entries does the cell refer to?\n"
        + "\n".join(lines)
        + "\nReply with ONLY the id of the best match, or NONE if none of them fit."
    )
    content = [{"type": "text", "text": prompt},
               {"type": "image_url",
                "image_url": {"url": f"data:image/jpeg;base64,{b64}", "detail": "high"}}]
    try:
        client = _make_llm_client(body.model)
        resp = _llm_complete(client, body.model, [{"role": "user", "content": content}], 256)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc))
    raw = (resp.choices[0].message.content or "").strip()
    choice = next((c.get("id") for c in body.candidates
                   if c.get("id") and c.get("id") in raw), None)
    return {"choice": choice, "raw": raw}


class PrepareRequest(BaseModel):
    max_iter:      Optional[int]   = None
    base_lr:       Optional[float] = None
    ims_per_batch: Optional[int]   = None
    num_workers:   Optional[int]   = None


@app.post("/api/project/{name}/prepare")
def api_prepare(name: str, body: Optional[PrepareRequest] = None):
    """Convert annotated LabelMe JSONs → COCO JSON + generate training scripts.
    Optional solver params (max_iter / base_lr / ims_per_batch / num_workers) are
    hand-edited in the dashboard and baked into the generated train.sh."""
    from app.coco_convert import prepare_training_data
    body = body or PrepareRequest()
    try:
        cfg  = load_config(name)
        pdir = project_dir(name)
        result = prepare_training_data(
            project_name    = name,
            ann_dir         = pdir / "annotations",
            labels          = cfg["labels"],
            intermediate_dir= pdir / "intermediate",
            base_yaml_path  = BASE_YAML,
            max_iter        = body.max_iter      or 2000,
            base_lr         = body.base_lr       or 0.00125,
            ims_per_batch   = body.ims_per_batch or 2,
            num_workers     = 2 if body.num_workers is None else body.num_workers,
        )
        return {"ok": True, **result}
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Routes — train / infer with SSE streaming
# ---------------------------------------------------------------------------

class TrainRequest(BaseModel):
    passphrase:     Optional[str] = None
    keep_container: bool = False        # leave the container running after the job

class InferRequest(BaseModel):
    passphrase:          Optional[str] = None
    skip_image_upload:   bool = False
    keep_container:      bool = False


_SSE_DONE = object()


def _safe_next(it):
    """Return next item, or _SSE_DONE sentinel — never raises.

    An exception escaping a streaming generator after the response has started
    kills the connection with no message (ASGI: "response already started"),
    so any failure is converted into a {"type": "error"} SSE event the client
    can show. The generator is dead after raising, so the next call returns
    _SSE_DONE and the stream ends cleanly."""
    try:
        return next(it)
    except StopIteration:
        return _SSE_DONE
    except HTTPException as exc:
        return json.dumps({"type": "error", "error": str(exc.detail)})
    except Exception as exc:
        return json.dumps({"type": "error", "error": f"{type(exc).__name__}: {exc}"})


async def _sse_stream(gen):
    """Wrap a sync generator as an SSE response."""
    import asyncio
    import json as _json

    async def event_gen():
        loop = asyncio.get_event_loop()
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            it = iter(gen)
            while True:
                try:
                    line = await loop.run_in_executor(pool, _safe_next, it)
                    if line is _SSE_DONE:
                        yield "data: {\"done\": true}\n\n"
                        break
                    payload = _json.dumps({"line": line.rstrip("\n")})
                    yield f"data: {payload}\n\n"
                except Exception as e:
                    yield f"data: {_json.dumps({'error': str(e)})}\n\n"
                    break

    return StreamingResponse(event_gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache",
                                      "X-Accel-Buffering": "no"})


def _push_training_data_gen(name: str, srv: dict, passphrase: str,
                            skip_images: bool = False):
    """Generator: push images + COCO JSON + config + scripts, yielding progress lines."""
    from app import ssh_ops
    pdir   = project_dir(name)
    inter  = pdir / "intermediate"
    remote = srv["remote_path"].rstrip("/")

    yield f"[push] remote_path = {remote}"
    yield f"[push] Connecting to {srv['host']} as {srv['user']}..."
    c = ssh_ops._client(srv["host"], srv["user"], srv["key_path"], passphrase)
    sftp = c.open_sftp()
    try:
        dirs = [
            f"{remote}/{name}/images",
            f"{remote}/layout-model-training/configs/{name}",
            f"{remote}/layout-model-training/tools",
            f"{remote}/layout-model-training/scripts",
        ]
        for d in dirs:
            yield f"[push] mkdir -p {d}"
            ssh_ops._sftp_mkdir_p(sftp, d)

        IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff"}
        images = [p for p in (pdir / "annotations").iterdir()
                  if p.suffix.lower() in IMAGE_EXTS]
        total = len(images)
        if skip_images:
            yield f"[push] Skipping image upload ({total} images assumed already on server)."
        else:
            yield f"[push] Uploading {total} image(s) → {remote}/{name}/images/"
            for i, img in enumerate(images, 1):
                dest = f"{remote}/{name}/images/{img.name}"
                sftp.put(str(img), dest)
                if i % 5 == 0 or i == total:
                    yield f"[push]   {i}/{total}  {img.name} → {dest}"

        dest = f"{remote}/{name}/annotations.json"
        yield f"[push] {inter/'annotations.json'} → {dest}"
        sftp.put(str(inter / "annotations.json"), dest)

        cfg_local = inter / "configs" / name / "fast_rcnn_R_50_FPN_3x.yaml"
        dest = f"{remote}/layout-model-training/configs/{name}/fast_rcnn_R_50_FPN_3x.yaml"
        yield f"[push] {cfg_local} → {dest}"
        sftp.put(str(cfg_local), dest)

        dest = f"{remote}/layout-model-training/tools/infer_layout.py"
        yield f"[push] infer_layout.py → {dest}"
        sftp.put(str(Path(__file__).parent / "infer_layout.py"), dest)

        def sftp_put_lf(local: Path, remote_dest: str):
            """Upload a text file with LF line endings (strip Windows CRLF)."""
            content = local.read_bytes().replace(b"\r\n", b"\n").replace(b"\r", b"\n")
            with sftp.open(remote_dest, "wb") as fh:
                fh.write(content)

        dest_train = f"{remote}/layout-model-training/scripts/{name}.sh"
        dest_infer = f"{remote}/layout-model-training/scripts/{name}_infer.sh"
        yield f"[push] train.sh → {dest_train}"
        sftp_put_lf(inter / "train.sh", dest_train)
        yield f"[push] infer.sh → {dest_infer}"
        sftp_put_lf(inter / "infer.sh", dest_infer)

        yield "[push] All files uploaded successfully."
        yield f"[push] Verifying: ls {remote}/layout-model-training/scripts/"
        _, stdout, _ = c.exec_command(f"ls -1 {remote}/layout-model-training/scripts/")
        for ln in stdout.read().decode().splitlines():
            yield f"[push]   {ln}"
    finally:
        sftp.close()
        c.close()


# ── Detached-job plumbing shared by Train and Fine-tune ──────────────────────
# The .pid file lives on the /workspace bind mount, so it SURVIVES container
# restarts — and container PIDs are small numbers that get reused. A bare
# `kill -0 <old pid>` therefore eventually matches some unrelated process and
# the endpoint "re-attaches" to a dead log forever (the fine-tune-blocked-
# until-container-killed bug). Three defenses:
#   1. the launch wrapper DELETES the pid file when the job exits,
#   2. the running-probe verifies /proc/<pid>/cmdline names our script,
#   3. the tail loop / still-running probe treat a missing pid file as done.

def _job_running_cmd(container: str, pid_path: str, script_path: str) -> str:
    import posixpath as pp
    script_name = pp.basename(script_path)
    return (
        f"docker exec {container} bash -c "
        f"'[ -f {pid_path} ] && pid=$(cat {pid_path}) && "
        f"kill -0 $pid 2>/dev/null && "
        f"tr \"\\0\" \" \" </proc/$pid/cmdline 2>/dev/null | grep -qF \"{script_name}\" "
        f"&& echo RUNNING || echo STOPPED' 2>/dev/null || echo STOPPED"
    )

def _job_launch_cmd(container: str, script_path: str, log_path: str,
                    pid_path: str, self_stop: bool = True) -> str:
    # self_stop: after the job exits, touch the sentinel that makes the
    # container's keep-alive loop terminate → the container stops itself and
    # frees the GPU even if no browser/SSE stream survived to do the cleanup.
    finale = "; touch /tmp/dedust_selfstop" if self_stop else ""
    return (
        f"docker start {container} && "
        f"docker exec {container} bash -c "
        f"'mkdir -p /workspace/layout-model-training/logs && "
        f"nohup bash -c \"bash {script_path} >{log_path} 2>&1; rm -f {pid_path}{finale}\" "
        f">/dev/null 2>&1 & echo $! >{pid_path}; "
        f"echo LAUNCHED:$(cat {pid_path})'"
    )

def _job_tail_cmd(container: str, log_path: str, pid_path: str,
                  done_msg: str) -> str:
    return (
        f"docker exec {container} bash -c '"
        f"tail -n 0 -f {log_path} & TAIL=$!; "
        f"pid=$(cat {pid_path} 2>/dev/null); "
        f"[ -n \"$pid\" ] && while [ -f {pid_path} ] && kill -0 $pid 2>/dev/null; "
        f"do sleep 5; done; "
        f"sleep 2; kill $TAIL 2>/dev/null; "
        f"echo \"{done_msg}\"'"
    )

def _job_still_running_cmd(container: str, pid_path: str) -> str:
    return (
        f"docker exec {container} bash -c '"
        f"{{ [ -f {pid_path} ] && pid=$(cat {pid_path}) && "
        f"kill -0 $pid 2>/dev/null && echo RUNNING; }} || echo DONE'"
        f" 2>/dev/null || echo DONE"
    )


def _job_instant_death_check(_quick, launch_result: str, container: str,
                             log_path: str, tag: str):
    """A healthy launch echoes LAUNCHED:<pid>. An empty pid means the job died
    within milliseconds (bad path, missing scaffolding, …) and its wrapper
    already removed the pid file — surface the log instead of a silent
    'complete'."""
    if re.search(r"LAUNCHED:\d+", launch_result or ""):
        return
    yield f"[{tag}] ⚠ Launch returned no PID — the job likely died instantly. Log tail:"
    try:
        tail = _quick(f"docker exec {container} tail -n 30 {log_path} 2>/dev/null || true")
        for ln in (tail or "").splitlines():
            yield f"[{tag}]   {ln}"
    except Exception as e:
        yield f"[{tag}]   (could not read log: {e})"


def _gpu_busy_warning(_quick, tag: str):
    """Preflight: warn (don't block) when other processes already hold the GPU.
    On shared hosts in exclusive-process mode a second CUDA client dies with
    the cryptic 'No CUDA GPUs are available' — this names the real culprit."""
    try:
        procs = _quick("nvidia-smi --query-compute-apps=pid,process_name "
                       "--format=csv,noheader 2>/dev/null")
    except Exception:
        return
    procs = (procs or "").strip()
    if procs:
        yield f"[{tag}] ⚠ GPU already in use by other process(es):"
        for ln in procs.splitlines()[:5]:
            yield f"[{tag}]     {ln.strip()}"
        yield (f"[{tag}] ⚠ If the job fails with 'No CUDA GPUs are available', "
               f"another container holds the GPU (exclusive mode) — "
               f"`docker ps` on the server, stop the culprit, retry.")


@app.post("/api/project/{name}/train")
async def api_train(name: str, body: TrainRequest = TrainRequest()):
    """Push data to server then run training inside Docker. Streams log via SSE.

    Training is launched detached (nohup + disown) so it survives webapp
    restarts and browser disconnects.  The log is streamed via tail -f;
    closing the browser does NOT kill training.  Clicking Train again while
    training is running re-attaches to the existing log instead of starting
    a second job.
    """
    from app import ssh_ops

    try:
        cfg = load_config(name)
        srv = _server_cfg(name)
        passphrase = body.passphrase or srv.get("passphrase")

        pdir  = project_dir(name)
        inter = pdir / "intermediate"

        script_path = f"/workspace/layout-model-training/scripts/{name}.sh"
        log_path    = f"/workspace/layout-model-training/logs/{name}_train.log"
        pid_path    = f"/workspace/layout-model-training/logs/{name}_train.pid"

        def _quick(cmd: str) -> str:
            """Run a non-streaming SSH command, return stdout."""
            c = ssh_ops._client(srv["host"], srv["user"], srv["key_path"], passphrase)
            _, out, _ = c.exec_command(cmd)
            result = out.read().decode(errors="replace").strip()
            c.close()
            return result

        def full_gen():
            # ── Check if training is already running ──────────────────────────
            already_running = False
            try:
                status = _quick(_job_running_cmd(_train_container(srv),
                                                 pid_path, script_path))
                already_running = "RUNNING" in status
            except Exception:
                already_running = False

            if already_running:
                yield "[train] Training already running — re-attaching to log..."
            else:
                # ── Validate + push data ──────────────────────────────────────
                if not (inter / "annotations.json").exists():
                    yield "ERROR: Run 'Prepare training data' first."
                    return
                yield from _push_training_data_gen(name, srv, passphrase)

                # ── Launch detached training ──────────────────────────────────
                yield from _gpu_busy_warning(_quick, "train")
                yield "[train] Launching detached training (safe to close browser)..."
                launch_result = _quick(_job_launch_cmd(_train_container(srv),
                                                       script_path, log_path, pid_path,
                                                       self_stop=not body.keep_container))
                yield f"[train] {launch_result.strip()}"
                yield from _job_instant_death_check(_quick, launch_result,
                                                    _train_container(srv), log_path, "train")

            # ── Stream log, stop when process exits ───────────────────────────
            yield "[train] Streaming log — closing browser won't stop training..."
            tail_cmd = _job_tail_cmd(_train_container(srv), log_path, pid_path,
                                     "[Training complete — process has exited]")
            yield from ssh_ops.stream_command(
                srv["host"], srv["user"], srv["key_path"], tail_cmd, passphrase)

            # Stream ended naturally → the training process has exited (a browser
            # disconnect closes this generator before reaching here, so a detached
            # job is never killed). Verify the PID is gone, then stop the
            # container to free the GPU / host resources.
            if not body.keep_container:
                still = _quick(_job_still_running_cmd(_train_container(srv), pid_path))
                if "RUNNING" in still:
                    yield "[train] Stream ended but training still running — container left up."
                else:
                    yield f"[train] Training finished — stopping container '{_train_container(srv)}'..."
                    _quick(f"docker stop {_train_container(srv)}")
                    yield f"[train] Container '{_train_container(srv)}' stopped."

        return await _sse_stream(full_gen())

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _push_infer_data_gen(name: str, srv: dict, passphrase: str, skip_images: bool = False):
    """Generator: push images + scripts + config for inference (no annotations.json needed).

    Uses predict_remote_path (if set) as the Docker-container root for placing
    the infer script — necessary when the predict container mounts a parent
    directory compared to the training container.
    skip_images=True skips the (slow) image upload and only pushes config/scripts.
    """
    from app import ssh_ops
    import posixpath as pp
    pdir         = project_dir(name)
    inter        = pdir / "intermediate"
    remote       = srv["remote_path"].rstrip("/")        # where data lives on host
    predict_root = srv.get("predict_remote_path", remote).rstrip("/")  # predict container ws root on host

    yield f"[push] remote_path (data)    = {remote}"
    yield f"[push] predict_remote_path   = {predict_root}"
    yield f"[push] Connecting to {srv['host']} as {srv['user']}..."
    c = ssh_ops._client(srv["host"], srv["user"], srv["key_path"], passphrase)

    # Detect the actual bind mount so we get the correct ws_prefix regardless
    # of how remote_path / predict_remote_path are configured.
    _, _out, _ = c.exec_command(
        f"docker inspect {_predict_container(srv)} "
        "--format '{{range .Mounts}}{{.Source}}:{{.Destination}}{{println}}{{end}}'"
    )
    container_ws_host = predict_root
    for _bind in _out.read().decode().strip().splitlines():
        _parts = _bind.strip().split(':')
        if len(_parts) >= 2 and _parts[1].rstrip('/') == '/workspace':
            container_ws_host = _parts[0].rstrip('/')
            break
    if remote.startswith(container_ws_host + "/"):
        ws_prefix = remote[len(container_ws_host) + 1:]
    else:
        ws_prefix = ""

    yield f"[push] container /workspace ← {container_ws_host}"
    yield f"[push] container ws prefix   = '{ws_prefix}'"
    sftp = c.open_sftp()
    try:
        # All paths that must be visible inside the predict container go under predict_root.
        # Images live under remote (data workspace, mounted into both containers).
        dirs = [
            f"{remote}/{name}/images",
            f"{predict_root}/layout-model-training/configs/{name}",
            f"{predict_root}/layout-model-training/tools",
            f"{predict_root}/layout-model-training/scripts",
        ]
        for d in dirs:
            yield f"[push] mkdir -p {d}"
            ssh_ops._sftp_mkdir_p(sftp, d)

        IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff"}
        if skip_images:
            images = [p for p in (pdir / "annotations").iterdir()
                      if p.suffix.lower() in IMAGE_EXTS]
            yield f"[push] Skipping image upload ({len(images)} images already on server)."
        else:
            images = [p for p in (pdir / "annotations").iterdir()
                      if p.suffix.lower() in IMAGE_EXTS]
            total = len(images)
            yield f"[push] Uploading {total} image(s) → {remote}/{name}/images/"
            for i, img in enumerate(images, 1):
                dest = f"{remote}/{name}/images/{img.name}"
                sftp.put(str(img), dest)
                if i % 10 == 0 or i == total:
                    yield f"[push]   {i}/{total}  {img.name}"

        cfg_local = inter / "configs" / name / "fast_rcnn_R_50_FPN_3x.yaml"
        dest = f"{predict_root}/layout-model-training/configs/{name}/fast_rcnn_R_50_FPN_3x.yaml"
        yield f"[push] config YAML → {dest}"
        sftp.put(str(cfg_local), dest)

        dest = f"{predict_root}/layout-model-training/tools/infer_layout.py"
        yield f"[push] infer_layout.py → {dest}"
        sftp.put(str(Path(__file__).parent / "infer_layout.py"), dest)

        def sftp_put_lf(local: Path, remote_dest: str):
            content = local.read_bytes().replace(b"\r\n", b"\n").replace(b"\r", b"\n")
            with sftp.open(remote_dest, "wb") as fh:
                fh.write(content)

        def sftp_put_infer_sh(local: Path, remote_dest: str):
            """Upload infer.sh, patching all /workspace/ paths to include ws_prefix.
            Every path in the generated infer.sh lives under the prefixed workspace,
            so we replace all /workspace/ occurrences at once.
            """
            content = local.read_bytes().replace(b"\r\n", b"\n").replace(b"\r", b"\n")
            if ws_prefix:
                content = content.replace(b"/workspace/",
                                          f"/workspace/{ws_prefix}/".encode())
            with sftp.open(remote_dest, "wb") as fh:
                fh.write(content)

        dest_infer = f"{predict_root}/layout-model-training/scripts/{name}_infer.sh"
        yield f"[push] infer.sh → {dest_infer}" + (f" (ws_prefix='{ws_prefix}')" if ws_prefix else "")
        sftp_put_infer_sh(inter / "infer.sh", dest_infer)

        yield "[push] All files uploaded successfully."
        yield f"[push] Verifying scripts dir ({predict_root}/layout-model-training/scripts/):"
        _, stdout, _ = c.exec_command(f"ls -la {predict_root}/layout-model-training/scripts/")
        for ln in stdout.read().decode().splitlines():
            yield f"[push]   {ln}"
    finally:
        sftp.close()
        c.close()


@app.post("/api/project/{name}/infer")
async def api_infer(name: str, body: InferRequest = InferRequest()):
    """Run inference on the server. Streams log via SSE."""
    from app import ssh_ops
    try:
        srv          = _server_cfg(name)
        passphrase   = body.passphrase or srv.get("passphrase")
        remote       = srv["remote_path"].rstrip("/")
        predict_root = srv.get("predict_remote_path", remote).rstrip("/")

        # Detect bind mount to build the correct in-container script path.
        def _detect_script_path() -> str:
            try:
                c = ssh_ops._client(srv["host"], srv["user"], srv["key_path"], passphrase)
                _, _out, _ = c.exec_command(
                    f"docker inspect {_predict_container(srv)} "
                    "--format '{{range .Mounts}}{{.Source}}:{{.Destination}}{{println}}{{end}}'"
                )
                container_ws_host = predict_root
                for _bind in _out.read().decode().strip().splitlines():
                    _parts = _bind.strip().split(':')
                    if len(_parts) >= 2 and _parts[1].rstrip('/') == '/workspace':
                        container_ws_host = _parts[0].rstrip('/')
                        break
                c.close()
                if remote.startswith(container_ws_host + "/"):
                    pfx = remote[len(container_ws_host) + 1:] + "/"
                else:
                    pfx = ""
                return f"/workspace/{pfx}layout-model-training/scripts/{name}_infer.sh"
            except Exception:
                return f"/workspace/layout-model-training/scripts/{name}_infer.sh"

        def full_gen():
            yield from _push_infer_data_gen(name, srv, passphrase,
                                            skip_images=body.skip_image_upload)
            script_path = _detect_script_path()
            _cn = _predict_container(srv)
            # Verify the script is actually visible inside the container before
            # running — guards against a /workspace mount that doesn't match the
            # upload path (the cross-user clobbering failure mode). The sentinel
            # touch after the script makes the container stop ITSELF when the
            # job ends, even if the browser/SSE stream is long gone.
            _fin = "" if body.keep_container else "; ec=$?; touch /tmp/dedust_selfstop; exit $ec"
            docker_cmd = (
                f"docker start {_cn} && "
                f"if docker exec {_cn} test -f {script_path}; then "
                f"docker exec {_cn} bash -c 'bash {script_path}{_fin}'; "
                f"else echo '[ERROR] {script_path} not found inside container {_cn}. "
                f"Its /workspace mount does not match where files were uploaded. "
                f"Rebuild the predict container in Docker settings so /workspace maps "
                f"to this project remote_path.'; exit 1; fi"
            )
            yield f"[push] Starting Docker inference (script: {script_path})..."
            yield from ssh_ops.stream_command(srv["host"], srv["user"],
                                              srv["key_path"], docker_cmd, passphrase)
            # Stream ended naturally → the foreground inference exec finished.
            yield from _stop_container_gen(srv, passphrase, _predict_container(srv),
                                           "infer", body.keep_container)

        return await _sse_stream(full_gen())

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _push_infer_from_gen(new_name: str, source_name: str,
                         srv: dict, passphrase: str,
                         skip_images: bool = False,
                         threshold: float = 0.3):
    """Generator: run inference on new_name's images using source_name's trained model.

    Identical to _push_infer_data_gen except the --weights and --config paths
    point at the source project, while --images and --output point at the new project.
    The source project's config YAML is expected to already be on the server
    (from a previous prepare+push cycle on that project).
    """
    from app import ssh_ops
    import posixpath as pp

    new_cfg  = load_config(new_name)
    src_cfg  = load_config(source_name)

    # Label compatibility check
    new_labels = new_cfg.get("labels", [])
    src_labels = src_cfg.get("labels", [])
    if set(new_labels) != set(src_labels):
        raise HTTPException(
            status_code=400,
            detail=(f"Label mismatch: '{new_name}' has {new_labels} "
                    f"but '{source_name}' was trained with {src_labels}. "
                    f"Projects must share the same label set."),
        )

    pdir         = project_dir(new_name)
    remote       = srv["remote_path"].rstrip("/")
    predict_root = srv.get("predict_remote_path", remote).rstrip("/")

    if remote.startswith(predict_root + "/"):
        ws_prefix = remote[len(predict_root) + 1:]
    else:
        ws_prefix = ""

    yield f"[infer-from] source model : {source_name}"
    yield f"[infer-from] target project: {new_name}"
    yield f"[infer-from] remote_path   : {remote}"
    yield f"[infer-from] ws_prefix     : '{ws_prefix}'"
    yield f"[infer-from] Connecting to {srv['host']} as {srv['user']}..."

    c    = ssh_ops._client(srv["host"], srv["user"], srv["key_path"], passphrase)
    sftp = c.open_sftp()
    try:
        # Determine the actual host directory that the predicting container mounts as /workspace.
        # This may differ from predict_root/remote if predict_remote_path isn't configured.
        _, _out, _ = c.exec_command(
            f"docker inspect {_predict_container(srv)} "
            "--format '{{range .Mounts}}{{.Source}}:{{.Destination}}{{println}}{{end}}'"
        )
        container_ws_host = predict_root  # fallback
        for _bind in _out.read().decode().strip().splitlines():
            _parts = _bind.strip().split(':')
            if len(_parts) >= 2 and _parts[1].rstrip('/') == '/workspace':
                container_ws_host = _parts[0].rstrip('/')
                break
        # Recompute ws_prefix from the real container mount rather than config assumptions
        if remote.startswith(container_ws_host + "/"):
            ws_prefix = remote[len(container_ws_host) + 1:]
        else:
            ws_prefix = ""
        yield f"[infer-from] container /workspace ← {container_ws_host}"
        yield f"[infer-from] effective ws_prefix  : '{ws_prefix}'"

        # Ensure directories exist for the new project
        dirs = [
            f"{remote}/{new_name}/images",
            f"{remote}/{new_name}/predictions",
            f"{predict_root}/layout-model-training/tools",
            f"{predict_root}/layout-model-training/scripts",
        ]
        for d in dirs:
            yield f"[infer-from] mkdir -p {d}"
            ssh_ops._sftp_mkdir_p(sftp, d)

        # Verify source model weights exist on the server
        host_weights = (f"{remote}/layout-model-training/outputs/{source_name}/"
                        f"fast_rcnn_R_50_FPN_3x/model_final.pth")
        _, out, _ = c.exec_command(f"test -f {host_weights} && echo OK || echo MISSING")
        status = out.read().decode().strip()
        if status != "OK":
            raise HTTPException(
                status_code=404,
                detail=f"Model weights not found on server: {host_weights}\n"
                       f"Run 'Train model' on project '{source_name}' first.",
            )
        yield f"[infer-from] ✓ weights found: {host_weights}"

        # Upload images for the new project
        IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff"}
        images = [p for p in (pdir / "annotations").iterdir()
                  if p.suffix.lower() in IMAGE_EXTS]
        if skip_images:
            yield f"[infer-from] Skipping image upload ({len(images)} images already on server)."
        else:
            total = len(images)
            yield f"[infer-from] Uploading {total} image(s) → {remote}/{new_name}/images/"
            for i, img in enumerate(images, 1):
                sftp.put(str(img), f"{remote}/{new_name}/images/{img.name}")
                if i % 10 == 0 or i == total:
                    yield f"[infer-from]   {i}/{total}  {img.name}"

        # Upload the inference tool
        dest = f"{predict_root}/layout-model-training/tools/infer_layout.py"
        yield f"[infer-from] infer_layout.py → {dest}"
        sftp.put(str(Path(__file__).parent / "infer_layout.py"), dest)

        # Build an infer script: source model weights/config, new project images/output
        pfx = (ws_prefix + "/") if ws_prefix else ""
        config_path  = f"/workspace/{pfx}layout-model-training/configs/{source_name}/fast_rcnn_R_50_FPN_3x.yaml"
        weights_path = f"/workspace/{pfx}layout-model-training/outputs/{source_name}/fast_rcnn_R_50_FPN_3x/model_final.pth"
        images_path  = f"/workspace/{pfx}{new_name}/images"
        output_path  = f"/workspace/{pfx}{new_name}/predictions"
        labels_str   = " ".join(src_labels)

        tool_path = f"/workspace/{pfx}layout-model-training/tools/infer_layout.py"
        script = (
            f"#!/bin/bash\nset -e\n"
            f"echo '=== EconAI: {new_name} inference from {source_name} model ==='\n"
            f"python3 {tool_path} \\\n"
            f"    --config  {config_path} \\\n"
            f"    --weights {weights_path} \\\n"
            f"    --images  {images_path} \\\n"
            f"    --output  {output_path} \\\n"
            f"    --labels  {labels_str} \\\n"
            f"    --threshold {threshold}\n"
            f"echo '=== Inference complete ==='\n"
        ).encode()

        script_dest = f"{predict_root}/layout-model-training/scripts/{new_name}_infer_from_{source_name}.sh"
        yield f"[infer-from] script → {script_dest}"
        with sftp.open(script_dest, "wb") as fh:
            fh.write(script)

        yield "[infer-from] All files uploaded successfully."
    finally:
        sftp.close()
        c.close()

    # Translate the host script path to the container path via the bind mount
    if script_dest.startswith(container_ws_host + "/"):
        container_script = "/workspace/" + script_dest[len(container_ws_host) + 1:]
    else:
        container_script = f"/workspace/{(ws_prefix + '/') if ws_prefix else ''}layout-model-training/scripts/{new_name}_infer_from_{source_name}.sh"
    yield f"[infer-from] container script path: {container_script}"
    yield f"__docker_cmd__:{container_script}"


class InferFromRequest(BaseModel):
    passphrase:        Optional[str] = None
    skip_image_upload: bool = False
    threshold:         float = 0.1
    keep_container:    bool = False


@app.post("/api/project/{name}/infer-from/{source}")
async def api_infer_from(name: str, source: str, body: InferFromRequest = InferFromRequest()):
    """Run inference on name's images using source project's trained model. Streams log via SSE."""
    from app import ssh_ops
    try:
        srv = _server_cfg(name)
        passphrase = body.passphrase or srv.get("passphrase")

        def full_gen():
            docker_cmd = None
            for line in _push_infer_from_gen(name, source, srv,
                                             passphrase, body.skip_image_upload,
                                             body.threshold):
                if line.startswith("__docker_cmd__:"):
                    docker_cmd = line[len("__docker_cmd__:"):]
                else:
                    yield line
            if docker_cmd:
                container_script = docker_cmd
                yield "[infer-from] Starting Docker inference..."
                _cn = _predict_container(srv)
                _fin = "" if body.keep_container else "; ec=$?; touch /tmp/dedust_selfstop; exit $ec"
                cmd = (
                    f"docker start {_cn} && "
                    f"if docker exec {_cn} test -f {container_script}; then "
                    f"docker exec {_cn} bash -c 'bash {container_script}{_fin}'; "
                    f"else echo '[ERROR] {container_script} not found inside container {_cn}. "
                    f"Its /workspace mount does not match where files were uploaded. "
                    f"Rebuild the predict container in Docker settings so /workspace maps "
                    f"to this project remote_path.'; exit 1; fi"
                )
                yield from ssh_ops.stream_command(
                    srv["host"], srv["user"], srv["key_path"], cmd, passphrase)
                yield from _stop_container_gen(srv, passphrase,
                                               _predict_container(srv), "infer-from",
                                               body.keep_container)

        return await _sse_stream(full_gen())

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class FinetuneFromRequest(BaseModel):
    passphrase:        Optional[str] = None
    max_iter:          int = 500          # short: we start from good weights
    base_lr:           float = 0.00025    # gentle: don't wreck what the model knows
    skip_image_upload: bool = False
    keep_container:    bool = False


@app.post("/api/project/{name}/finetune-from/{source}")
async def api_finetune_from(name: str, source: str,
                            body: FinetuneFromRequest = FinetuneFromRequest()):
    """Fine-tune: train `name`'s own model on its (hand-corrected) annotations,
    warm-started from `source`'s trained weights (MODEL.WEIGHTS override)
    instead of the generic COCO backbone. Short + low-LR by default. The result
    lands in outputs/{name}/ — the source model is never touched. Launched
    detached (like Train) so it survives browser disconnects; streams the log.
    """
    from app import ssh_ops
    from app.coco_convert import prepare_training_data

    try:
        try:
            cfg     = load_config(name)
            src_cfg = load_config(source)
        except FileNotFoundError as e:
            raise HTTPException(status_code=404, detail=str(e))
        if set(cfg.get("labels", [])) != set(src_cfg.get("labels", [])):
            raise HTTPException(
                status_code=400,
                detail=(f"Label mismatch: '{name}' has {cfg.get('labels')} but "
                        f"'{source}' was trained with {src_cfg.get('labels')}. "
                        f"Projects must share the same label set."))

        srv        = _server_cfg(name)
        passphrase = body.passphrase or srv.get("passphrase")
        pdir       = project_dir(name)
        inter      = pdir / "intermediate"
        remote     = srv["remote_path"].rstrip("/")

        max_iter = max(1, int(body.max_iter or 500))
        base_lr  = float(body.base_lr or 0.00025)

        # Regenerate COCO + config locally from the CURRENT (corrected)
        # annotations. Standard solver params here so the normal train.sh that
        # gets pushed alongside is not silently rewritten with fine-tune values.
        prepare_training_data(
            project_name     = name,
            ann_dir          = pdir / "annotations",
            labels           = cfg["labels"],
            intermediate_dir = inter,
            base_yaml_path   = BASE_YAML,
        )

        # The fine-tune script: identical to train.sh except MODEL.WEIGHTS
        # points at the source model and the solver is short + gentle.
        src_weights = (f"/workspace/layout-model-training/outputs/{source}/"
                       f"fast_rcnn_R_50_FPN_3x/model_final.pth")
        ft_script = f"""#!/bin/bash
set -e
echo "=== Dedust: fine-tune {name} from {source} model ==="
echo "Running cocosplit..."
cd /workspace/layout-model-training
python3 utils/cocosplit.py \\
    --annotation-path /workspace/{name}/annotations.json \\
    --train            /workspace/{name}/train.json \\
    --test             /workspace/{name}/test.json \\
    --split-ratio      0.8 \\
    --having-annotations

echo "=== Cleaning previous checkpoints of {name} ==="
rm -f /workspace/layout-model-training/outputs/{name}/fast_rcnn_R_50_FPN_3x/*.pth
rm -f /workspace/layout-model-training/outputs/{name}/fast_rcnn_R_50_FPN_3x/last_checkpoint

echo "=== Starting fine-tune (warm start from {source}) ==="
cd /workspace/layout-model-training/tools
python3 train_net.py \\
    --dataset_name          {name}-layout \\
    --json_annotation_train /workspace/{name}/train.json \\
    --image_path_train      /workspace/{name}/images \\
    --json_annotation_val   /workspace/{name}/test.json \\
    --image_path_val        /workspace/{name}/images \\
    --config-file           /workspace/layout-model-training/configs/{name}/fast_rcnn_R_50_FPN_3x.yaml \\
    OUTPUT_DIR  /workspace/layout-model-training/outputs/{name}/fast_rcnn_R_50_FPN_3x/ \\
    MODEL.WEIGHTS {src_weights} \\
    SOLVER.IMS_PER_BATCH 2 \\
    SOLVER.BASE_LR {base_lr} \\
    SOLVER.MAX_ITER {max_iter} \\
    DATALOADER.NUM_WORKERS 2
echo "=== Fine-tune complete ==="
"""

        script_path = f"/workspace/layout-model-training/scripts/{name}_finetune_from_{source}.sh"
        log_path    = f"/workspace/layout-model-training/logs/{name}_ft.log"
        pid_path    = f"/workspace/layout-model-training/logs/{name}_ft.pid"

        def _quick(cmd: str) -> str:
            c = ssh_ops._client(srv["host"], srv["user"], srv["key_path"], passphrase)
            _, out, _ = c.exec_command(cmd)
            result = out.read().decode(errors="replace").strip()
            c.close()
            return result

        def full_gen():
            # ── Re-attach if a fine-tune is already running ────────────────────
            already_running = False
            try:
                status = _quick(_job_running_cmd(_train_container(srv),
                                                 pid_path, script_path))
                already_running = "RUNNING" in status
            except Exception:
                already_running = False

            if already_running:
                yield "[ft] Fine-tune already running — re-attaching to log..."
            else:
                yield f"[ft] source model : {source}"
                yield f"[ft] target project: {name}"
                yield f"[ft] MAX_ITER={max_iter}  BASE_LR={base_lr}"

                # Source weights must exist on the server
                host_weights = (f"{remote}/layout-model-training/outputs/{source}/"
                                f"fast_rcnn_R_50_FPN_3x/model_final.pth")
                st = _quick(f"test -f {host_weights} && echo OK || echo MISSING")
                if st != "OK":
                    yield (f"ERROR: model weights not found on server: {host_weights} — "
                           f"run 'Train model' on '{source}' first.")
                    return
                yield f"[ft] ✓ weights found: {host_weights}"

                # Push images + COCO + config + tools (same as Train)
                yield from _push_training_data_gen(name, srv, passphrase,
                                                   skip_images=body.skip_image_upload)

                # Upload the fine-tune script (LF endings)
                c2 = ssh_ops._client(srv["host"], srv["user"], srv["key_path"], passphrase)
                sftp2 = c2.open_sftp()
                try:
                    dest = f"{remote}/layout-model-training/scripts/{name}_finetune_from_{source}.sh"
                    yield f"[ft] finetune script → {dest}"
                    with sftp2.open(dest, "wb") as fh:
                        fh.write(ft_script.replace("\r\n", "\n").encode())
                finally:
                    sftp2.close()
                    c2.close()

                # ── Launch detached (survives browser disconnects) ─────────────
                yield from _gpu_busy_warning(_quick, "ft")
                yield "[ft] Launching detached fine-tune (safe to close browser)..."
                launch_result = _quick(_job_launch_cmd(_train_container(srv),
                                                       script_path, log_path, pid_path,
                                                       self_stop=not body.keep_container))
                yield f"[ft] {launch_result.strip()}"
                yield from _job_instant_death_check(_quick, launch_result,
                                                    _train_container(srv), log_path, "ft")

            # ── Stream log until the process exits ─────────────────────────────
            yield "[ft] Streaming log — closing browser won't stop the fine-tune..."
            tail_cmd = _job_tail_cmd(_train_container(srv), log_path, pid_path,
                                     "[Fine-tune complete — process has exited]")
            yield from ssh_ops.stream_command(
                srv["host"], srv["user"], srv["key_path"], tail_cmd, passphrase)

            if not body.keep_container:
                still = _quick(_job_still_running_cmd(_train_container(srv), pid_path))
                if "RUNNING" in still:
                    yield "[ft] Stream ended but fine-tune still running — container left up."
                else:
                    yield f"[ft] Finished — stopping container '{_train_container(srv)}'..."
                    _quick(f"docker stop {_train_container(srv)}")
                    yield f"[ft] Container '{_train_container(srv)}' stopped."
                    yield (f"[ft] Done. '{name}' now has its own fine-tuned model — "
                           f"'Run inference' on this project will use it.")

        return await _sse_stream(full_gen())

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/project/{name}/pull-predictions")
def api_pull_predictions(name: str, body: TrainRequest = TrainRequest()):
    """Pull predicted JSONs from server into local predictions/ folder (never touches annotations/)."""
    from app import ssh_ops
    try:
        srv       = _server_cfg(name)
        pdir      = project_dir(name)
        remote    = srv["remote_path"].rstrip("/")
        pred_dir  = pdir / "predictions"
        pred_dir.mkdir(parents=True, exist_ok=True)
        result = ssh_ops.pull_folder(
            srv["host"], srv["user"], srv["key_path"],
            f"{remote}/{name}/predictions",
            pred_dir,
            body.passphrase or srv.get("passphrase"),
        )
        return {**result, "local_path": str(pred_dir)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/project/{name}/apply-predictions")
def api_apply_predictions(name: str):
    """Copy predicted shapes into annotation JSONs that have no shapes yet."""
    import json as _json
    pdir     = project_dir(name)
    pred_dir = pdir / "predictions"
    ann_dir  = pdir / "annotations"
    if not pred_dir.exists():
        raise HTTPException(status_code=400, detail="No predictions pulled yet")
    applied = skipped = 0
    for pred_file in pred_dir.glob("*.json"):
        ann_file = ann_dir / pred_file.name
        if not ann_file.exists():
            continue
        ann = _json.loads(ann_file.read_text(encoding="utf-8"))
        if ann.get("shapes"):  # already has hand annotations — skip
            skipped += 1
            continue
        pred = _json.loads(pred_file.read_text(encoding="utf-8"))
        ann["shapes"] = pred.get("shapes", [])
        _write_json(ann_file, ann)
        applied += 1
    return {"applied": applied, "skipped_had_annotations": skipped}


# ---------------------------------------------------------------------------
# Routes — perspective correction
# ---------------------------------------------------------------------------

class PerspectiveRequest(BaseModel):
    folder: str
    stem:   str
    points: list = []  # [[x,y]×4] for manual; empty [] triggers auto-detection
    save:   bool = False


def _auto_detect_page_quad(img_np):
    """
    Automatically find the four corners of a document page.
    Returns a list of four [x, y] points (any order) or None.

    Strategy 1 — Canny edges → largest quad-shaped contour.
      Works well when the page has a visible border against the scanner bed.
    Strategy 2 — HoughLinesP intersections.
      Works when the page fills the frame but has prominent table/text lines.
    """
    import cv2
    import numpy as np

    H, W = img_np.shape[:2]
    gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)

    # ── Strategy 1: largest quad contour ──────────────────────────────────────
    for lo, hi in [(30, 100), (50, 150), (10, 50)]:
        edges = cv2.Canny(blur, lo, hi)
        edges = cv2.dilate(edges, np.ones((3, 3), np.uint8))
        cnts, _ = cv2.findContours(edges, cv2.RETR_LIST,
                                   cv2.CHAIN_APPROX_SIMPLE)
        if not cnts:
            continue
        for c in sorted(cnts, key=cv2.contourArea, reverse=True)[:10]:
            if cv2.contourArea(c) < 0.15 * H * W:
                break  # sorted, so no point continuing
            peri = cv2.arcLength(c, True)
            for eps in [0.01, 0.02, 0.03, 0.05]:
                approx = cv2.approxPolyDP(c, eps * peri, True)
                if len(approx) == 4:
                    return [p[0].tolist() for p in approx]

    # ── Strategy 2: Hough line intersections ──────────────────────────────────
    edges = cv2.Canny(blur, 50, 150)
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180,
                            threshold=max(50, min(W, H) // 8),
                            minLineLength=min(W, H) // 5,
                            maxLineGap=40)
    if lines is None:
        return None

    segs = lines.reshape(-1, 4).tolist()
    h_segs, v_segs = [], []
    for x1, y1, x2, y2 in segs:
        ang = np.degrees(np.arctan2(abs(y2 - y1), abs(x2 - x1) + 1e-9))
        (h_segs if ang < 20 else v_segs if ang > 70 else []).append(
            (x1, y1, x2, y2))

    if len(h_segs) < 2 or len(v_segs) < 2:
        return None

    h_segs.sort(key=lambda s: (s[1] + s[3]) / 2)
    v_segs.sort(key=lambda s: (s[0] + s[2]) / 2)

    def _avg(segs):
        return (float(np.mean([s[0] for s in segs])),
                float(np.mean([s[1] for s in segs])),
                float(np.mean([s[2] for s in segs])),
                float(np.mean([s[3] for s in segs])))

    def _isect(a, b):
        x1, y1, x2, y2 = a
        x3, y3, x4, y4 = b
        d = (x1 - x2) * (y3 - y4) - (y1 - y2) * (x3 - x4)
        if abs(d) < 1e-8:
            return None
        t = ((x1 - x3) * (y3 - y4) - (y1 - y3) * (x3 - x4)) / d
        return [x1 + t * (x2 - x1), y1 + t * (y2 - y1)]

    n_h = max(1, len(h_segs) // 5)
    n_v = max(1, len(v_segs) // 5)
    top_l  = _avg(h_segs[:n_h])
    bot_l  = _avg(h_segs[-n_h:])
    left_l = _avg(v_segs[:n_v])
    rgt_l  = _avg(v_segs[-n_v:])

    corners = [
        _isect(top_l, left_l), _isect(top_l, rgt_l),
        _isect(bot_l, rgt_l),  _isect(bot_l, left_l),
    ]
    if any(c is None for c in corners):
        return None
    # Sanity: all corners within ±30 % of the image bounds
    m = 0.3
    if not all(-W * m <= c[0] <= W * (1 + m) and
               -H * m <= c[1] <= H * (1 + m) for c in corners):
        return None

    return corners   # [TL, TR, BR, BL]


@app.post("/api/page/perspective")
def api_perspective(body: PerspectiveRequest):
    """
    Perspective (trapezoid→rectangle) correction.

    points=[]   → auto-detect page corners, return preview + detected_points.
    points=[×4] → use the supplied corners (manual or re-submit of detected).
    save=False  → base64 JPEG preview only.
    save=True   → overwrite image, clear shapes, update JSON dimensions.
    """
    import base64, io, math
    import cv2
    import numpy as np
    from PIL import Image

    d        = _resolve_folder(body.folder)
    img_path = _find_image(d, body.stem)
    jf       = d / f"{body.stem}.json"

    if img_path is None:
        raise HTTPException(status_code=404, detail="Image not found")
    if not jf.exists():
        raise HTTPException(status_code=404, detail="JSON not found")
    if body.points and len(body.points) != 4:
        raise HTTPException(status_code=400, detail="Supply 0 or exactly 4 points")

    try:
        img = Image.open(str(img_path)).convert("RGB")
    except Exception:
        import traceback
        raise HTTPException(status_code=500, detail=traceback.format_exc())

    img_np = np.array(img)
    H_img, W_img = img_np.shape[:2]

    # ── Determine the four source corners ─────────────────────────────────────
    if not body.points:
        raw = _auto_detect_page_quad(img_np)
        if raw is None:
            raise HTTPException(
                status_code=422,
                detail="Could not automatically detect the page boundary. "
                       "Try again or mark the four corners manually.")
        detected_points = [[float(v) for v in p] for p in raw]
    else:
        detected_points = [[float(v) for v in p] for p in body.points]

    # Sort into TL, TR, BR, BL regardless of supplied order
    pts = [tuple(p) for p in detected_points]
    by_y   = sorted(pts, key=lambda p: p[1])
    top    = sorted(by_y[:2], key=lambda p: p[0])
    bottom = sorted(by_y[2:], key=lambda p: p[0])
    tl, tr = top[0],    top[1]
    bl, br = bottom[0], bottom[1]

    w_top  = math.dist(tl, tr);  w_bot  = math.dist(bl, br)
    h_left = math.dist(tl, bl);  h_right = math.dist(tr, br)
    dst_w  = int(max(w_top, w_bot))
    dst_h  = int(max(h_left, h_right))

    if dst_w < 2 or dst_h < 2:
        raise HTTPException(status_code=400, detail="Degenerate quadrilateral")

    # ── Build homography and warp the full image ───────────────────────────────
    src_arr = np.float32([tl, tr, br, bl])
    dst_arr = np.float32([(0, 0), (dst_w, 0), (dst_w, dst_h), (0, dst_h)])
    H_mat   = cv2.getPerspectiveTransform(src_arr, dst_arr)

    # Project all four image corners through H to find the full output canvas.
    img_corners = np.float32(
        [[0, 0], [W_img, 0], [W_img, H_img], [0, H_img]]
    ).reshape(-1, 1, 2)
    wc = cv2.perspectiveTransform(img_corners, H_mat).reshape(-1, 2)

    min_x, min_y = float(wc[:, 0].min()), float(wc[:, 1].min())
    max_x, max_y = float(wc[:, 0].max()), float(wc[:, 1].max())

    # Translate so all pixels land in non-negative coordinates.
    T = np.array([[1, 0, -min_x], [0, 1, -min_y], [0, 0, 1]],
                 dtype=np.float64)
    H_eff = (T @ H_mat.astype(np.float64)).astype(np.float32)
    out_w  = int(round(max_x - min_x))
    out_h  = int(round(max_y - min_y))

    try:
        out_np = cv2.warpPerspective(img_np, H_eff, (out_w, out_h),
                                     flags=cv2.INTER_CUBIC)
        out = Image.fromarray(out_np)
    except Exception:
        import traceback
        raise HTTPException(status_code=500, detail=traceback.format_exc())

    # ── Save or preview ────────────────────────────────────────────────────────
    if body.save:
        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except Exception as exc:
            raise HTTPException(status_code=500,
                                detail=f"Annotation JSON corrupt: {exc}")
        fmt = {"jpg": "JPEG", "jpeg": "JPEG", "png": "PNG",
               "tif": "TIFF", "tiff": "TIFF"}.get(
            img_path.suffix.lstrip(".").lower(), "JPEG")
        save_kw = {"quality": 92} if fmt == "JPEG" else {}
        out.save(str(img_path), format=fmt, **save_kw)
        data["shapes"]      = []
        data["imageWidth"]  = out_w
        data["imageHeight"] = out_h
        _write_json(jf, data)
        return {"ok": True, "width": out_w, "height": out_h}
    else:
        buf = io.BytesIO()
        out.save(buf, format="JPEG", quality=88)
        buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode()
        return {"ok": True, "preview": b64,
                "width": out_w, "height": out_h,
                "detected_points": detected_points}


# ---------------------------------------------------------------------------
# Routes — Quality Audit
# ---------------------------------------------------------------------------

@app.get("/api/audit/random")
def api_audit_random(folder: str = Query(...), mode: str = Query("both")):
    """Return a randomly chosen shape that has OCR/LLM output but no human correction."""
    import random, base64, io
    from PIL import Image

    d = _resolve_folder(folder)
    json_files = list(d.glob("*.json"))
    if not json_files:
        raise HTTPException(status_code=404, detail="No pages found")
    random.shuffle(json_files)

    def _scan_page(jf):
        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except Exception:
            return []
        hits = []
        for idx, shape in enumerate(data.get("shapes", [])):
            has_ocr   = bool(
                shape.get("easyocr_output", {}).get("ocr_text") or
                shape.get("tesseract_output", {}).get("ocr_text")
            )
            has_llm   = bool(shape.get("openai_output", {}).get("response"))
            has_human = bool(shape.get("human_output", {}).get("human_corrected_text"))
            if has_human:
                continue
            if mode == "ocr" and not has_ocr:
                continue
            if mode == "llm" and not has_llm:
                continue
            if mode == "both" and not (has_ocr and has_llm):
                continue
            hits.append((jf.stem, idx))
        return hits

    # Fast path: try random pages one by one; stop as soon as we have candidates
    candidates = []
    for jf in json_files:
        candidates = _scan_page(jf)
        if candidates:
            break

    if not candidates:
        raise HTTPException(status_code=404, detail="No qualifying shapes found")

    stem, idx = random.choice(candidates)
    jf   = d / f"{stem}.json"
    data = json.loads(jf.read_text(encoding="utf-8"))
    shape = data["shapes"][idx]

    ocr_text = (
        shape.get("easyocr_output", {}).get("ocr_text") or
        shape.get("tesseract_output", {}).get("ocr_text") or ""
    )
    llm_text = shape.get("openai_output", {}).get("response") or ""
    label    = shape.get("label", "")

    # Crop cell image and detect row boundaries
    image_b64 = None
    line_rows  = []          # [[top, bottom], ...] in crop-pixel space
    img_path  = _find_image(d, stem)
    if img_path is not None:
        try:
            pts = shape.get("points", [])
            if pts:
                xs = [p[0] for p in pts]
                ys = [p[1] for p in pts]
                pad  = 4
                x1   = max(0, int(min(xs)) - pad)
                y1   = max(0, int(min(ys)) - pad)
                x2   = int(max(xs)) + pad
                y2   = int(max(ys)) + pad
                img  = Image.open(str(img_path)).convert("RGB")
                crop = img.crop((x1, y1, x2, y2))
                buf  = io.BytesIO()
                crop.save(buf, format="JPEG", quality=90)
                buf.seek(0)
                image_b64 = base64.b64encode(buf.read()).decode()
                # Detect row boundaries using the shadow page (same as OCR line-by-line)
                try:
                    shadow = _get_shadow_page(folder, stem, img_path)
                    shadow_crop = shadow.crop((x1, y1, x2, y2))
                    rows = _detect_text_rows(shadow_crop, cell_height=26)
                    line_rows = [list(r) for r in rows]
                except Exception:
                    line_rows = []
        except Exception:
            image_b64 = None

    return {
        "stem":       stem,
        "idx":        idx,
        "label":      label,
        "ocr_text":   ocr_text,
        "llm_text":   llm_text,
        "image_b64":  image_b64,
        "line_rows":  line_rows,
    }


_AUDIT_STATS_DEFAULT = {
    "ocr": {"msd_sum": 0, "msd_count": 0, "cer_sum": 0, "cer_count": 0},
    "llm": {"msd_sum": 0, "msd_count": 0, "cer_sum": 0, "cer_count": 0},
}


@app.get("/api/audit/stats")
def api_audit_get_stats(folder: str = Query(...)):
    """Return accumulated audit stats for the project."""
    stats_file = _resolve_folder(folder).parent / "audit_stats.json"
    if not stats_file.exists():
        return _AUDIT_STATS_DEFAULT.copy()
    try:
        return json.loads(stats_file.read_text(encoding="utf-8"))
    except Exception:
        return _AUDIT_STATS_DEFAULT.copy()


@app.post("/api/audit/stats")
async def api_audit_update_stats(folder: str = Query(...), request: Request = None):
    """Save audit stats for the project."""
    body = await request.json()
    stats_file = _resolve_folder(folder).parent / "audit_stats.json"
    _write_json(stats_file, body)
    return {"ok": True}


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@app.get("/api/export/excel")
def api_export_excel(
    folder:      str  = Query(...),
    scope:       str  = Query("page"),      # "page" | "document"
    stem:        str  = Query(None),        # required when scope="page"
    dual:        bool = Query(False),       # legacy: equivalent to pattern="1,1"
    pattern:     str  = Query(""),          # 1/0 page pattern, e.g. "1,1,0,0"; empty = all pages vertically
    layer:       str  = Query("best_llm"), # "ocr"|"llm"|"human"|"best_ocr"|"best_llm"
    types:       str  = Query(""),         # comma-sep label types; empty = all
    col_headers: str  = Query(""),         # comma-sep column header labels; empty = no header row
    stems:       str  = Query(""),         # comma-sep stems to include; empty = all (document scope)
    col_filter:  str  = Query(""),         # comma-sep 1-indexed column numbers; empty = all columns
    page_from:   int  = Query(None),       # 1-indexed first page of range (inclusive); alternative to stems
    page_to:     int  = Query(None),       # 1-indexed last  page of range (inclusive)
    rows_only:   bool = Query(False),      # export only cells that have an internal row structure
    clip_col:    bool = Query(False),      # add a "Clip" meta column with each row's clip id(s)
    clips_only:  bool = Query(False),      # export only rows that carry a clip
    auth_cols:   bool = Query(True),       # insert resolved-authority name+ID columns next to each resolved column
):
    """
    Generate an .xlsx preserving spatial layout.
    Each text line inside a lattice cell becomes its own Excel row.
    When cells in the same lattice row have different line counts, the
    shorter ones are padded with blank cells coloured light-red.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500,
            detail="openpyxl not installed — run: pip install openpyxl")

    import io as _io, re
    from collections import defaultdict

    d = _resolve_folder(folder)
    selected_types = {t.strip() for t in types.split(",") if t.strip()} if types else set()

    # Number of meta columns before the data: Row# always, + Clip when asked.
    # Data columns live at col_idx + 1 + META_COLS (+ group col_offset).
    META_COLS = 2 if clip_col else 1

    # ── Styles ───────────────────────────────────────────────────────────────
    from openpyxl.styles import Font
    _align      = Alignment(wrap_text=True, vertical="top", horizontal="left")
    _red_fill   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    _blue_fill  = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    _src_fill   = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
    _src_font   = Font(italic=True, bold=True, color="444444")

    def natural_key(p):
        return [int(c) if c.isdigit() else c.lower()
                for c in re.split(r'(\d+)', str(p))]

    def get_text(shape):
        human = (shape.get("human_output") or {}).get("human_corrected_text") or ""
        # structural blank → missing (empty), unless a human explicitly overrode it
        if shape.get("blank"):
            return human.strip()
        ocr   = ((shape.get("tesseract_output") or {}).get("ocr_text") or
                 (shape.get("easyocr_output")   or {}).get("ocr_text") or "")
        llm   = (shape.get("openai_output") or {}).get("response") or ""
        pdf   = shape.get("pdf_text") or ""
        if layer == "human":    return human.strip()
        if layer == "ocr":      return ocr.strip()
        if layer == "llm":      return llm.strip()
        if layer == "pdf":      return pdf.strip()
        if layer == "best_ocr": return (human or ocr or llm).strip()
        if layer == "best_pdf": return (human or llm or ocr or pdf).strip()
        return (human or llm or ocr).strip()   # best_llm (default)

    def text_to_lines(text):
        """Split text into lines, stripping trailing blank lines."""
        lines = text.split("\n")
        while lines and not lines[-1].strip():
            lines.pop()
        return lines or [""]

    def get_row_lines(shape):
        """Per-internal-row layer pick from row_struct, mirroring get_text's
        layer logic at row level.  None when the shape has no usable rows —
        the caller then falls back to the flat-text logic."""
        rows = (shape.get("row_struct") or {}).get("rows") or []
        if not rows:
            return None

        def pick(r):
            h = (r.get("human") or "").strip()
            if r.get("blank"):      # structural blank → missing (human wins)
                return h
            o = (r.get("ocr")   or "").strip()
            l = (r.get("llm")   or "").strip()
            p = (r.get("pdf")   or "").strip()
            if layer == "human":    return h
            if layer == "ocr":      return o
            if layer == "llm":      return l
            if layer == "pdf":      return p
            if layer == "best_ocr": return h or o or l
            if layer == "best_pdf": return h or l or o or p
            return h or l or o      # best_llm (default)

        lines = [pick(r) for r in rows]
        return lines if any(lines) else None

    def _spatial_cluster_rows(items, tol):
        """Group items by Y proximity. Returns row_idx on each item."""
        items.sort(key=lambda c: c["top_y"])
        row_groups: list = []
        for c in items:
            placed = False
            for grp in row_groups:
                avg_top = sum(r["top_y"] for r in grp) / len(grp)
                avg_bot = sum(r["bot_y"] for r in grp) / len(grp)
                if avg_top - tol <= c["cy"] <= avg_bot + tol:
                    grp.append(c); placed = True; break
            if not placed:
                row_groups.append([c])
        for ri, grp in enumerate(row_groups):
            for c in grp:
                c["row_idx"] = ri

    def _spatial_cluster_cols(items):
        """Assign col_idx by clustering cx values."""
        if not items:
            return
        med_w    = sorted(c["w"] for c in items)[len(items) // 2]
        thresh_x = max(3, med_w * 0.5)
        all_cx   = sorted({c["cx"] for c in items})
        col_centers: list = []
        grp = [all_cx[0]]
        for cx in all_cx[1:]:
            if cx - grp[-1] <= thresh_x:
                grp.append(cx)
            else:
                col_centers.append(sum(grp) / len(grp)); grp = [cx]
        col_centers.append(sum(grp) / len(grp))
        for c in items:
            c["col_idx"] = min(range(len(col_centers)),
                               key=lambda i: abs(col_centers[i] - c["cx"]))

    def _auth_pair(c, lines):
        """Per-line resolved (names, ids) for one cell, padded/truncated to
        len(lines); (None, None) when nothing is resolved. Internal-row
        authority wins (index-aligned with the row lines); a whole-cell
        authority goes on the first line."""
        ra = c.get("rows_auth") or []
        if c.get("row_lines") and any(ra):
            names = [((a or {}).get("name") or "") for a in ra]
            ids   = [((a or {}).get("id")   or "") for a in ra]
        elif c.get("auth"):
            names = [(c["auth"].get("name") or "")]
            ids   = [(c["auth"].get("id")   or "")]
        else:
            return None, None
        n = len(lines)
        return (names + [""] * n)[:n], (ids + [""] * n)[:n]

    def shapes_to_cells(shapes):
        """
        Return list of dicts: {row_idx, col_idx, lines[], w_px}.

        Lattice path (preferred): shapes with super_row/super_column use those
        directly (1-indexed → 0-indexed).  Text is expanded line-by-line so the
        lattice row structure is preserved in Excel.

        Spatial path (no lattice): each annotation = ONE Excel cell regardless
        of how many text lines it contains.  Row/col are assigned by spatial
        clustering of bounding-box positions.  This prevents multi-line cells
        from bleeding over adjacent annotations.
        """
        raw = []
        for sh in shapes:
            if selected_types and sh.get("label", "") not in selected_types:
                continue
            if rows_only and not ((sh.get("row_struct") or {}).get("rows")):
                continue
            row_lines = get_row_lines(sh)   # internal row structure wins
            text      = get_text(sh)
            if not text and not row_lines:
                continue
            pts = sh.get("points", [])
            if not pts:
                continue
            xs = [p[0] for p in pts]; ys = [p[1] for p in pts]
            x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
            raw.append(dict(text=text, row_lines=row_lines,
                            cx=(x1+x2)/2, cy=(y1+y2)/2,
                            top_y=y1, bot_y=y2,
                            h=max(1, y2-y1), w=max(1, x2-x1),
                            label=sh.get("label", "?"),
                            clip=sh.get("clip"),
                            table=sh.get("table") or 0,
                            super_row=sh.get("super_row"),
                            super_col=sh.get("super_column"),
                            auth=sh.get("authority"),
                            rows_auth=[r.get("authority") for r in
                                       (sh.get("row_struct") or {}).get("rows") or []]))
        print(f"[EXCEL] shapes_to_cells: {len(shapes)} shapes in, {len(raw)} with text", flush=True)
        if not raw:
            return []

        have_coords = [c for c in raw if c["super_row"] is not None and c["super_col"] is not None]

        # ── Spatial path: no lattice on this page ────────────────────────────
        if not have_coords:
            med_h   = sorted(c["h"] for c in raw)[len(raw) // 2]
            row_tol = max(5, med_h * 0.4)
            _spatial_cluster_rows(raw, row_tol)
            _spatial_cluster_cols(raw)
            # Each shape → one cell; text kept as a single value (no line expansion)
            winner: dict = {}
            for c in raw:
                k = (c["row_idx"], c["col_idx"])
                if k not in winner or c["h"] > winner[k]["h"]:
                    winner[k] = c
            if winner:
                min_row = min(c["row_idx"] for c in winner.values())
                if min_row != 0:
                    for c in winner.values(): c["row_idx"] -= min_row
            out = []
            for c in winner.values():
                # one cell, no line expansion
                lines = [c["text"] or "\n".join(c["row_lines"] or [])]
                names, ids = _auth_pair(dict(c, row_lines=None), lines)
                out.append(dict(row_idx=c["row_idx"], col_idx=c["col_idx"],
                                lines=lines, w_px=c["w"], clip=c.get("clip"),
                                auth_names=names, auth_ids=ids))
            return out

        # ── Lattice path: use stored super_row / super_column ────────────────
        no_coords = [c for c in raw if c["super_row"] is None or c["super_col"] is None]

        for c in have_coords:
            c["row_idx"] = int(c["super_row"]) - 1
            c["col_idx"] = int(c["super_col"]) - 1

        tables = sorted({c["table"] for c in have_coords})

        if len(tables) <= 1:
            # ── single lattice: keep the grid from the lattice cells only ──
            winner = {}
            for c in have_coords:
                k = (c["row_idx"], c["col_idx"])
                if k not in winner or c["h"] > winner[k]["h"]:
                    winner[k] = c
            if winner:
                min_row = min(c["row_idx"] for c in winner.values())
                if min_row != 0:
                    for c in winner.values(): c["row_idx"] -= min_row
            cells = list(winner.values())
        else:
            # ── multiple lattices on the page: stack tables top-to-bottom ──
            def _table_top(t): return min(c["top_y"] for c in have_coords if c["table"] == t)
            cells, row_offset = [], 0
            for t in sorted(tables, key=_table_top):
                winner = {}
                for c in (x for x in have_coords if x["table"] == t):
                    k = (c["row_idx"], c["col_idx"])
                    if k not in winner or c["h"] > winner[k]["h"]:
                        winner[k] = c
                ws = list(winner.values())
                if not ws:
                    continue
                min_row = min(c["row_idx"] for c in ws)
                top = 0
                for c in ws:
                    c["row_idx"] = c["row_idx"] - min_row + row_offset
                    top = max(top, c["row_idx"])
                cells.extend(ws)
                row_offset = top + 2          # one blank row between stacked tables

        # ── Interleave non-lattice annotations (titles / headers / free text) by
        # vertical position so they aren't dropped on lattice pages. Each becomes
        # its own row at column 0, inserted between the lattice rows it sits among.
        # (Skipped when there are none → lattice-only pages are unchanged.)
        if no_coords:
            row_y = {}
            for c in cells:
                row_y.setdefault(c["row_idx"], []).append(c["top_y"])
            row_y = {r: min(v) for r, v in row_y.items()}
            for c in cells:
                c["_pos"] = float(c["row_idx"])
            for j, c in enumerate(sorted(no_coords, key=lambda x: x["top_y"])):
                c["col_idx"] = 0
                rank = sum(1 for y in row_y.values() if y <= c["top_y"])
                c["_pos"] = (rank - 0.5) + j * 1e-6      # between lattice rows, unique & y-ordered
            all_cells = cells + no_coords
            order = sorted({c["_pos"] for c in all_cells})
            remap = {p: i for i, p in enumerate(order)}
            for c in all_cells:
                c["row_idx"] = remap[c["_pos"]]
            cells = all_cells

        _n_free = sum(1 for c in cells if c["super_row"] is None or c["super_col"] is None)
        print(f"[EXCEL] shapes_to_cells returning {len(cells)} cells "
              f"({len(cells) - _n_free} lattice + {_n_free} non-lattice interleaved)", flush=True)
        out = []
        for c in cells:
            # Internal row structure is authoritative when present
            lines = c["row_lines"] if c["row_lines"] else text_to_lines(c["text"])
            names, ids = _auth_pair(c, lines)
            out.append(dict(row_idx=c["row_idx"], col_idx=c["col_idx"],
                            lines=lines, w_px=c["w"], clip=c.get("clip"),
                            # free = interleaved non-lattice annotation (title/footer/…);
                            # exempt from the column filter — it has no real column.
                            free=(c["super_row"] is None or c["super_col"] is None),
                            auth_names=names, auth_ids=ids))
        return out

    # ── Helpers for stacking / horizontal page groups ────────────────────────

    def row_rank_list(cells):
        """Ordered list of (row_idx, height) for the lattice rows that are
        actually printed (rows with no surviving cells simply don't appear)."""
        rmap: dict = defaultdict(list)
        for c in cells:
            rmap[c["row_idx"]].append(c)
        return [(ridx, max(len(c["lines"]) for c in rmap[ridx]))
                for ridx in sorted(rmap)]

    def write_cells(ws, cells, col_offset=0, base_row=1, align_pad=0, row_heights=None):
        """
        Write cells to ws starting at base_row.
        align_pad blank rows (light-blue) are prepended before content — used to
        align the lattice with a paired page in dual-page mode.
        row_heights forces minimum heights per lattice row so paired pages stay
        line-for-line aligned in dual mode.
        Within each lattice row, cells shorter than max_lines are padded with
        blank light-red cells to flag line-count mismatches.
        """
        if not cells:
            return

        # ── Alignment padding (blue) ─────────────────────────────────────
        # Data columns are now shifted +1 to make room for the meta (row#) column.
        if align_pad > 0:
            min_ci = min(c["col_idx"] for c in cells)
            max_ci = max(c["col_idx"] for c in cells)
            for pad_r in range(base_row, base_row + align_pad):
                for ci in range(min_ci, max_ci + 1):
                    ws.cell(row=pad_r,
                            column=ci + 1 + META_COLS + col_offset).fill = _blue_fill

        # ── Cell content ─────────────────────────────────────────────────
        rmap: dict = defaultdict(list)
        for c in cells:
            rmap[c["row_idx"]].append(c)

        excel_row = base_row + align_pad
        for row_idx in sorted(rmap):
            row_cells = rmap[row_idx]
            max_lines = max(len(c["lines"]) for c in row_cells)
            if row_heights:
                max_lines = max(max_lines, row_heights.get(row_idx, 0))
            # This lattice row's clip id(s): distinct clips among its cells
            row_clip = ";".join(sorted({str(c["clip"]) for c in row_cells
                                        if c.get("clip") is not None}))
            for line_i in range(max_lines):
                # Write lattice row number in meta column on every sub-row
                meta_c = ws.cell(row=excel_row + line_i, column=col_offset + 1)
                meta_c.value     = row_idx + 1   # 1-based
                meta_c.alignment = _align
                if clip_col:
                    clip_c = ws.cell(row=excel_row + line_i, column=col_offset + 2)
                    clip_c.value     = row_clip
                    clip_c.alignment = _align
                for c in row_cells:
                    col  = c["col_idx"] + 1 + META_COLS + col_offset   # after meta col(s)
                    xcel = ws.cell(row=excel_row + line_i, column=col)
                    xcel.alignment = _align
                    if line_i < len(c["lines"]):
                        xcel.value = c["lines"][line_i]
                    else:
                        xcel.value = ""
                        xcel.fill  = _red_fill
            excel_row += max_lines

        # ── Column widths based on maximum content length ────────────────
        # Meta column: fit the largest row number
        meta_letter = get_column_letter(col_offset + 1)
        max_row_num = (max(rmap.keys()) + 1) if rmap else 1
        meta_w = max(4.0, float(len(str(max_row_num)) + 1))
        if ws.column_dimensions[meta_letter].width < meta_w:
            ws.column_dimensions[meta_letter].width = meta_w
        # Data columns: max character length of any line in that column
        col_max_len: dict = defaultdict(int)
        for c in cells:
            for line in c["lines"]:
                col_max_len[c["col_idx"]] = max(col_max_len[c["col_idx"]], len(str(line)))
        for col_idx, max_len in col_max_len.items():
            width  = max(4.0, min(60.0, float(max_len + 2)))
            letter = get_column_letter(col_idx + 1 + META_COLS + col_offset)
            if ws.column_dimensions[letter].width < width:
                ws.column_dimensions[letter].width = width

    def write_source_row(ws, row, entries):
        """Write source file names in the meta columns before data rows.
        entries: list of (col_offset, name) — one per page in the group."""
        for col_offset, name in entries:
            c = ws.cell(row=row, column=col_offset + 1)
            c.value     = name
            c.fill      = _src_fill
            c.font      = _src_font
            c.alignment = _align

    def max_col_of(cells):
        return max((c["col_idx"] for c in cells), default=-1)

    # ── Column-header support ────────────────────────────────────────────────
    hdr_list = [h.strip() for h in col_headers.split(",") if h.strip()] \
               if col_headers.strip() else []
    _hdr_font = Font(bold=True)
    _hdr_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    _hdr_font_white = Font(bold=True, color="FFFFFF")

    def write_header_row(ws, col_offset=0):
        """Write 'Row' (+ 'Clip') meta header + hdr_list data headers at row 1."""
        def _hdr(col, text):
            cell = ws.cell(row=1, column=col)
            cell.value     = text
            cell.font      = _hdr_font_white
            cell.fill      = _hdr_fill
            cell.alignment = _align
        _hdr(col_offset + 1, "Row")
        if clip_col:
            _hdr(col_offset + 2, "Clip")
        # User-specified data column headers (after the meta column(s)),
        # shifted around any inserted authority columns
        for i, label in enumerate(hdr_list):
            _hdr(_newcol(i) + 1 + META_COLS + col_offset, label)
        # Authority column headers (always labeled — they're synthetic)
        for a in _acols:
            src = hdr_list[a] if a < len(hdr_list) else f"col {a + 1}"
            _hdr(_newcol(a) + 2 + META_COLS + col_offset, f"{src} → name")
            _hdr(_newcol(a) + 3 + META_COLS + col_offset, f"{src} → id")

    data_start = 2   # row 1 is always the header row (meta col "Row" + optional data headers)

    # ── Column filter ────────────────────────────────────────────────────────
    # col_filter is a comma-sep list of 1-indexed column numbers from the client.
    # We parse them to a 0-indexed set, then strip and re-index after shapes_to_cells.
    # col_filter: comma-sep 1-indexed column numbers/ranges; empty = all columns.
    # Supports open-ended ranges: "4-" means col 4 to infinity; "-3" means 1-3.
    _col_filter_ranges: list | None = None   # list of (lo, hi) 0-indexed, hi=None=∞
    if col_filter.strip():
        _col_filter_ranges = []
        for part in col_filter.split(","):
            part = part.strip()
            if not part:
                continue
            try:
                if "-" in part:
                    lo_s, _, hi_s = part.partition("-")
                    lo = (int(lo_s) - 1) if lo_s.strip() else 0
                    hi = (int(hi_s) - 1) if hi_s.strip() else None   # None = ∞
                    _col_filter_ranges.append((lo, hi))
                else:
                    v = int(part) - 1
                    _col_filter_ranges.append((v, v))
            except ValueError:
                pass

    def _col_matches(col_idx):
        return any(lo <= col_idx <= (hi if hi is not None else col_idx)
                   for lo, hi in _col_filter_ranges)

    def apply_col_filter(cells):
        """Keep only columns matching _col_filter_ranges and re-index col_idx.

        Non-lattice annotations interleaved on lattice pages (free=True) are
        kept no matter what — they sit at column 0 only because they have no
        real column, so the filter must not drop them."""
        if _col_filter_ranges is None:
            return cells
        kept = [c for c in cells if not c.get("free") and _col_matches(c["col_idx"])]
        # Re-index: sort the kept col indices, map old → new
        old_cols = sorted({c["col_idx"] for c in kept})
        remap = {old: new for new, old in enumerate(old_cols)}
        out = [{**c, "col_idx": remap[c["col_idx"]]} for c in kept]
        out.extend({**c, "col_idx": 0} for c in cells if c.get("free"))
        return out

    def apply_clips_only(cells):
        """Keep only rows that carry a clip on at least one of their cells."""
        if not clips_only:
            return cells
        rows_with_clip = {c["row_idx"] for c in cells if c.get("clip") is not None}
        return [c for c in cells if c["row_idx"] in rows_with_clip]

    # ── Authority columns: resolved name + ID inserted next to each source
    # column that carries any resolution (per internal row when present,
    # whole-cell otherwise). The set of authority-bearing columns is computed
    # GLOBALLY over all exported pages so pattern groups stay column-aligned
    # even when early pages are still unresolved. Opt out with auth_cols=false.
    _acols: list = []          # post-filter data col indices carrying authority

    def _newcol(col):
        return col + 2 * sum(1 for a in _acols if a < col)

    def compute_auth_cols(cells_lists):
        if not auth_cols:
            return
        seen = set()
        for cells in cells_lists:
            for c in cells:
                if not c.get("free") and c.get("auth_names"):
                    seen.add(c["col_idx"])
        _acols.extend(sorted(seen))

    def inject_auth_cells(cells):
        """Shift data columns to make room and add the name/ID cells."""
        if not _acols:
            return cells
        out = []
        for c in cells:
            nc = dict(c)
            if not nc.get("free"):
                nc["col_idx"] = _newcol(c["col_idx"])
            out.append(nc)
            if c.get("free") or not c.get("auth_names"):
                continue
            base = _newcol(c["col_idx"])
            out.append(dict(row_idx=c["row_idx"], col_idx=base + 1,
                            lines=c["auth_names"], w_px=1, clip=None, free=False))
            out.append(dict(row_idx=c["row_idx"], col_idx=base + 2,
                            lines=c["auth_ids"], w_px=1, clip=None, free=False))
        return out

    def load_shapes(jf):
        try:
            return json.loads(jf.read_text(encoding="utf-8")).get("shapes", [])
        except Exception:
            return []

    # ── Collect page files ────────────────────────────────────────────────────
    if scope == "page":
        if not stem:
            raise HTTPException(status_code=400, detail="stem required for scope=page")
        jfiles = [d / f"{stem}.json"]
    else:
        jfiles = sorted(d.glob("*.json"), key=lambda f: natural_key(f.stem))
        if stems.strip():
            # Client sent an explicit ordered list of stems — filter and preserve that order
            stem_list  = [s.strip() for s in stems.split(",") if s.strip()]
            stem_index = {s: i for i, s in enumerate(stem_list)}
            jfiles = sorted(
                (jf for jf in jfiles if jf.stem in stem_index),
                key=lambda jf: stem_index[jf.stem],
            )
        elif page_from is not None or page_to is not None:
            # Numeric range against the full sorted JSON list (no image required)
            lo = max(0, (page_from or 1) - 1)
            hi = (page_to or len(jfiles)) - 1
            jfiles = jfiles[lo : hi + 1]

    wb  = openpyxl.Workbook()
    wb.remove(wb.active)

    if scope == "page":
        # ── Single page: one sheet (no source row needed) ─────────────────
        print(f"[EXCEL] processing page {jfiles[0].name}", flush=True)
        ws = wb.create_sheet(title=stem[:31])
        page_cells = apply_clips_only(apply_col_filter(shapes_to_cells(load_shapes(jfiles[0]))))
        compute_auth_cols([page_cells])
        write_header_row(ws)
        write_cells(ws, inject_auth_cells(page_cells), base_row=data_start)

    else:
        # ── Whole document: 1/0 page pattern over the selected sequence ───
        # Each cycle of the pattern consumes len(pattern) pages; the 1-pages
        # of a cycle are printed side by side horizontally, the 0-pages are
        # skipped, then the next cycle continues below.  Empty pattern = "1"
        # (every page, stacked vertically).  Legacy dual=true maps to "1,1".
        bits = [1 if p.strip() == "1" else 0
                for p in pattern.split(",") if p.strip() != ""] if pattern.strip() else []
        if not bits:
            bits = [1, 1] if dual else [1]

        # Pre-compute every printed page's cells once so the authority-bearing
        # column set can be established globally BEFORE any group is written
        # (pattern groups must agree on column positions across pages).
        cells_by_jf = {}
        for k, jf in enumerate(jfiles):
            if bits[k % len(bits)] == 1:
                print(f"[EXCEL] processing {jf.name}", flush=True)
                cells_by_jf[jf] = apply_clips_only(apply_col_filter(
                    shapes_to_cells(load_shapes(jf))))
        compute_auth_cols(cells_by_jf.values())

        ws      = wb.create_sheet(title="Export")
        cur_row = data_start
        headers_written = False
        i = 0
        while i < len(jfiles):
            chunk = jfiles[i : i + len(bits)]
            i += len(bits)
            printed = [jf for jf, b in zip(chunk, bits) if b == 1]
            if not printed:
                continue
            cells_list = [inject_auth_cells(cells_by_jf[jf]) for jf in printed]
            if not any(cells_list):
                continue

            # Column offset of each page in the group: meta col + data + gap
            offsets, coff = [], 0
            for cells in cells_list:
                offsets.append(coff)
                coff += max_col_of(cells) + META_COLS + 2

            if not headers_written:
                for o in offsets:
                    write_header_row(ws, col_offset=o)
                headers_written = True

            # Source banner BEFORE the group data
            write_source_row(ws, cur_row,
                             [(o, jf.stem) for o, jf in zip(offsets, printed)])

            # Rank-based row sync: the k-th PRINTED lattice row of every page
            # in the group starts on the same Excel row; the next rank starts
            # after the tallest k-th row (shorter rows get red-padded blanks)
            ranks   = [row_rank_list(cells) for cells in cells_list]
            n_ranks = max((len(r) for r in ranks), default=0)
            heights = [max((r[k][1] for r in ranks if k < len(r)), default=1)
                       for k in range(n_ranks)]
            hmaps   = [{ridx: heights[k] for k, (ridx, _h) in enumerate(r)}
                       for r in ranks]

            for cells, o, hmap in zip(cells_list, offsets, hmaps):
                write_cells(ws, cells, col_offset=o,
                            base_row=cur_row + 1, row_heights=hmap)

            cur_row += 1 + sum(heights)   # 1 source row + data rows

    # ── Companion sheet: resolved authority entities ──────────────────────────
    # One row per resolved unit (whole cell, or each internal row), keeping the
    # ORIGINAL text alongside the resolved name + ID so nothing is lost.
    def _row_layer_text(r):
        h = (r.get("human") or "").strip(); o = (r.get("ocr") or "").strip()
        l = (r.get("llm") or "").strip();   p = (r.get("pdf") or "").strip()
        if layer == "human":    return h
        if layer == "ocr":      return o
        if layer == "llm":      return l
        if layer == "pdf":      return p
        if layer == "best_ocr": return h or o or l
        if layer == "best_pdf": return h or l or o or p
        return h or l or o

    resolved_recs = []
    for jf in jfiles:
        for sh in load_shapes(jf):
            if selected_types and sh.get("label", "") not in selected_types:
                continue
            tbl = sh.get("table") or 0
            sr, sc = sh.get("super_row"), sh.get("super_column")
            rows = (sh.get("row_struct") or {}).get("rows") or []
            if rows:
                for r in rows:
                    a = r.get("authority")
                    if not a:
                        continue
                    resolved_recs.append((jf.stem, tbl, sr, sc, r.get("n"),
                                          _row_layer_text(r), a.get("name"), a.get("id"),
                                          a.get("type"), a.get("score"), a.get("source")))
            else:
                a = sh.get("authority")
                if not a:
                    continue
                resolved_recs.append((jf.stem, tbl, sr, sc, None,
                                      get_text(sh), a.get("name"), a.get("id"),
                                      a.get("type"), a.get("score"), a.get("source")))

    if resolved_recs:
        rs = wb.create_sheet(title="Resolved")
        hdr = ["page", "table", "row", "col", "internal_row", "original",
               "resolved_name", "resolved_id", "entity_type", "score", "source"]
        for ci, h in enumerate(hdr, 1):
            c = rs.cell(row=1, column=ci, value=h); c.font = _src_font; c.fill = _src_fill
        for ri, rec in enumerate(resolved_recs, 2):
            for ci, v in enumerate(rec, 1):
                rs.cell(row=ri, column=ci, value=v)

    # ── Companion sheet: structured (JSON) records, one row per annotation ────
    def _flatten(obj, prefix=""):
        """Flatten a nested dict to dotted keys; lists → JSON string."""
        flat = {}
        if isinstance(obj, dict):
            for k, v in obj.items():
                key = f"{prefix}{k}"
                if isinstance(v, dict):
                    flat.update(_flatten(v, key + "."))
                elif isinstance(v, list):
                    flat[key] = json.dumps(v, ensure_ascii=False)
                else:
                    flat[key] = v
        return flat

    struct_recs, struct_keys = [], []
    for jf in jfiles:
        for sh in load_shapes(jf):
            if selected_types and sh.get("label", "") not in selected_types:
                continue
            st = sh.get("structured") or {}
            rec = st.get("data") if st.get("data") is not None else st.get("llm")
            if not isinstance(rec, dict):
                continue
            flat = _flatten(rec)
            for k in flat:
                if k not in struct_keys:
                    struct_keys.append(k)
            struct_recs.append({"page": jf.stem, "table": sh.get("table") or 0,
                                "row": sh.get("super_row"), "col": sh.get("super_column"),
                                "schema": st.get("schema_name"), **flat})
    if struct_recs:
        ss = wb.create_sheet(title="Structured")
        hdr = ["page", "table", "row", "col", "schema"] + struct_keys
        for ci, h in enumerate(hdr, 1):
            c = ss.cell(row=1, column=ci, value=h); c.font = _src_font; c.fill = _src_fill
        for ri, rec in enumerate(struct_recs, 2):
            for ci, h in enumerate(hdr, 1):
                v = rec.get(h)
                ss.cell(row=ri, column=ci, value=v if (v is None or isinstance(v, (str, int, float, bool))) else str(v))

    if not wb.worksheets:
        wb.create_sheet("Empty")

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"{stem}.xlsx" if (scope == "page" and stem) else "export.xlsx"
    import datetime as _dt
    headers = {
        "Content-Disposition": f'attachment; filename="{fname}"',
        "X-EconAI-Version": "taller-wins-probe",
        "X-EconAI-Time": _dt.datetime.utcnow().isoformat(),
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ---------------------------------------------------------------------------
# Structured export — gather per-annotation JSON records across pages, with
# "propagate forward" for title-type labels (country headers etc.).
# ---------------------------------------------------------------------------

class JsonExportRequest(BaseModel):
    stems:       List[str] = []          # ordered pages to export from
    label_modes: dict = {}               # label -> "export" | "ignore" | "propagate"
    mode:        str = "single"          # "single" (one file) | "per_annotation" (zip)


def _best_shape_text(sh: dict) -> str:
    h = ((sh.get("human_output") or {}).get("human_corrected_text") or "").strip()
    l = ((sh.get("openai_output") or {}).get("response") or "").strip()
    o = ((sh.get("tesseract_output") or {}).get("ocr_text")
         or (sh.get("easyocr_output") or {}).get("ocr_text") or "").strip()
    p = (sh.get("pdf_text") or "").strip()
    return h or l or o or p


def _shape_topleft(sh: dict):
    pts = sh.get("points") or []
    if not pts:
        return (0.0, 0.0)
    ys = [pt[1] for pt in pts]; xs = [pt[0] for pt in pts]
    return (min(ys), min(xs))


@app.post("/api/export/json")
def api_export_json(folder: str = Query(...), body: JsonExportRequest = ...):
    """Export structured (JSON) records across the selected pages.

    Per label, `label_modes` says: export (emit the shape's structured record if
    it has one), ignore (skip), or propagate (a non-structured title annotation
    whose text is carried into every later record under a key named after the
    label, until the next annotation of that label resets it). Records are taken
    in reading order (top→bottom, then left→right) within each page, pages in the
    given order. Returns one JSON file (mode=single) or a zip of one file per
    record (mode=per_annotation)."""
    import io as _io
    d = _resolve_folder(folder)
    modes = body.label_modes or {}
    stems = [s for s in (body.stems or []) if s] or \
            [jf.stem for jf in sorted(d.glob("*.json"), key=lambda p: _page_sort_key(p.stem))]

    carry: dict = {}                 # label -> propagated value (string)
    records = []                     # list of {data, stem, idx}
    for stem in stems:
        jf = d / f"{stem}.json"
        if not jf.exists():
            continue
        try:
            shapes = json.loads(jf.read_text(encoding="utf-8")).get("shapes", [])
        except Exception:
            continue
        order = sorted(range(len(shapes)), key=lambda i: _shape_topleft(shapes[i]))
        for i in order:
            sh = shapes[i]
            mode = modes.get(sh.get("label", ""))
            if mode == "propagate":
                txt = _best_shape_text(sh)
                if txt:
                    carry[sh.get("label", "")] = txt
                else:
                    carry.pop(sh.get("label", ""), None)
            elif mode == "export":
                st = sh.get("structured") or {}
                rec = st.get("data") if st.get("data") is not None else st.get("llm")
                if rec is None:
                    continue                       # "export if exists" — none here
                merged = {**carry, **rec} if isinstance(rec, dict) else {**carry, "value": rec}
                records.append({"data": merged, "stem": stem, "idx": i})
            # else "ignore" / unset → skip

    import datetime as _dt
    if body.mode == "per_annotation":
        import zipfile
        buf = _io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for r in records:
                zf.writestr(f"{r['stem']}__{r['idx']}.json",
                            json.dumps(r["data"], ensure_ascii=False, indent=2))
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/zip", headers={
            "Content-Disposition": 'attachment; filename="structured_export.zip"',
            "X-EconAI-Records": str(len(records)),
        })

    payload = json.dumps([r["data"] for r in records], ensure_ascii=False, indent=2)
    return StreamingResponse(_io.BytesIO(payload.encode("utf-8")),
        media_type="application/json", headers={
            "Content-Disposition": 'attachment; filename="structured_export.json"',
            "X-EconAI-Records": str(len(records)),
        })


# ---------------------------------------------------------------------------
# Root — redirect to dashboard
# ---------------------------------------------------------------------------

@app.get("/")
def root():
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/static/dashboard.html")
