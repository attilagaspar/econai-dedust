"""Excel export: lattice layout, non-lattice interleaving, and the column
filter exemption for free annotations.

Guards two real past bugs:
- free annotations (title/footer) on lattice pages were dropped entirely;
- once interleaved at column 0, they were then dropped by any Columns
  filter that excluded column 1 (e.g. "2").
"""
import io

import openpyxl


def _export(client, page_folder, **params):
    p = {"folder": str(page_folder), "scope": "page", "stem": "p1",
         "layer": "human"}
    p.update(params)
    r = client.get("/api/export/excel", params=p)
    assert r.status_code == 200, r.text
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    return wb


def _all_values(ws):
    return {str(c.value) for row in ws.iter_rows() for c in row
            if c.value not in (None, "")}


def test_lattice_and_free_annotations_both_exported(client, page_folder):
    wb = _export(client, page_folder)
    vals = _all_values(wb.worksheets[0])
    for expected in ("A1", "B1", "A2", "B2", "TITLE ROW", "FOOTER ROW"):
        assert expected in vals, f"missing {expected!r} in export"


def test_free_annotations_ordered_around_lattice(client, page_folder):
    ws = _export(client, page_folder).worksheets[0]
    order = [str(c.value) for row in ws.iter_rows() for c in row
             if str(c.value) in ("TITLE ROW", "A1", "A2", "FOOTER ROW")]
    assert order == ["TITLE ROW", "A1", "A2", "FOOTER ROW"]


def test_col_filter_keeps_free_annotations(client, page_folder):
    """Columns filter '2' → only lattice column 2 survives, but the free
    title/footer rows must still be exported."""
    vals = _all_values(_export(client, page_folder, col_filter="2").worksheets[0])
    assert "B1" in vals and "B2" in vals          # column 2 kept
    assert "A1" not in vals and "A2" not in vals  # column 1 filtered out
    assert "TITLE ROW" in vals and "FOOTER ROW" in vals   # free cells exempt


def test_col_filter_open_range(client, page_folder):
    vals = _all_values(_export(client, page_folder, col_filter="2-").worksheets[0])
    assert "B1" in vals and "A1" not in vals
    assert "TITLE ROW" in vals


def test_types_filter(client, page_folder):
    vals = _all_values(_export(client, page_folder, types="cell").worksheets[0])
    assert "A1" in vals
    assert "TITLE ROW" not in vals


# ── Authority columns in the layout sheet ────────────────────────────────────
# For every source column carrying authority resolutions, two columns are
# inserted right next to it: resolved name + stable ID (per internal row when
# a row structure exists, whole-cell otherwise). Opt out with auth_cols=false.

import json

import pytest


@pytest.fixture()
def auth_folder(tmp_path):
    tmp_path = tmp_path / "proj" / "annotations"
    tmp_path.mkdir(parents=True)

    def rows(specs):
        out = []
        for i, (human, auth) in enumerate(specs):
            r = {"n": i + 1, "y0": 100.0 + 30 * i, "y1": 130.0 + 30 * i,
                 "ocr": "", "llm": "", "human": human}
            if auth:
                r["authority"] = {"id": auth[1], "name": auth[0],
                                  "type": "settlement", "score": 100,
                                  "source": "auto"}
            out.append(r)
        return out

    shapes = [
        {"label": "cell", "points": [[100, 100], [300, 190]],
         "shape_type": "rectangle", "flags": {},
         "super_row": 1, "super_column": 1, "table": 0,
         "human_output": {"human_corrected_text": "Tóthfalu\nx\nKérsziget"},
         "row_struct": {"version": 1, "origin": "test", "rows": rows([
             ("Tóthfalu", ("Tótfalu", "M001")),
             ("x", None),
             ("Kérsziget", ("Kérsziget", "M002"))])}},
        {"label": "cell", "points": [[300, 100], [500, 190]],
         "shape_type": "rectangle", "flags": {},
         "super_row": 1, "super_column": 2, "table": 0,
         "human_output": {"human_corrected_text": "12\n34\n56"}},
        {"label": "cell", "points": [[100, 200], [300, 290]],
         "shape_type": "rectangle", "flags": {},
         "super_row": 2, "super_column": 1, "table": 0,
         "human_output": {"human_corrected_text": "Egyfalu"},
         "authority": {"id": "M003", "name": "Egyfalu", "type": "settlement",
                       "score": 95, "source": "human"}},
        {"label": "cell", "points": [[300, 200], [500, 290]],
         "shape_type": "rectangle", "flags": {},
         "super_row": 2, "super_column": 2, "table": 0,
         "human_output": {"human_corrected_text": "78"}},
    ]
    doc = {"version": "5.0.1", "flags": {}, "shapes": shapes,
           "imagePath": "p1.jpg", "imageHeight": 400, "imageWidth": 600}
    (tmp_path / "p1.json").write_text(json.dumps(doc, ensure_ascii=False),
                                      encoding="utf-8")
    from PIL import Image
    Image.new("RGB", (600, 400), "white").save(tmp_path / "p1.jpg")
    return tmp_path


def _grid(ws):
    return [[c.value for c in row] for row in ws.iter_rows()]


def test_auth_cols_inserted_adjacent(client, auth_folder):
    ws = _export(client, auth_folder).worksheets[0]
    g = _grid(ws)
    # data col 1 = excel col 2 (after Row meta); name col 3; id col 4;
    # data col 2 shifted to excel col 5
    r = 1  # first data row (row 0 is the header)
    assert [g[r][1], g[r][2], g[r][3], g[r][4]] == ["Tóthfalu", "Tótfalu", "M001", "12"]
    # unresolved row: empty name/id (openpyxl reads written "" back as None)
    assert [g[r+1][1], g[r+1][2], g[r+1][3]] == ["x", None, None]
    assert [g[r+2][1], g[r+2][2], g[r+2][3]] == ["Kérsziget", "Kérsziget", "M002"]
    # whole-cell authority: on the first (only) line of the lattice row 2 cell
    assert [g[r+3][1], g[r+3][2], g[r+3][3], g[r+3][4]] == ["Egyfalu", "Egyfalu", "M003", "78"]


def test_auth_cols_headers(client, auth_folder):
    ws = _export(client, auth_folder, col_headers="place,value").worksheets[0]
    g = _grid(ws)
    assert g[0][1:5] == ["place", "place → name", "place → id", "value"]


def test_auth_cols_opt_out(client, auth_folder):
    ws = _export(client, auth_folder, auth_cols="false").worksheets[0]
    g = _grid(ws)
    assert [g[1][1], g[1][2]] == ["Tóthfalu", "12"]      # no inserted columns
    vals = _all_values(ws)
    assert "M001" not in vals and "M003" not in vals


def test_auth_cols_respect_col_filter(client, auth_folder):
    # keep only column 2 (no authority there) → no auth columns at all
    ws = _export(client, auth_folder, col_filter="2").worksheets[0]
    vals = _all_values(ws)
    assert "12" in vals and "M001" not in vals and "Tótfalu" not in vals


# ── Pattern groups: horizontal slots aligned across cycles ───────────────────

@pytest.fixture()
def pattern_folder(tmp_path):
    proj = tmp_path / "proj" / "annotations"
    proj.mkdir(parents=True)

    def page(stem, n_cols):
        shapes = [{"label": "cell", "points": [[c * 100, 50], [c * 100 + 90, 100]],
                   "shape_type": "rectangle", "flags": {},
                   "super_row": 1, "super_column": c, "table": 0,
                   "human_output": {"human_corrected_text": f"{stem}c{c}"}}
                  for c in range(1, n_cols + 1)]
        (proj / f"{stem}.json").write_text(json.dumps(
            {"shapes": shapes, "imagePath": f"{stem}.jpg",
             "imageWidth": 800, "imageHeight": 200}), encoding="utf-8")
        from PIL import Image
        Image.new("RGB", (800, 200), "white").save(proj / f"{stem}.jpg")

    page("p1", 1)   # cycle 1 left — NARROW
    page("p2", 2)   # cycle 1 right
    page("p3", 3)   # cycle 2 left — wider
    page("p4", 2)   # cycle 2 right
    return proj


def test_pattern_slots_align_across_cycles(client, pattern_folder):
    """A narrow left page must not shift its cycle's right page west: the
    right slot starts at the same column in every cycle (widest-left rule)."""
    p = {"folder": str(pattern_folder), "scope": "document",
         "pattern": "1,1", "layer": "human"}
    r = client.get("/api/export/excel", params=p)
    assert r.status_code == 200, r.text
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.worksheets[0]
    col_of = {}
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("p") and "c" in c.value:
                col_of.setdefault(c.value, c.column)
            if c.value in ("p1", "p2", "p3", "p4"):
                col_of.setdefault("banner_" + c.value, c.column)
    assert col_of["banner_p1"] == col_of["banner_p3"] == 1
    # right pages start at the same column, sized by the WIDEST left page (p3)
    assert col_of["banner_p2"] == col_of["banner_p4"]
    assert col_of["p2c1"] == col_of["p4c1"]
    assert col_of["banner_p2"] > col_of["banner_p1"] + 3   # room for p3's 3 cols
